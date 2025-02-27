import streamlit as st
import pandas as pd
from openpyxl import load_workbook

# Expand the working space to full width
st.set_page_config(layout="wide")
st.title("Excel File Uploader with Condensed Visualization, Validation, and Global Data Extraction")

# Upload the Excel file
uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

# Define required columns for each sheet
required_columns = {
    "Item Attributes": ["Bid ID", "Incumbent", "Capacity Group"],
    "Supplier Bid Attributes": ["Supplier Name", "Bid ID"],
    "Price": ["Supplier Name", "Bid ID", "Price"],
    "Demand": ["Bid ID", "Demand"],
    "Rebate Tiers": ["Supplier Name", "Min", "Max", "Percentage", "Scope Attribute", "Scope Value"],
    "Discount Tiers": ["Supplier Name", "Min", "Max", "Percentage", "Scope Attribute", "Scope Value"],
    "Baseline Price": ["Bid ID", "Baseline Price"],
    "Per Item Capacity": ["Supplier Name", "Bid ID", "Capacity"],
    "Global Capacity": ["Supplier Name", "Capacity Group", "Capacity"]
}

# Dictionary to hold DataFrames for global extraction
sheet_dfs = {}

if uploaded_file is not None:
    # Load the workbook
    workbook = load_workbook(filename=uploaded_file, data_only=True)
    sheet_names = workbook.sheetnames
    st.write("Worksheets found:", sheet_names)
    
    # Use three columns per row for side-by-side display
    num_cols = 3
    for i in range(0, len(sheet_names), num_cols):
        cols = st.columns(num_cols)
        for j, sheet_name in enumerate(sheet_names[i:i+num_cols]):
            with cols[j]:
                with st.expander(f" {sheet_name}", expanded=False):
                    sheet = workbook[sheet_name]
                    data = []
                    for row in sheet.iter_rows(values_only=True):
                        data.append(list(row))
                    
                    if data:
                        # Use first row as header if all cells are strings
                        if all(isinstance(x, str) for x in data[0]):
                            df = pd.DataFrame(data[1:], columns=data[0])
                        else:
                            df = pd.DataFrame(data)
                        
                        # Clean header names
                        df.columns = [str(col).strip() for col in df.columns]
                        
                        # Save DataFrame for global extraction
                        sheet_dfs[sheet_name] = df.copy()
                        
                        st.dataframe(df)
                        
                        # Validate required columns if defined for this sheet
                        if sheet_name in required_columns:
                            missing_cols = [col for col in required_columns[sheet_name] if col not in df.columns]
                            if missing_cols:
                                st.error(f"Missing required columns: {', '.join(missing_cols)}")
                            
                            # Additional check for Item Attributes: recommend one of 'BusinessUnit' or 'Facility'
                            if sheet_name == "Item Attributes":
                                if not any(col in df.columns for col in ["BusinessUnit", "Facility"]):
                                    st.warning("Missing both 'BusinessUnit' and 'Facility'. One of these is recommended.")

# Global Data Extraction in a thinner sidebar panel
st.sidebar.markdown("## Global Data Extraction")
if sheet_dfs:
    selected_sheet = st.sidebar.selectbox("Select a sheet", options=list(sheet_dfs.keys()))
    df_selected = sheet_dfs[selected_sheet]
    selected_column = st.sidebar.selectbox("Select column", options=df_selected.columns.tolist(), key="global_col")
    
    # Get unique values for filtering (with an option to show all data)
    unique_values = df_selected[selected_column].dropna().unique().tolist()
    unique_values = ["<Show All>"] + unique_values
    selected_value = st.sidebar.selectbox(f"Select value from '{selected_column}'", options=unique_values, key="global_val")
    
    if selected_value == "<Show All>":
        result_df = df_selected
    else:
        result_df = df_selected[df_selected[selected_column] == selected_value]
    
    st.sidebar.write(f"Extracted Data from '{selected_sheet}' where {selected_column} = {selected_value}:")
    st.sidebar.dataframe(result_df)
