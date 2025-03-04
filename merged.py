import os
import uuid
import pandas as pd
import pulp
import streamlit as st
from openpyxl import load_workbook

#############################################
# REQUIRED COLUMNS & HELPER FUNCTIONS FOR EXCEL INTAKE
#############################################

REQUIRED_COLUMNS = {
    "Item Attributes": ["Bid ID", "Incumbent", "Capacity Group"],
    "Supplier Bid Attributes": ["Supplier Name", "Bid ID"],
    "Price": ["Supplier Name", "Bid ID", "Price"],
    "Demand": ["Bid ID", "Demand"],
    "Rebate Tiers": ["Supplier Name", "Min", "Max", "Percentage", "Scope Attribute", "Scope Value"],
    "Discount Tiers": ["Supplier Name", "Min", "Max", "Percentage", "Scope Attribute", "Scope Value"],
    "Baseline Price": ["Bid ID", "Baseline Price"],
    "Per Item Capacity": ["Supplier Name", "Bid ID", "Capacity"],
    "Global Capacity": ["Supplier Name", "Capacity Group", "Capacity"]
}

def load_excel_sheets(uploaded_file):
    sheet_dfs = {}
    workbook = load_workbook(filename=uploaded_file, data_only=True)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        data = [list(row) for row in sheet.iter_rows(values_only=True)]
        if data:
            if all(isinstance(x, str) for x in data[0]):
                df = pd.DataFrame(data[1:], columns=[str(x).strip() for x in data[0]])
            else:
                df = pd.DataFrame(data)
            sheet_dfs[sheet_name] = df.copy()
    return sheet_dfs

def validate_sheet(df, sheet_name):
    required = REQUIRED_COLUMNS.get(sheet_name, [])
    missing = [col for col in required if col not in df.columns]
    return missing

def df_to_dict_item_attributes(df):
    df.columns = [col.strip() for col in df.columns]
    bid_id_col = None
    for col in df.columns:
        if col.lower() == "bid id":
            bid_id_col = col
            break
    if bid_id_col is None:
        raise Exception("The 'Bid ID' column is missing in the Item Attributes sheet.")
    df = df.dropna(subset=[bid_id_col])
    df[bid_id_col] = df[bid_id_col].astype(str).str.strip()
    return {bid: row.to_dict() for bid, row in df.set_index(bid_id_col).iterrows()}

def df_to_dict_supplier_bid_attributes(df):
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=["Supplier Name", "Bid ID"])
    df["Supplier Name"] = df["Supplier Name"].astype(str).str.strip()
    df["Bid ID"] = df["Bid ID"].astype(str).str.strip()
    result = {}
    for _, row in df.iterrows():
        key = (row["Supplier Name"], row["Bid ID"])
        result[key] = row.drop(labels=["Supplier Name", "Bid ID"]).to_dict()
    return result

def df_to_dict_price(df):
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=["Supplier Name", "Bid ID", "Price"])
    df["Supplier Name"] = df["Supplier Name"].astype(str).str.strip()
    df["Bid ID"] = df["Bid ID"].astype(str).str.strip()
    result = {}
    for _, row in df.iterrows():
        result[(row["Supplier Name"], row["Bid ID"])] = float(row["Price"])
    return result

def df_to_dict_demand(df):
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=["Bid ID", "Demand"])
    df["Bid ID"] = df["Bid ID"].astype(str).str.strip()
    return {bid: float(demand) for bid, demand in df.set_index("Bid ID")["Demand"].to_dict().items()}

def df_to_dict_baseline_price(df):
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=["Bid ID", "Baseline Price"])
    df["Bid ID"] = df["Bid ID"].astype(str).str.strip()
    return {bid: float(price) for bid, price in df.set_index("Bid ID")["Baseline Price"].to_dict().items()}

def df_to_dict_per_item_capacity(df):
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=["Supplier Name", "Bid ID", "Capacity"])
    df["Supplier Name"] = df["Supplier Name"].astype(str).str.strip()
    df["Bid ID"] = df["Bid ID"].astype(str).str.strip()
    result = {}
    for _, row in df.iterrows():
        result[(row["Supplier Name"], row["Bid ID"])] = float(row["Capacity"])
    return result

def df_to_dict_global_capacity(df):
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=["Supplier Name", "Capacity Group", "Capacity"])
    df["Supplier Name"] = df["Supplier Name"].astype(str).str.strip()
    df["Capacity Group"] = df["Capacity Group"].astype(str).str.strip()
    result = {}
    for _, row in df.iterrows():
        result[(row["Supplier Name"], row["Capacity Group"])] = float(row["Capacity"])
    return result

def df_to_dict_tiers(df):
    df.columns = [col.strip() for col in df.columns]
    df = df.dropna(subset=["Supplier Name", "Min", "Max", "Percentage"])
    df["Supplier Name"] = df["Supplier Name"].astype(str).str.strip()
    tiers = {}
    for supplier in df["Supplier Name"].unique():
        sub_df = df[df["Supplier Name"] == supplier]
        tier_list = []
        for _, row in sub_df.iterrows():
            tier_list.append((float(row["Min"]), float(row["Max"]), float(row["Percentage"]),
                              row.get("Scope Attribute", None), row.get("Scope Value", None)))
        tiers[supplier] = tier_list
    return tiers

#############################################
# COMPUTE HELPER FUNCTIONS (VOLUME & SPEND)
#############################################

def compute_U_volume(per_item_cap, suppliers):
    total = {}
    for s in suppliers:
        tot = sum(per_item_cap.get((s, bid), 0) for (sup, bid) in per_item_cap.keys() if sup == s)
        total[s] = tot
    return total

def compute_U_spend(per_item_cap, price_data, suppliers):
    total = {}
    for s in suppliers:
        tot = sum(price_data.get((s, bid), 0) * per_item_cap.get((s, bid), 0)
                  for (sup, bid) in per_item_cap.keys() if sup == s)
        total[s] = tot
    return total

#############################################
# OPTIMIZATION MODEL FUNCTION
#############################################

def run_optimization(use_global, capacity_data, demand_data, item_attr_data, price_data,
                     rebate_tiers, discount_tiers, baseline_price_data, per_item_capacity,
                     supplier_bid_attributes, suppliers, rules=[]):
    debug = True
    items_dynamic = list(demand_data.keys())
    T = {}
    for bid in items_dynamic:
        incumbent = item_attr_data[bid].get("Incumbent")
        for s in suppliers:
            if s != incumbent:
                T[(bid, s)] = pulp.LpVariable(f"T_{bid}_{s}", cat='Binary')
    U_volume = compute_U_volume(per_item_capacity, suppliers)
    U_spend = compute_U_spend(per_item_capacity, price_data, suppliers)
    
    lp_problem = pulp.LpProblem("Sourcing_with_MultiTier_Rebates_Discounts", pulp.LpMinimize)
    
    x = {(s, bid): pulp.LpVariable(f"x_{s}_{bid}", lowBound=0, cat='Continuous')
         for s in suppliers for bid in items_dynamic}
    S0 = {s: pulp.LpVariable(f"S0_{s}", lowBound=0, cat='Continuous') for s in suppliers}
    S = {s: pulp.LpVariable(f"S_{s}", lowBound=0, cat='Continuous') for s in suppliers}
    V = {s: pulp.LpVariable(f"V_{s}", lowBound=0, cat='Continuous') for s in suppliers}
    
    for bid in items_dynamic:
        for s in suppliers:
            if (bid, s) in T:
                lp_problem += x[(s, bid)] <= demand_data[bid] * T[(bid, s)], f"Transition_{bid}_{s}"
    
    z_discount = {}
    for s in suppliers:
        tiers = discount_tiers.get(s, [])
        z_discount[s] = {k: pulp.LpVariable(f"z_discount_{s}_{k}", cat='Binary') for k in range(len(tiers))}
        lp_problem += pulp.lpSum(z_discount[s][k] for k in range(len(tiers))) == 1, f"DiscountTierSelect_{s}"
    
    y_rebate = {}
    for s in suppliers:
        tiers = rebate_tiers.get(s, [])
        y_rebate[s] = {k: pulp.LpVariable(f"y_rebate_{s}_{k}", cat='Binary') for k in range(len(tiers))}
        lp_problem += pulp.lpSum(y_rebate[s][k] for k in range(len(tiers))) == 1, f"RebateTierSelect_{s}"
    
    d = {s: pulp.LpVariable(f"d_{s}", lowBound=0, cat='Continuous') for s in suppliers}
    rebate_var = {s: pulp.LpVariable(f"rebate_{s}", lowBound=0, cat='Continuous') for s in suppliers}
    
    lp_problem += pulp.lpSum(S[s] - rebate_var[s] for s in suppliers), "Total_Effective_Cost"
    
    for bid in items_dynamic:
        lp_problem += pulp.lpSum(x[(s, bid)] for s in suppliers) == demand_data[bid], f"Demand_{bid}"
    
    if use_global:
        supplier_capacity_groups = {}
        for s in suppliers:
            supplier_capacity_groups[s] = {}
        for bid in items_dynamic:
            group = item_attr_data[bid].get("Capacity Group")
            if group is not None:
                for s in suppliers:
                    supplier_capacity_groups[s].setdefault(group, []).append(bid)
        for s in suppliers:
            for group, bid_list in supplier_capacity_groups[s].items():
                cap = capacity_data.get((s, group), 1e9)
                lp_problem += pulp.lpSum(x[(s, bid)] for bid in bid_list) <= cap, f"GlobalCapacity_{s}_{group}"
    else:
        for s in suppliers:
            for bid in items_dynamic:
                cap = capacity_data.get((s, bid), 1e9)
                lp_problem += x[(s, bid)] <= cap, f"PerItemCapacity_{s}_{bid}"
    
    for s in suppliers:
        lp_problem += S0[s] == pulp.lpSum(price_data.get((s, bid), 0) * x[(s, bid)] for bid in items_dynamic), f"BaseSpend_{s}"
        lp_problem += V[s] == pulp.lpSum(x[(s, bid)] for bid in items_dynamic), f"Volume_{s}"
    
    M = 1e9
    small_value = 1e-3
    
    for s in suppliers:
        tiers = discount_tiers.get(s, [])
        M_discount = U_spend[s]
        for k, tier in enumerate(tiers):
            Dmin, Dmax, Dperc, scope_attr, scope_value = tier
            if scope_attr is None or scope_value is None:
                vol_expr = pulp.lpSum(x[(s, bid)] for bid in items_dynamic)
            else:
                vol_expr = pulp.lpSum(x[(s, bid)] for bid in items_dynamic if item_attr_data[bid].get(scope_attr) == scope_value)
            lp_problem += vol_expr >= Dmin * z_discount[s][k], f"DiscountTierMin_{s}_{k}"
            if Dmax < float('inf'):
                lp_problem += vol_expr <= Dmax + M_discount * (1 - z_discount[s][k]), f"DiscountTierMax_{s}_{k}"
            lp_problem += d[s] >= Dperc * S0[s] - M_discount * (1 - z_discount[s][k]), f"DiscountTierLower_{s}_{k}"
            lp_problem += d[s] <= Dperc * S0[s] + M_discount * (1 - z_discount[s][k]), f"DiscountTierUpper_{s}_{k}"
    
    for s in suppliers:
        lp_problem += S[s] == S0[s] - d[s], f"EffectiveSpend_{s}"
    
    for s in suppliers:
        tiers = rebate_tiers.get(s, [])
        M_rebate = U_spend[s]
        for k, tier in enumerate(tiers):
            Rmin, Rmax, Rperc, scope_attr, scope_value = tier
            if scope_attr is None or scope_value is None:
                vol_expr = pulp.lpSum(x[(s, bid)] for bid in items_dynamic)
            else:
                vol_expr = pulp.lpSum(x[(s, bid)] for bid in items_dynamic if item_attr_data[bid].get(scope_attr) == scope_value)
            lp_problem += vol_expr >= Rmin * y_rebate[s][k], f"RebateTierMin_{s}_{k}"
            if Rmax < float('inf'):
                lp_problem += vol_expr <= Rmax + M_rebate * (1 - y_rebate[s][k]), f"RebateTierMax_{s}_{k}"
            lp_problem += rebate_var[s] >= Rperc * S[s] - M_rebate * (1 - y_rebate[s][k]), f"RebateTierLower_{s}_{k}"
            lp_problem += rebate_var[s] <= Rperc * S[s] + M_rebate * (1 - y_rebate[s][k]), f"RebateTierUpper_{s}_{k}"
    
    # CUSTOM RULES PROCESSING
    for r_idx, rule in enumerate(rules):
        if rule["rule_type"] == "# of suppliers":
            if rule["grouping"] == "Bid ID" and rule["operator"] == "Exactly" and rule["rule_input"] == "1":
                if rule["grouping_scope"] == "Apply to all items individually":
                    bids = sorted(list(item_attr_data.keys()))
                else:
                    bids = [rule["grouping_scope"]]
                for bid in bids:
                    w = {}
                    for s in suppliers:
                        w[(s, bid)] = pulp.LpVariable(f"w_{r_idx}_{bid}_{s}", cat='Binary')
                        lp_problem += x[(s, bid)] <= M * w[(s, bid)], f"RuleSupplierIndicator_{r_idx}_{bid}_{s}"
                        lp_problem += x[(s, bid)] >= small_value * w[(s, bid)], f"RuleSupplierIndicatorLB_{r_idx}_{bid}_{s}"
                    lp_problem += pulp.lpSum(w[(s, bid)] for s in suppliers) == 1, f"RuleSingleSupplier_{r_idx}_{bid}"
                    if debug:
                        print(f"DEBUG: Enforcing exactly one supplier for Bid {bid} via rule {r_idx}")
            else:
                if rule["grouping"] == "All" or rule["grouping_scope"] == "All":
                    items_group = items_dynamic
                elif rule["grouping_scope"] == "Apply to all items individually":
                    if rule["grouping"] == "Bid ID":
                        items_group = sorted(list(item_attr_data.keys()))
                    else:
                        items_group = sorted({str(item_attr_data[bid].get(rule["grouping"], "")).strip() 
                                               for bid in item_attr_data if str(item_attr_data[bid].get(rule["grouping"], "")).strip() != ""})
                else:
                    items_group = [bid for bid in items_dynamic if str(item_attr_data[bid].get(rule["grouping"], "")).strip() == str(rule["grouping_scope"]).strip()]
                try:
                    supplier_target = int(rule["rule_input"])
                except:
                    continue
                operator = rule["operator"]
                w = {}
                for s in suppliers:
                    w[s] = pulp.LpVariable(f"w_{r_idx}_{s}", cat='Binary')
                    lp_problem += pulp.lpSum(x[(s, bid)] for bid in items_group) <= M * w[s], f"SupplierIndicator_{r_idx}_{s}"
                    lp_problem += pulp.lpSum(x[(s, bid)] for bid in items_group) >= small_value * w[s], f"SupplierIndicatorLB_{r_idx}_{s}"
                total_suppliers = pulp.lpSum(w[s] for s in suppliers)
                if operator == "At least":
                    lp_problem += total_suppliers >= supplier_target, f"Rule_{r_idx}"
                elif operator == "At most":
                    lp_problem += total_suppliers <= supplier_target, f"Rule_{r_idx}"
                elif operator == "Exactly":
                    lp_problem += total_suppliers == supplier_target, f"Rule_{r_idx}"
        elif rule["rule_type"] == "% of Volume Awarded":
            if rule["supplier_scope"] == "New Suppliers" and rule["grouping"] == "All":
                try:
                    percentage = float(rule["rule_input"]) / 100.0
                except:
                    continue
                total_new_suppliers_vol = pulp.lpSum(
                    pulp.lpSum(x[(s, bid)] for s in suppliers if s != item_attr_data[bid].get("Incumbent"))
                    for bid in items_dynamic
                )
                lp_problem += total_new_suppliers_vol <= percentage * sum(demand_data[bid] for bid in items_dynamic), f"Rule_{r_idx}_AggregateNewSuppliers"
                if debug:
                    print(f"DEBUG: Enforcing aggregate new suppliers volume <= {percentage * sum(demand_data[bid] for bid in items_dynamic):.2f}")
            else:
                if rule["grouping"] == "All" or rule["grouping_scope"] == "All":
                    items_group = items_dynamic
                elif rule["grouping_scope"] == "Apply to all items individually":
                    if rule["grouping"] == "Bid ID":
                        items_group = sorted(list(item_attr_data.keys()))
                    else:
                        items_group = sorted({str(item_attr_data[bid].get(rule["grouping"], "")).strip() 
                                               for bid in item_attr_data if str(item_attr_data[bid].get(rule["grouping"], "")).strip() != ""})
                else:
                    items_group = [bid for bid in items_dynamic if str(item_attr_data[bid].get(rule["grouping"], "")).strip() == str(rule["grouping_scope"]).strip()]
                try:
                    percentage = float(rule["rule_input"]) / 100.0
                except:
                    continue
                operator = rule["operator"]
                lhs = pulp.lpSum(x[(rule["supplier_scope"], bid)] for bid in items_group) if rule["supplier_scope"] else 0
                total_vol = pulp.lpSum(x[(s, bid)] for s in suppliers for bid in items_group)
                if operator == "At least":
                    lp_problem += lhs >= percentage * total_vol, f"Rule_{r_idx}"
                elif operator == "At most":
                    lp_problem += lhs <= percentage * total_vol, f"Rule_{r_idx}"
                elif operator == "Exactly":
                    lp_problem += lhs == percentage * total_vol, f"Rule_{r_idx}"
        elif rule["rule_type"] == "# of transitions":
            if rule["grouping"] == "All" or rule["grouping_scope"] == "All":
                items_group = items_dynamic
            elif rule["grouping_scope"] == "Apply to all items individually":
                if rule["grouping"] == "Bid ID":
                    items_group = sorted(list(item_attr_data.keys()))
                else:
                    items_group = sorted({str(item_attr_data[bid].get(rule["grouping"], "")).strip() 
                                           for bid in item_attr_data if str(item_attr_data[bid].get(rule["grouping"], "")).strip() != ""})
            else:
                items_group = [bid for bid in items_dynamic if str(item_attr_data[bid].get(rule["grouping"], "")).strip() == str(rule["grouping_scope"]).strip()]
            try:
                transitions_target = int(rule["rule_input"])
            except:
                continue
            operator = rule["operator"]
            total_transitions = pulp.lpSum(T[(bid, s)] for (bid, s) in T if bid in items_group)
            if operator == "At least":
                lp_problem += total_transitions >= transitions_target, f"Rule_{r_idx}"
            elif operator == "At most":
                lp_problem += total_transitions <= transitions_target, f"Rule_{r_idx}"
            elif operator == "Exactly":
                lp_problem += total_transitions == transitions_target, f"Rule_{r_idx}"
            if debug:
                print(f"DEBUG: Enforcing total transitions {operator} {transitions_target} over items {items_group}")
        elif rule["rule_type"] == "Bid Exclusions":
            if rule["grouping"] == "All" or rule["grouping_scope"] in ["All", "Apply to all items individually"]:
                items_group = items_dynamic
            else:
                items_group = [bid for bid in items_dynamic if str(item_attr_data[bid].get(rule["grouping"], "")).strip() == str(rule["grouping_scope"]).strip()]
            bid_group = rule.get("bid_grouping", None)
            if bid_group is None:
                continue
            is_numeric = True
            for bid in items_group:
                for s in suppliers:
                    val = supplier_bid_attributes.get((s, bid), {}).get(bid_group, None)
                    if val is not None:
                        try:
                            float(val)
                        except:
                            is_numeric = False
                        break
                break
            for bid in items_group:
                for s in suppliers:
                    bid_val = supplier_bid_attributes.get((s, bid), {}).get(bid_group, None)
                    if bid_val is None:
                        continue
                    exclude = False
                    if is_numeric:
                        try:
                            bid_val_num = float(bid_val)
                            threshold = float(rule["rule_input"])
                        except:
                            continue
                        op = rule["operator"]
                        if op == "At most" and bid_val_num > threshold:
                            exclude = True
                        elif op == "At least" and bid_val_num < threshold:
                            exclude = True
                        elif op == "Exactly" and bid_val_num != threshold:
                            exclude = True
                    else:
                        if bid_val.strip() == rule.get("bid_exclusion_value", "").strip():
                            exclude = True
                    if exclude:
                        lp_problem += x[(s, bid)] == 0, f"BidExclusion_{r_idx}_{bid}_{s}"
                        if debug:
                            print(f"DEBUG: Excluding Bid {bid} for supplier {s} due to Bid Exclusions rule {r_idx} on {bid_group} with value {bid_val}")
    
    if debug:
        constraint_names = list(lp_problem.constraints.keys())
        duplicates = set([n for n in constraint_names if constraint_names.count(n) > 1])
        if duplicates:
            print("DEBUG: Duplicate constraint names found:", duplicates)
        else:
            print("DEBUG: No duplicate constraint names.")
        print("DEBUG: Total constraints added:", len(constraint_names))
    
    lp_problem.solve()
    model_status = pulp.LpStatus[lp_problem.status]
    
    feasibility_notes = ""
    if model_status == "Infeasible":
        feasibility_notes += "Model is infeasible. Likely causes include:\n"
        feasibility_notes += " - Insufficient supplier capacity relative to demand.\n"
        feasibility_notes += " - Custom rule constraints conflicting with overall volume/demand.\n"
        for bid in items_dynamic:
            if use_global:
                group = item_attr_data[bid].get("Capacity Group", "Unknown")
                total_cap = sum(capacity_data.get((s, group), 0) for s in suppliers)
                feasibility_notes += f"Bid {bid}: demand = {demand_data[bid]}, capacity for group {group} = {total_cap}\n"
            else:
                total_cap = sum(capacity_data.get((s, bid), 0) for s in suppliers)
                feasibility_notes += f"Bid {bid}: demand = {demand_data[bid]}, capacity = {total_cap}\n"
        feasibility_notes += "Please review supplier capacities, demand, and custom rule constraints."
    else:
        feasibility_notes = "Model is optimal."
    
    excel_rows = []
    letter_list = list("ABCDEFGHIJKLMNOPQRSTUVWXYZ")
    for bid in items_dynamic:
        awarded_list = []
        for s in suppliers:
            award_val = pulp.value(x[(s, bid)])
            if award_val is None:
                award_val = 0
            if award_val > 0:
                awarded_list.append((s, award_val))
        if not awarded_list:
            awarded_list = [("None", 0)]
        awarded_list.sort(key=lambda tup: (-tup[1], tup[0]))
        for i, (s, award_val) in enumerate(awarded_list):
            bid_split = letter_list[i] if i < len(letter_list) else f"Split{i+1}"
            orig_price = price_data.get((s, bid), 0)
            active_discount = 0
            for k, tier in enumerate(discount_tiers.get(s, [])):
                if pulp.value(z_discount[s][k]) is not None and pulp.value(z_discount[s][k]) >= 0.5:
                    active_discount = tier[2]
                    break
            discounted_price = orig_price * (1 - active_discount)
            awarded_spend = discounted_price * award_val
            base_price = baseline_price_data.get(bid, 0)
            baseline_spend = base_price * award_val
            baseline_savings = baseline_spend - awarded_spend
            if use_global:
                group = item_attr_data[bid].get("Capacity Group", "Unknown")
                awarded_capacity = capacity_data.get((s, group), 0)
            else:
                awarded_capacity = capacity_data.get((s, bid), 0)
            active_rebate = 0
            for k, tier in enumerate(rebate_tiers.get(s, [])):
                if pulp.value(y_rebate[s][k]) is not None and pulp.value(y_rebate[s][k]) >= 0.5:
                    active_rebate = tier[2]
                    break
            rebate_savings = awarded_spend * active_rebate
            facility_val = item_attr_data[bid].get("Facility", "")
            row = {
                "Bid ID": bid,
                "Capacity Group": item_attr_data[bid].get("Capacity Group", "") if use_global else "",
                "Bid ID Split": bid_split,
                "Facility": facility_val,
                "Incumbent": item_attr_data[bid].get("Incumbent", ""),
                "Baseline Price": base_price,
                "Baseline Spend": baseline_spend,
                "Awarded Supplier": s,
                "Original Awarded Supplier Price": orig_price,
                "Percentage Volume Discount": f"{active_discount*100:.0f}%",
                "Discounted Awarded Supplier Price": discounted_price,
                "Awarded Supplier Spend": awarded_spend,
                "Awarded Volume": award_val,
                "Awarded Supplier Capacity": awarded_capacity,
                "Baseline Savings": baseline_savings,
                "Rebate %": f"{active_rebate*100:.0f}%",
                "Rebate Savings": rebate_savings
            }
            excel_rows.append(row)
    
    df_results = pd.DataFrame(excel_rows)
    if use_global:
        cols = ["Bid ID", "Capacity Group", "Bid ID Split", "Facility", "Incumbent", "Baseline Price", "Baseline Spend",
                "Awarded Supplier", "Original Awarded Supplier Price", "Percentage Volume Discount",
                "Discounted Awarded Supplier Price", "Awarded Supplier Spend", "Awarded Volume",
                "Awarded Supplier Capacity", "Baseline Savings", "Rebate %", "Rebate Savings"]
    else:
        cols = ["Bid ID", "Bid ID Split", "Facility", "Incumbent", "Baseline Price", "Baseline Spend",
                "Awarded Supplier", "Original Awarded Supplier Price", "Percentage Volume Discount",
                "Discounted Awarded Supplier Price", "Awarded Supplier Spend", "Awarded Volume",
                "Awarded Supplier Capacity", "Baseline Savings", "Rebate %", "Rebate Savings"]
    df_results = df_results[cols]
    
    home_dir = os.path.expanduser("~")
    downloads_folder = os.path.join(home_dir, "Downloads")
    temp_lp_file = os.path.join(downloads_folder, "temp_model.lp")
    lp_problem.writeLP(temp_lp_file)
    with open(temp_lp_file, "r") as f:
        lp_text = f.read()
    df_lp = pd.DataFrame({"LP Model": [lp_text]})
    
    output_file = os.path.join(downloads_folder, "optimization_results.xlsx")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df_results.to_excel(writer, sheet_name="Results", index=False)
        pd.DataFrame({"Feasibility Notes": [feasibility_notes]}).to_excel(writer, sheet_name="Feasibility Notes", index=False)
        df_lp.to_excel(writer, sheet_name="LP Model", index=False)
    
    return output_file, feasibility_notes, model_status

#############################################
# Helper: Generate human-readable rule description
#############################################
def rule_to_text(rule):
    rt = rule["rule_type"]
    op = rule["operator"]
    inp = rule["rule_input"]
    group = rule["grouping"]
    group_scope = rule["grouping_scope"]
    supp_scope = rule.get("supplier_scope", "All Suppliers")
    if rt == "% of Volume Awarded":
        return f"{op} {inp}% of volume is awarded to {supp_scope} in grouping '{group_scope}' ({group})."
    elif rt == "# of Volume Awarded":
        return f"{op} {inp} volume units are awarded to {supp_scope} in grouping '{group_scope}' ({group})."
    elif rt == "# of transitions":
        return f"{op} {inp} transitions are enforced in grouping '{group_scope}' ({group})."
    elif rt == "# of suppliers":
        return f"{op} {inp} unique suppliers are awarded in grouping '{group_scope}' ({group})."
    elif rt == "Supplier Exclusion":
        return f"Exclude supplier {supp_scope} from grouping '{group_scope}' ({group})."
    elif rt == "Bid Exclusions":
        bid_grp = rule.get("bid_grouping", "N/A")
        bid_val = rule.get("bid_exclusion_value", "N/A")
        return f"Exclude bids with {bid_grp} {op} {inp} (or value '{bid_val}')."
    else:
        return str(rule)

#############################################
# STREAMLIT USER INTERFACE
#############################################

st.set_page_config(layout="centered")
st.title("Scenario Builder & Optimization")

uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

if uploaded_file is not None:
    st.markdown("### Uploaded Worksheets")
    sheet_dfs = load_excel_sheets(uploaded_file)
    for sheet_name, df in sheet_dfs.items():
        missing = validate_sheet(df, sheet_name)
        if missing:
            st.error(f"Missing required columns in '{sheet_name}': {', '.join(missing)}")
        # else:
        #     st.success(f"'{sheet_name}' has all required columns.")
    
    # Capacity Input Type (moved above custom rules)
    capacity_mode = st.radio("Select Capacity Input Type:", ("Global Capacity", "Per Item Capacity"))
    use_global = (capacity_mode == "Global Capacity")
    
    # --- Custom Rules Section ---
    st.markdown("### Custom Rules")
    col_rt, col_op = st.columns(2)
    with col_rt:
        rule_type = st.selectbox("Rule Type", options=["% of Volume Awarded", "# of Volume Awarded", "# of transitions", "# of suppliers", "Supplier Exclusion", "Bid Exclusions"])
    with col_op:
        operator = st.selectbox("Operator", options=["At least", "At most", "Exactly"])
    rule_input = st.text_input("Rule Input")
    
    if "Item Attributes" in sheet_dfs:
        temp_item_attr = df_to_dict_item_attributes(sheet_dfs["Item Attributes"])
        sample_keys = list(next(iter(temp_item_attr.values())).keys())
        grouping_options = ["All", "Bid ID"] + sample_keys
    else:
        grouping_options = ["All", "Bid ID"]
    col_group, col_group_scope = st.columns(2)
    with col_group:
        grouping = st.selectbox("Grouping", options=grouping_options)
    with col_group_scope:
        def update_grouping_scope(grouping, item_attr_data):
            if grouping == "Bid ID":
                vals = sorted(list(item_attr_data.keys()))
            else:
                vals = sorted({str(item_attr_data[bid].get(grouping, "")).strip() 
                               for bid in item_attr_data if str(item_attr_data[bid].get(grouping, "")).strip() != ""})
            return ["Apply to all items individually"] + vals
        if grouping != "All":
            grouping_scope = st.selectbox("Grouping Scope", options=update_grouping_scope(grouping, temp_item_attr))
        else:
            grouping_scope = "All"
    
    # Auto-populate Supplier Scope from uploaded data
    suppliers_auto = st.session_state.get("suppliers", [])
    if suppliers_auto:
        supplier_scope = st.selectbox("Supplier Scope", options=["All"] + suppliers_auto)
        if supplier_scope == "All":
            supplier_scope = None
    else:
        supplier_scope = st.text_input("Supplier Scope (or leave blank)", value="")
    
    bid_grouping = None
    bid_exclusion_value = None
    if rule_type == "Bid Exclusions":
        bid_grouping = st.selectbox("Bid Grouping", options=["Milage", "Origin Country"])
        if bid_grouping != "Milage":
            bid_exclusion_value = st.text_input("Bid Exclusion Value", value="")
    
    col_add, col_clear = st.columns(2)
    with col_add:
        if st.button("Add Rule"):
            rule = {
                "rule_type": rule_type,
                "operator": operator,
                "rule_input": rule_input,
                "grouping": grouping,
                "grouping_scope": grouping_scope,
                "supplier_scope": supplier_scope
            }
            if rule_type == "Bid Exclusions":
                rule["bid_grouping"] = bid_grouping
                rule["bid_exclusion_value"] = bid_exclusion_value
            if "rules_list" not in st.session_state:
                st.session_state.rules_list = []
            st.session_state.rules_list.append(rule)
            st.success("Rule added.")
    with col_clear:
        if st.button("Clear Rules"):
            st.session_state.rules_list = []
            st.success("All rules cleared.")
    
    if "rules_list" in st.session_state and st.session_state.rules_list:
        st.markdown("#### Current Rules")
        for i, r in enumerate(st.session_state.rules_list):
            col_rule, col_del = st.columns([0.95, 0.05])
            with col_rule:
                st.write(f"{i+1}. {rule_to_text(r)}")
            with col_del:
                if st.button("X", key=f"delete_rule_{i}"):
                    st.session_state.rules_list.pop(i)
                    # The page will automatically rerun after state update.
    
    # --- End Custom Rules Section ---
    
    required_sheet_names = ["Item Attributes", "Price", "Demand", "Rebate Tiers", "Discount Tiers",
                            "Baseline Price", "Per Item Capacity", "Global Capacity", "Supplier Bid Attributes"]
    if all(sheet in sheet_dfs for sheet in required_sheet_names):
        try:
            item_attr_dict = df_to_dict_item_attributes(sheet_dfs["Item Attributes"])
            price_dict = df_to_dict_price(sheet_dfs["Price"])
            demand_dict = df_to_dict_demand(sheet_dfs["Demand"])
            baseline_price_dict = df_to_dict_baseline_price(sheet_dfs["Baseline Price"])
            per_item_capacity_dict = df_to_dict_per_item_capacity(sheet_dfs["Per Item Capacity"])
            global_capacity_dict = df_to_dict_global_capacity(sheet_dfs["Global Capacity"])
            rebate_tiers_dict = df_to_dict_tiers(sheet_dfs["Rebate Tiers"])
            discount_tiers_dict = df_to_dict_tiers(sheet_dfs["Discount Tiers"])
            supplier_bid_attr_dict = df_to_dict_supplier_bid_attributes(sheet_dfs["Supplier Bid Attributes"])
            
            suppliers = sheet_dfs["Price"]["Supplier Name"].dropna().astype(str).str.strip().unique().tolist()
            st.session_state.suppliers = suppliers
            st.success("Data extraction complete. Ready to run optimization.")
            
            if st.button("Run Optimization"):
                with st.spinner("Running optimization..."):
                    output_file, feasibility_notes, model_status = run_optimization(
                        use_global,
                        global_capacity_dict if use_global else per_item_capacity_dict,
                        demand_dict,
                        item_attr_dict,
                        price_dict,
                        rebate_tiers_dict,
                        discount_tiers_dict,
                        baseline_price_dict,
                        per_item_capacity_dict,
                        supplier_bid_attr_dict,
                        suppliers,
                        rules=st.session_state.rules_list
                    )
                st.success(f"Model Status: {model_status}")
                st.markdown("#### Feasibility Notes")
                st.text(feasibility_notes)
                st.markdown("#### Optimization Results")
                df_results = pd.read_excel(output_file, sheet_name="Results")
                st.dataframe(df_results)
                with open(output_file, "rb") as f:
                    st.download_button("Download Results", f, file_name="optimization_results.xlsx")
        except Exception as e:
            st.error(f"Error preparing data for optimization: {e}")
