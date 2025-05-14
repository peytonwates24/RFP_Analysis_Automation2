import streamlit as st
import pandas as pd
from io import BytesIO
import os
from modules.config import *
from modules.authentication import *
from modules.utils import *
from modules.data_loader import *
from modules.analysis import *
from modules.presentations import *
from modules.projects import *
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.styles import Font
from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from pptx import Presentation
from modules.optimization import *
import logging
import uuid 
from modules.ppt_scenario import *



def apply_custom_css():
    """Apply custom CSS for styling the app."""
    st.markdown(
        f"""
        <style>
        /* Set a subtle background color */
        body {{
            background-color: #f0f2f6;
        }}

        /* Remove the default main menu and footer for a cleaner look */
        #MainMenu {{visibility: hidden;}}
        footer {{visibility: hidden;}}

        /* Style for the header */
        .header {{
            display: flex;
            align-items: center;
            padding: 10px 20px;
            background-color: #ffffff;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
            margin-bottom: 20px;
        }}
        .header img {{
            height: 50px;
            margin-right: 20px;
        }}
        .header .page-title {{
            font-size: 24px;
            font-weight: bold;
        }}
        </style>
        """,
        unsafe_allow_html=True
    )


# Streamlit App
def main():
    # Apply custom CSS
    apply_custom_css()

    # Initialize session state variables
    if 'authenticated' not in st.session_state:
        st.session_state.authenticated = False
    if 'username' not in st.session_state:
        st.session_state.username = ''
    if 'merged_data' not in st.session_state:
        st.session_state.merged_data = None
    if 'original_merged_data' not in st.session_state:
        st.session_state.original_merged_data = None
    if 'column_mapping' not in st.session_state:
        st.session_state.column_mapping = {}
    if 'columns' not in st.session_state:
        st.session_state.columns = []
    if 'baseline_data' not in st.session_state:
        st.session_state.baseline_data = None
    if 'exclusions_bob' not in st.session_state:
        st.session_state.exclusions_bob = []
    if 'exclusions_ais' not in st.session_state:
        st.session_state.exclusions_ais = []
    if 'current_section' not in st.session_state:
        st.session_state.current_section = 'home'
    if 'selected_project' not in st.session_state:
        st.session_state.selected_project = None
    if 'selected_subfolder' not in st.session_state:
        st.session_state.selected_subfolder = None
    if 'customizable_grouping_column' not in st.session_state:
        st.session_state.customizable_grouping_column = None
    if "rules_list" not in st.session_state:
        st.session_state.rules_list = []



    # Header with logo and page title
    st.markdown(
        f"""
        <div class="header">
            <img src="https://scfuturemakers.com/wp-content/uploads/2017/11/Georgia-Pacific_overview_video-854x480-c-default.jpg" alt="Logo">
            <div class="page-title">{st.session_state.current_section.capitalize()}</div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # Right side: User Info and Logout button
    col1, col2 = st.columns([3, 1])

    with col1:
        pass  # Logo and page title are handled above

    with col2:
        if st.session_state.authenticated:
            st.markdown(f"<div style='text-align: right;'>**Welcome, {st.session_state.username}!**</div>", unsafe_allow_html=True)
            logout = st.button("Logout", key='logout_button')
            if logout:
                st.session_state.authenticated = False
                st.session_state.username = ''
                st.success("You have been logged out.")
                logger.info(f"User logged out successfully.")
        else:
            # Display a button to navigate to 'My Projects'
            login = st.button("Login to My Projects", key='login_button')
            if login:
                st.session_state.current_section = 'analysis'

    # Sidebar navigation using buttons
    st.sidebar.title('Navigation')

    # Define a function to handle navigation button clicks
    def navigate_to(section):
        st.session_state.current_section = section
        st.session_state.selected_project = None
        st.session_state.selected_subfolder = None

    # Create buttons in the sidebar for navigation
    if st.sidebar.button('Home'):
        navigate_to('home')
    if st.sidebar.button('Scenario Reports'):
        navigate_to('upload')
    if st.sidebar.button('Scenario Optimizer'):
        navigate_to('optimization')
    #if st.sidebar.button('My Projects'):
    #    navigate_to('analysis')
    #if st.sidebar.button('Upload Database'):
    #    navigate_to('upload database')
    if st.sidebar.button('About'):
        navigate_to('about')

    # If in 'My Projects', display folder tree in sidebar
    if st.session_state.current_section == 'analysis' and st.session_state.authenticated:
        st.sidebar.markdown("---")
        st.sidebar.subheader("Your Projects")
        user_projects = get_user_projects(st.session_state.username)
        subfolders = ["Baseline", "Round 1 Analysis", "Round 2 Analysis", "Supplier Feedback", "Negotiations"]
        for project in user_projects:
            with st.sidebar.expander(project, expanded=False):
                for subfolder in subfolders:
                    button_label = f"{project}/{subfolder}"
                    # Use unique key
                    if st.button(button_label, key=f"{project}_{subfolder}_sidebar"):
                        st.session_state.current_section = 'project_folder'
                        st.session_state.selected_project = project
                        st.session_state.selected_subfolder = subfolder

    # Display content based on the current section
    section = st.session_state.current_section
    if section == 'home':
        st.write("Welcome to the Scourcing COE Analysis Tool.")

        if not st.session_state.authenticated:
            st.write("You are not logged in. Please navigate to 'My Projects' to log in and access your projects.")
        else:
            st.write(f"Welcome, {st.session_state.username}!")

    elif section == 'analysis':
        st.title('My Projects')
        if not st.session_state.authenticated:
            st.header("Log In to Access My Projects")
            # Create a login form
            with st.form(key='login_form'):
                username = st.text_input("Username")
                password = st.text_input("Password", type='password')
                submit_button = st.form_submit_button(label='Login')

            if submit_button:
                if 'credentials' in config and 'usernames' in config['credentials']:
                    if username in config['credentials']['usernames']:
                        stored_hashed_password = config['credentials']['usernames'][username]['password']
                        # Verify the entered password against the stored hashed password
                        if bcrypt.checkpw(password.encode('utf-8'), stored_hashed_password.encode('utf-8')):
                            st.session_state.authenticated = True
                            st.session_state.username = username
                            st.success(f"Logged in as {username}")
                            logger.info(f"User {username} logged in successfully.")
                        else:
                            st.error("Incorrect password. Please try again.")
                            logger.warning(f"Incorrect password attempt for user {username}.")
                    else:
                        st.error("Username not found. Please check and try again.")
                        logger.warning(f"Login attempt with unknown username: {username}")
                else:
                    st.error("Authentication configuration is missing.")
                    logger.error("Authentication configuration is missing in 'config.yaml'.")

        else:
            # Project Management UI
            st.subheader("Manage Your Projects")
            col_start, col_delete = st.columns([1, 1])

            with col_start:
                # Start a New Project Form
                with st.form(key='create_project_form'):
                    st.markdown("### Create New Project")
                    new_project_name = st.text_input("Enter Project Name")
                    create_project_button = st.form_submit_button(label='Create Project')
                    if create_project_button:
                        if new_project_name.strip() == "":
                            st.error("Project name cannot be empty.")
                            logger.warning("Attempted to create a project with an empty name.")
                        else:
                            success = create_project(st.session_state.username, new_project_name.strip())
                            if success:
                                st.success(f"Project '{new_project_name.strip()}' created successfully.")

            with col_delete:
                # Delete a Project Form
                with st.form(key='delete_project_form'):
                    st.markdown("### Delete Existing Project")
                    projects = get_user_projects(st.session_state.username)
                    if projects:
                        project_to_delete = st.selectbox("Select Project to Delete", projects)
                        confirm_delete = st.form_submit_button(label='Confirm Delete')
                        if confirm_delete:
                            confirm = st.checkbox("I confirm that I want to delete this project.", key='confirm_delete_checkbox')
                            if confirm:
                                success = delete_project(st.session_state.username, project_to_delete)
                                if success:
                                    st.success(f"Project '{project_to_delete}' deleted successfully.")
                            else:
                                st.warning("Deletion not confirmed.")
                    else:
                        st.info("No projects available to delete.")
                        logger.info("No projects available to delete for the user.")

            st.markdown("---")
            st.subheader("Your Projects")

            projects = get_user_projects(st.session_state.username)
            if projects:
                for project in projects:
                    with st.container():
                        st.markdown(f"### {project}")
                        subfolders = ["Baseline", "Round 1 Analysis", "Round 2 Analysis", "Supplier Feedback", "Negotiations"]
                        cols = st.columns(len(subfolders))
                        for idx, subfolder in enumerate(subfolders):
                            button_label = f"{project}/{subfolder}"
                            # Use unique key
                            if cols[idx].button(button_label, key=f"{project}_{subfolder}_main"):
                                st.session_state.current_section = 'project_folder'
                                st.session_state.selected_project = project
                                st.session_state.selected_subfolder = subfolder
            else:
                st.info("No projects found. Start by creating a new project.")

    elif section == 'optimization':
        st.title("Scenario Builder & Optimization")

        ################################################################
        # 1) DEFAULT SCENARIOS (Best of Best = no rules, As-is = 1 rule)
        ################################################################
        if "scenario_definitions" not in st.session_state:
            st.session_state["scenario_definitions"] = {}

            # Best of Best scenario with no rules
            st.session_state["scenario_definitions"]["Best of Best"] = []

            # As-is scenario => 100% volume must go to Incumbent
            st.session_state["scenario_definitions"]["As-is"] = [
                {
                    "rule_type": "% of Volume Awarded",
                    "operator": "Exactly",
                    "rule_input": "100",   # Interpreted as 100%
                    "grouping": "All",
                    "grouping_scope": "All",
                    "supplier_scope": "Incumbent"
                }
            ]

        # Let the user upload Excel
        # ──────────────────────────────────────────────────────────────────────────────
        # Upload workbook & auto-convert if the user supplied the new 5-tab template
        # ──────────────────────────────────────────────────────────────────────────────
        uploaded_file = st.file_uploader("Upload an Excel file", type=["xlsx"])

        if uploaded_file is not None:
            st.markdown("### Uploaded Worksheets")
            sheet_dfs = load_excel_sheets(uploaded_file)

            # ──────────────────────────────────────────────────────────────
            # 1)  EXPLODE a compact **Bid Data** tab → classic tabs
            #     (Price · Demand · Supplier Bid Attributes)
            #     Baseline/Current prices are *not* required here any more.
            # ──────────────────────────────────────────────────────────────
            if "Bid Data" in sheet_dfs:
                classic_tabs = {"Price", "Demand", "Supplier Bid Attributes"}
                # purge duplicate classic versions if user kept them by mistake
                for dup in classic_tabs & sheet_dfs.keys():
                    sheet_dfs.pop(dup, None)

                sheet_dfs.update(split_bid_data_sheet(sheet_dfs["Bid Data"]))


            # ──────────────────────────────────────────────────────────────
            # 2)  If the **Baseline Price** tab is missing but we *do* have an
            #     Item Attributes sheet that contains Baseline Price / Current Price,
            #     build the legacy tab on-the-fly so the optimiser still works.
            # ──────────────────────────────────────────────────────────────
            if "Baseline Price" not in sheet_dfs and "Item Attributes" in sheet_dfs:
                ia_df = sheet_dfs["Item Attributes"]
                if {"Bid ID", "Baseline Price", "Current Price"} <= set(ia_df.columns):
                    sheet_dfs["Baseline Price"] = (
                        ia_df[["Bid ID", "Baseline Price", "Current Price"]]
                        .drop_duplicates("Bid ID")
                        .copy()
                    )

            # ──────────────────────────────────────────────────────────────
            # 3)  Map modern tier names → legacy names the optimiser expects
            # ──────────────────────────────────────────────────────────────
            if "Rebates" in sheet_dfs and "Rebate Tiers" not in sheet_dfs:
                sheet_dfs["Rebate Tiers"] = sheet_dfs["Rebates"]
            if "Volume Discounts" in sheet_dfs and "Discount Tiers" not in sheet_dfs:
                sheet_dfs["Discount Tiers"] = sheet_dfs["Volume Discounts"]

            # ──────────────────────────────────────────────────────────────
            # 4)  Basic column validation (after all auto-conversions)
            # ──────────────────────────────────────────────────────────────
            for sheet_name, df in sheet_dfs.items():
                missing = validate_sheet(df, sheet_name)
                if missing:
                    st.error(
                        f"Missing required column(s) in **{sheet_name}**: "
                        f"{', '.join(missing)}"
                    )

            # ──────────────────────────────────────────────────────────────
            # 5)  Detect whether freight columns exist (either template)
            # ──────────────────────────────────────────────────────────────
            price_sheet_name = "Price" if "Price" in sheet_dfs else "Bid Data"
            price_cols       = sheet_dfs[price_sheet_name].columns.str.strip().tolist()
            freight_enabled  = all(c in price_cols for c in ["Supplier Freight", "KBX"])
            st.info(
                f"Freight logic **{'ENABLED' if freight_enabled else 'DISABLED'}** "
                f"(detected in *{price_sheet_name}* tab)"
            )


            ################################################################
            # 2) CUSTOM RULES EXPANDER (with Save & Single-run)
            ################################################################
            with st.expander("Custom Rules (Current Rule Set)", expanded=True):
                st.markdown("#### Build & Manage Your Current Rule Set")

                # Left side: rule-building UI; Middle: scenario name; Right: "Save Scenario"
                col1, col2, col3 = st.columns([2, 1, 1])

                # -- (A) The rule-building UI in col1 --
                with col1:
                    rule_descriptions = {
                        "% of Volume Awarded": "Operator + % ensures the specified supplier scope receives that portion of volume.",
                        "# of Volume Awarded": "Operator + number of units awarded to the supplier scope.",
                        "# of Transitions": "Limits total transitions away from incumbents within a grouping.",
                        "# of Suppliers": "Restricts how many unique suppliers are used in a grouping.",
                        "Supplier Exclusion": "Excludes a chosen supplier from the grouping entirely; no rule input needed.",
                        "Exclude Bids": "Removes bids if a certain attribute meets the chosen operator and threshold/value.",
                        "# Minimum Volume Awarded": "Requires at least X units for a specific supplier scope.",
                        "% Minimum Volume Awarded": "Requires at least a certain % of volume for a supplier scope."
                    }

                    rule_type = st.selectbox(
                        "Rule Type", 
                        options=[
                            "% of Volume Awarded", "# of Volume Awarded", "# of Transitions",
                            "# of Suppliers", "Supplier Exclusion", "Exclude Bids"
                        ],
                        key="rule_type_select"
                    )
                    st.markdown(f"<small>*{rule_descriptions.get(rule_type, '')}*</small>", unsafe_allow_html=True)

                    col_op, col_in = st.columns(2)
                    with col_op:
                        if rule_type == "Supplier Exclusion":
                            operator = "Exactly"
                            st.selectbox("Operator", ["Exactly"], disabled=True, key="operator_select")
                        elif rule_type in ["% Minimum Volume Awarded", "# Minimum Volume Awarded"]:
                            operator = "At least"
                            st.selectbox("Operator", ["At least"], disabled=True, key="operator_select")
                        elif rule_type == "Exclude Bids":
                            st.empty()
                            operator = ""
                        else:
                            operator = st.selectbox("Operator", ["At least", "At most", "Exactly"], key="operator_select")

                    with col_in:
                        rule_input = ""
                        if rule_type == "% of Volume Awarded":
                            pct_val = st.number_input("Rule Input (as %)", 0.0, 100.0, 50.0, step=0.01)
                            rule_input = f"{pct_val:.2f}%"
                        elif rule_type == "# of Volume Awarded":
                            vol_val = st.number_input("Rule Input (#units)", 0, 999999, 0)
                            rule_input = str(vol_val)
                        elif rule_type == "# of Transitions":
                            trans_val = st.number_input("Rule Input (#transitions)", 0, 999999, 0)
                            rule_input = str(trans_val)
                        elif rule_type == "% Minimum Volume Awarded":
                            pct_val = st.number_input("Rule Input (as %)", 0.0, 100.0, 50.0, step=0.01)
                            rule_input = f"{pct_val:.2f}%"
                        elif rule_type == "# Minimum Volume Awarded":
                            vol_val = st.number_input("Rule Input (#units)", 0, 999999, 0)
                            rule_input = str(vol_val)
                        elif rule_type == "Supplier Exclusion":
                            st.text_input("Rule Input", value="", disabled=True)
                            rule_input = ""
                        elif rule_type != "Exclude Bids":
                            rule_input = st.text_input("Rule Input")

                    # If "Exclude Bids," do numeric/text logic
                    bid_grouping = None
                    bid_exclusion_value = None
                    if rule_type == "Exclude Bids":
                        supplier_bid_attr_df = sheet_dfs.get("Supplier Bid Attributes")
                        if supplier_bid_attr_df is not None:
                            bid_attr_columns = [c for c in supplier_bid_attr_df.columns if c not in ["Supplier Name", "Bid ID"]]
                        else:
                            bid_attr_columns = []

                        cb1, cb2 = st.columns([1,1])
                        with cb1:
                            bid_grouping = st.selectbox("Bid Grouping", bid_attr_columns, key="bid_grouping_select")

                        is_numeric = False
                        if supplier_bid_attr_df is not None and bid_grouping in supplier_bid_attr_df.columns:
                            non_null = supplier_bid_attr_df[bid_grouping].dropna()
                            if not non_null.empty:
                                try:
                                    float(non_null.iloc[0])
                                    is_numeric = True
                                except:
                                    is_numeric = False

                        with cb2:
                            if is_numeric:
                                operator = st.selectbox("Operator", ["Greater than", "Less than", "Equal To"], key="operator_select_exclude_bids")
                            else:
                                operator = st.selectbox("Operator", ["Equal To"], disabled=True, key="operator_select_exclude_bids")

                        if is_numeric:
                            thr = st.number_input("Rule Input (numeric threshold)", 0.0, 9999999.0, 0.0, step=0.01, key="rule_input_bid_exclusion_numeric")
                            rule_input = str(thr)
                            bid_exclusion_value = None
                        else:
                            rule_input = ""
                            if supplier_bid_attr_df is not None and bid_grouping in supplier_bid_attr_df.columns:
                                unique_vals = sorted(supplier_bid_attr_df[bid_grouping].dropna().unique())
                            else:
                                unique_vals = []
                            bid_exclusion_value = st.selectbox("Bid Exclusion Value", unique_vals, key="bid_exclusion_value_select")

                    # Grouping
                    if "Item Attributes" in sheet_dfs:
                        item_attr_temp = df_to_dict_item_attributes(sheet_dfs["Item Attributes"])
                        sample_keys = list(next(iter(item_attr_temp.values())).keys())
                        grouping_options = ["All", "Bid ID"] + sample_keys
                    else:
                        grouping_options = ["All", "Bid ID"]

                    cgr1, cgr2 = st.columns(2)
                    with cgr1:
                        grouping_sel = st.selectbox("Grouping", grouping_options, key="grouping_select")
                    with cgr2:
                        def update_grouping_scope(g, item_attr):
                            if g == "Bid ID":
                                return sorted(list(item_attr.keys()))
                            else:
                                return sorted({
                                    str(item_attr[bid].get(g, "")).strip()
                                    for bid in item_attr if str(item_attr[bid].get(g, "")).strip() != ""
                                })

                        if grouping_sel != "All":
                            grouping_scope_opts = ["Apply to all items individually"] + update_grouping_scope(grouping_sel, item_attr_temp)
                            grouping_scope_sel = st.selectbox("Grouping Scope", grouping_scope_opts, key="grouping_scope_select")
                        else:
                            grouping_scope_sel = "All"

                    # Supplier scope
                    if rule_type in ["# of Transitions", "# of Suppliers"]:
                        supplier_scope = "All"
                        st.selectbox("Supplier Scope", ["All"], disabled=True)
                    elif rule_type == "Exclude Bids":
                        supplier_scope = st.selectbox("Supplier Scope", ["All"], disabled=True)
                    elif rule_type in ["% of Volume Awarded", "# of Volume Awarded"]:
                        if "Price" in sheet_dfs:
                            s_auto = sheet_dfs["Price"]["Supplier Name"].dropna().astype(str).str.strip().unique().tolist()
                        else:
                            s_auto = []
                        def_scopes = ["All"] + s_auto + ["New Suppliers", "Lowest cost supplier", "Second Lowest Cost Supplier", "Incumbent"]
                        supplier_scope = st.selectbox("Supplier Scope", def_scopes)
                    elif rule_type in ["% Minimum Volume Awarded", "# Minimum Volume Awarded"]:
                        if "Price" in sheet_dfs:
                            s_auto = sheet_dfs["Price"]["Supplier Name"].dropna().astype(str).str.strip().unique().tolist()
                        else:
                            s_auto = []
                        def_scopes = ["All"] + s_auto + ["New Suppliers", "Lowest cost supplier", "Second Lowest Cost Supplier", "Incumbent"]
                        supplier_scope = st.selectbox("Supplier Scope", def_scopes)
                    else:
                        if "Price" in sheet_dfs:
                            s_auto = sheet_dfs["Price"]["Supplier Name"].dropna().astype(str).str.strip().unique().tolist()
                        else:
                            s_auto = []
                        def_scopes = s_auto + ["New Suppliers", "Lowest cost supplier", "Second Lowest Cost Supplier", "Incumbent"]
                        supplier_scope = st.selectbox("Supplier Scope", def_scopes)
                        if supplier_scope == "All":
                            supplier_scope = None

                    rcol_add, rcol_clear = st.columns(2)
                    with rcol_add:
                        if st.button("Add Rule"):
                            errors = []
                            # Validation
                            if rule_type in ["% of Volume Awarded", "% Minimum volume awarded"]:
                                try:
                                    p_test = float(rule_input.rstrip("%"))
                                    if p_test < 0 or p_test > 100:
                                        errors.append("Percentage must be between 0 and 100.")
                                except:
                                    errors.append("Invalid percentage format (e.g. '50.00%').")

                            elif rule_type in ["# of Volume Awarded", "# of transitions", "# Minimum volume awarded"]:
                                try:
                                    float(rule_input)
                                except:
                                    errors.append("Please enter a valid numeric value for the rule input.")

                            if rule_type == "Exclude Bids":
                                sbad = df_to_dict_supplier_bid_attributes(sheet_dfs["Supplier Bid Attributes"])
                                if not is_bid_attribute_numeric(bid_grouping, sbad):
                                    if not bid_exclusion_value:
                                        errors.append("Please select a Bid Exclusion Value.")

                            if errors:
                                for e in errors:
                                    st.error(e)
                                st.stop()
                            else:
                                new_rule = {
                                    "rule_type": rule_type,
                                    "operator": operator,
                                    "rule_input": rule_input,
                                    "grouping": grouping_sel,
                                    "grouping_scope": grouping_scope_sel,
                                    "supplier_scope": supplier_scope
                                }
                                if rule_type == "Exclude Bids":
                                    new_rule["bid_grouping"] = bid_grouping
                                    new_rule["bid_exclusion_value"] = bid_exclusion_value

                                if "rules_list" not in st.session_state:
                                    st.session_state.rules_list = []
                                st.session_state.rules_list.append(new_rule)
                                st.success("Rule added.")

                    with rcol_clear:
                        if st.button("Clear Rules"):
                            st.session_state.rules_list = []
                            st.success("All rules cleared.")

                # (B) Middle & Right columns for scenario name & Save button
                with col2:
                    st.write("**Scenario Name**")
                    scenario_name = st.text_input("", label_visibility="collapsed")
                with col3:
                    st.write("")
                    if st.button("Save Scenario"):
                        if not scenario_name.strip():
                            st.error("Scenario name cannot be empty.")
                        elif not st.session_state.get("rules_list"):
                            st.error("No rules to save.")
                        else:
                            st.session_state["scenario_definitions"][scenario_name] = list(st.session_state["rules_list"])
                            st.success(f"Scenario '{scenario_name}' saved with {len(st.session_state['rules_list'])} rules.")
                # Show the current unsaved rules
                temp_item_attr = df_to_dict_item_attributes(sheet_dfs["Item Attributes"])
                if st.session_state.get("rules_list"):
                    st.markdown("#### Current Rule Set")
                    for i, rule_obj in enumerate(st.session_state["rules_list"]):
                        ccol_rule, ccol_delete = st.columns([0.95, 0.05])
                        with ccol_rule:
                            st.markdown(f"<div class='rule-text'>{expand_rule_text(rule_obj, temp_item_attr)}</div>", unsafe_allow_html=True)
                        with ccol_delete:
                            if st.button("X", key=f"delete_rule_{i}"):
                                st.session_state["rules_list"].pop(i)
                                st.rerun()

                # Single-run button for the current rule list
                st.markdown("#### Run Current Rule Set")
                if st.button("Run Current Ruleset"):
                    required_sheet_names = [
                        "Item Attributes", "Price", "Demand",
                        "Rebate Tiers", "Discount Tiers",        # legacy
                        "Rebates", "Volume Discounts",           # modern
                        "Baseline Price", "Capacity",
                        "Supplier Bid Attributes", "Bid Data"    # Bid Data is optional once split
                    ]
                    missing_sheets = [sheet for sheet in required_sheet_names if sheet not in sheet_dfs]
                    if missing_sheets:
                        st.error(f"Missing required sheets: {', '.join(missing_sheets)}. Cannot run optimization.")
                        st.stop()

                    try:
                        # =========== UPDATED df_to_dict_baseline_price ==============
                        # Now store both baseline + current in a dictionary
                        def df_to_dict_baseline_price(df):
                            d = {}
                            for _, row in df.iterrows():
                                bid = normalize_bid_id(row["Bid ID"])
                                baseline_val = row.get("Baseline Price", 0.0)
                                current_val  = row.get("Current Price", 0.0)
                                d[bid] = {
                                    "baseline": baseline_val,
                                    "current":  current_val
                                }
                            return d

                        # =========== Load data ==============
                        item_attr_dict = df_to_dict_item_attributes(sheet_dfs["Item Attributes"])
                        price_dict = df_to_dict_price(sheet_dfs["Price"])
                        demand_dict = df_to_dict_demand(sheet_dfs["Demand"])
                        baseline_price_dict = df_to_dict_baseline_price(sheet_dfs["Baseline Price"])  # Updated
                        capacity_dict = df_to_dict_capacity(sheet_dfs["Capacity"])
                        rebate_tiers_dict = df_to_dict_tiers(sheet_dfs["Rebate Tiers"])
                        discount_tiers_dict = df_to_dict_tiers(sheet_dfs["Discount Tiers"])
                        supplier_bid_attr_dict = df_to_dict_supplier_bid_attributes(sheet_dfs["Supplier Bid Attributes"])
                        suppliers = sheet_dfs["Price"]["Supplier Name"].dropna().astype(str).str.strip().unique().tolist()

                        # =========== Updated run_optimization with Current Price logic ===========
                        with st.spinner("Running optimization..."):
                            out_file, feasibility_notes, model_status = run_optimization(
                                capacity_dict,
                                demand_dict,
                                item_attr_dict,
                                price_dict,
                                rebate_tiers_dict,
                                discount_tiers_dict,
                                baseline_price_dict,            # includes baseline + current
                                rules=st.session_state.get("rules_list", []),
                                supplier_bid_attr_dict=supplier_bid_attr_dict,
                                suppliers=suppliers,
                                freight_enabled=freight_enabled 
                            )

                        st.success(f"Model Status: {model_status}")
                        if model_status == "Infeasible":
                            st.warning("This scenario was infeasible. No feasible solution found.")
                        st.markdown("#### Feasibility Notes")
                        st.text(feasibility_notes)
                        st.markdown("#### Optimization Results")
                        try:
                            # We assume run_optimization now produces columns:
                            #   Bid ID, Bid ID Split, Facility, Incumbent, Baseline Price, Current Price, ...
                            df_results = pd.read_excel(out_file, sheet_name="Results")
                            st.dataframe(df_results)
                            with open(out_file, "rb") as f:
                                st.download_button("Download Results", f, file_name="optimization_results.xlsx")
                        except Exception as e:
                            st.warning("Unable to read 'Results' sheet. Possibly no solution was found.")
                    except Exception as e:
                        st.error(f"Error preparing data for optimization: {e}")

            ################################################################
            # SCENARIO SUMMARY GROUPING (for PPT details) [Optional]
            ################################################################
            st.markdown("### Scenario Summary (PowerPoint) Configuration (Optional)")
            if st.session_state.merged_data is not None:
                grouping_options = st.session_state.merged_data.columns.tolist()
                st.selectbox("Group by for Scenario Detail (PPT)", grouping_options, key="scenario_detail_grouping")

                # (Optional) sub-summaries
                st.toggle("Include Sub-Scenario Summaries?", key="scenario_sub_summaries_on")
                if st.session_state["scenario_sub_summaries_on"]:
                    scenario_summary_fields = st.session_state.merged_data.columns.tolist()
                    st.selectbox("Scenario Summaries Selections", scenario_summary_fields, key="scenario_summary_selections")
                    sub_vals = st.session_state.merged_data[st.session_state.scenario_summary_selections].unique()
                    st.pills(
                        "Select scenario sub-summaries",
                        sub_vals,
                        selection_mode="multi",
                        key="sub_summary_selections"
                    )
            else:
                st.info("No merged_data available; cannot configure scenario detail grouping yet.")

            ################################################################
            # 3) SAVED SCENARIOS EXPANDER (with “Run All Scenarios”)
            ################################################################
            with st.expander("Saved Scenarios", expanded=True):
                st.markdown("#### Manage & Run All Scenarios")

                if not st.session_state["scenario_definitions"]:
                    st.info("No saved scenarios yet.")
                else:
                    for scen_name, rules_set in list(st.session_state["scenario_definitions"].items()):
                        row1, row2 = st.columns([0.8, 0.2])
                        with row1:
                            st.write(f"**Scenario:** {scen_name} ({len(rules_set)} rules)")
                        with row2:
                            if st.button(f"Delete {scen_name}"):
                                del st.session_state["scenario_definitions"][scen_name]
                                st.rerun()

                st.markdown("#### Select Output Format")
                output_format = st.radio(
                    "Choose output format(s)",
                    options=["Excel only", "PowerPoint only", "Both"],
                    index=0
                )

                if st.button("Run All Scenarios"):
                    if not st.session_state["scenario_definitions"]:
                        st.error("No scenarios have been saved.")
                        st.stop()

                    required_sheet_names = [
                        "Item Attributes", "Price", "Demand",
                        "Rebate Tiers", "Discount Tiers",        # legacy
                        "Rebates", "Volume Discounts",           # modern
                        "Baseline Price", "Capacity",
                        "Supplier Bid Attributes", "Bid Data"    # Bid Data is optional once split
                    ]
                    missing_sheets = [sheet for sheet in required_sheet_names if sheet not in sheet_dfs]
                    if missing_sheets:
                        st.error(f"Missing required sheets: {', '.join(missing_sheets)}. Cannot run optimization.")
                        st.stop()

                    # =========== UPDATED df_to_dict_baseline_price for multi-scenario =============
                    def df_to_dict_baseline_price(df):
                        d = {}
                        for _, row in df.iterrows():
                            bid = normalize_bid_id(row["Bid ID"])
                            baseline_val = row.get("Baseline Price", 0.0)
                            current_val  = row.get("Current Price", 0.0)
                            d[bid] = {
                                "baseline": baseline_val,
                                "current":  current_val
                            }
                        return d

                    try:
                        item_attr_dict = df_to_dict_item_attributes(sheet_dfs["Item Attributes"])
                        price_dict = df_to_dict_price(sheet_dfs["Price"])
                        demand_dict = df_to_dict_demand(sheet_dfs["Demand"])
                        baseline_price_dict = df_to_dict_baseline_price(sheet_dfs["Baseline Price"])  # new function
                        capacity_dict = df_to_dict_capacity(sheet_dfs["Capacity"])
                        rebate_tiers_dict = df_to_dict_tiers(sheet_dfs["Rebate Tiers"])
                        discount_tiers_dict = df_to_dict_tiers(sheet_dfs["Discount Tiers"])
                        supplier_bid_attr_dict = df_to_dict_supplier_bid_attributes(sheet_dfs["Supplier Bid Attributes"])
                        suppliers = sheet_dfs["Price"]["Supplier Name"].dropna().astype(str).str.strip().unique().tolist()
                    except Exception as e:
                        st.error(f"Error preparing data for scenario runs: {e}")
                        st.stop()

                    from io import BytesIO
                    scenario_results_dict = {}
                    infeasible_scenarios = []

                    with st.spinner("Running all scenarios..."):
                        for scenario_nm, the_rules in st.session_state["scenario_definitions"].items():
                            try:
                                out_file, feas, status = run_optimization(
                                    capacity_data       = capacity_dict,
                                    demand_data         = demand_dict,
                                    item_attr_data      = item_attr_dict,
                                    price_data          = price_dict,
                                    rebate_tiers        = rebate_tiers_dict,
                                    discount_tiers      = discount_tiers_dict,
                                    baseline_price_data = baseline_price_dict,
                                    rules               = the_rules,
                                    supplier_bid_attr_dict = supplier_bid_attr_dict,
                                    suppliers           = suppliers,
                                    freight_enabled     = freight_enabled   # ← NEW ARG
                                )

                                if status == "Infeasible":
                                    infeasible_scenarios.append(scenario_nm)

                                try:
                                    df_res = pd.read_excel(out_file, sheet_name="Results")
                                except:
                                    df_res = pd.DataFrame()

                                scenario_results_dict[scenario_nm] = df_res
                            except Exception as e:
                                st.warning(f"Scenario '{scenario_nm}' failed: {e}")
                                scenario_results_dict[scenario_nm] = pd.DataFrame()

                    excel_output = BytesIO()
                    ppt_output = BytesIO()
                    scenario_ppt_data = None

                    # ───────────── Compile Excel (freight-aware ordering) ─────────────
                    if output_format in ["Excel only", "Both"]:
                        excel_output = BytesIO()
                        with pd.ExcelWriter(excel_output, engine="openpyxl") as writer:
                            for s_name, df_scen in scenario_results_dict.items():

                                # always write a sheet so the workbook keeps its order
                                if df_scen.empty:
                                    df_scen.to_excel(
                                        writer,
                                        sheet_name=(s_name[:31] or "Sheet1"),
                                        index=False
                                    )
                                    continue

                                # ---------- column ordering ----------
                                base_cols = [
                                    "Bid ID", "Bid ID Split", "Facility", "Incumbent",
                                    "Baseline Price", "Current Price", "Baseline Spend",
                                    "Awarded Supplier", "Original Awarded Supplier Price",
                                    "Percentage Volume Discount", "Discounted Awarded Supplier Price"
                                ]
                                freight_cols = ["Freight Method", "Freight Amount", "Effective Supplier Price"]
                                tail_cols = [
                                    "Awarded Supplier Spend", "Awarded Volume",
                                    "Baseline Savings", "Current Price Savings",
                                    "Rebate %", "Rebate Savings"
                                ]

                                desired_cols = (
                                    base_cols +
                                    (freight_cols if freight_enabled else []) +   # ← NEW: include only when freight is ON
                                    tail_cols
                                )
                                ordered = [c for c in desired_cols if c in df_scen.columns]
                                ordered += [c for c in df_scen.columns if c not in ordered]

                                df_scen = df_scen[ordered]
                                df_scen.to_excel(
                                    writer,
                                    sheet_name=(s_name[:31] or "Sheet1"),
                                    index=False
                                )

                        excel_output.seek(0)
                        st.download_button(
                            "Download All Scenario Results (Excel)",
                            data=excel_output.getvalue(),
                            file_name="all_scenarios_results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        st.success("All scenario results compiled into an Excel file.")
                    else:
                        st.error("No scenario produced valid results. Excel was not generated.")


                    # (B) PowerPoint
                    if output_format in ["PowerPoint only", "Both"]:
                        # Build scenario_dataframes for PPT
                        scenario_dataframes = {}
                        for scen_name, df_res in scenario_results_dict.items():
                            scenario_dataframes[f"#{scen_name}"] = df_res.copy()

                        # Possibly rename columns
                        for key, dfp in scenario_dataframes.items():
                            if "Baseline Savings" not in dfp.columns and "Savings" in dfp.columns:
                                dfp["Baseline Savings"] = dfp["Savings"]
                            if "Supplier Name" in dfp.columns and "Awarded Supplier" not in dfp.columns:
                                dfp["Awarded Supplier"] = dfp["Supplier Name"]

                        script_dir = os.path.dirname(os.path.abspath(__file__))
                        template_file_path = os.path.join(script_dir, 'Slide template.pptx')

                        prs = create_scenario_summary_presentation(scenario_dataframes, template_file_path=template_file_path)
                        if prs is not None:
                            prs.save(ppt_output)
                            scenario_ppt_data = ppt_output.getvalue()
                            st.success("All scenario results compiled into a PowerPoint file.")

                            st.download_button(
                                label="Download Scenario Summary (PowerPoint)",
                                data=scenario_ppt_data,
                                file_name="scenario_summary.pptx",
                                mime="application/vnd.openxmlformats-officedocument.presentationml.presentation"
                            )
                        else:
                            st.error("Failed to generate PowerPoint summary.")

                    if infeasible_scenarios:
                        st.warning(
                            "Some scenarios were infeasible: "
                            + ", ".join(infeasible_scenarios)
                            + ". Their sheets/slides may be empty or partial."
                        )

            # Final check for any leftover required sheets
            required_sheet_names = [
                "Item Attributes", "Price", "Demand",
                "Rebate Tiers", "Discount Tiers",        # legacy
                "Rebates", "Volume Discounts",           # modern
                "Baseline Price", "Capacity",
                "Supplier Bid Attributes", "Bid Data"    # Bid Data is optional once split
            ]
            missing_sheets = [sheet for sheet in required_sheet_names if sheet not in sheet_dfs]
            if missing_sheets:
                st.error(
                    f"Missing required sheets: {', '.join(missing_sheets)}. "
                    "Please upload an Excel file that contains all of these sheets."
                )
                st.stop()





    elif section == 'upload':
        st.title('Start a New Analysis')
        st.write("Upload your data here.")

        # Select data input method
        data_input_method = st.radio(
            "Select Data Input Method",
            ('Separate Bid & Baseline files', 'Merged Data'),
            index=0,
            key='data_input_method'
        )

        # Use the selected method in your code
        if data_input_method == 'Separate Bid & Baseline files':
            # Existing steps for separate files
            st.header("Upload Baseline and Bid Files")

            # Upload baseline file
            baseline_file = st.file_uploader("Upload Baseline Sheet", type=["xlsx"])

            # Sheet selection for Baseline File
            baseline_sheet = None
            if baseline_file:
                try:
                    excel_baseline = pd.ExcelFile(baseline_file, engine='openpyxl')
                    baseline_sheet = st.selectbox(
                        "Select Baseline Sheet",
                        excel_baseline.sheet_names,
                        key='baseline_sheet_selection'
                    )
                except Exception as e:
                    st.error(f"Error reading baseline file: {e}")
                    logger.error(f"Error reading baseline file: {e}")

            num_files = st.number_input("Number of Bid Sheets to Upload", min_value=1, step=1)

            bid_files_suppliers = []
            for i in range(int(num_files)):
                bid_file = st.file_uploader(f"Upload Bid Sheet {i + 1}", type=["xlsx"], key=f'bid_file_{i}')
                supplier_name = st.text_input(f"Supplier Name for Bid Sheet {i + 1}", key=f'supplier_name_{i}')
                # Sheet selection for each Bid File
                bid_sheet = None
                if bid_file and supplier_name:
                    try:
                        excel_bid = pd.ExcelFile(bid_file, engine='openpyxl')
                        bid_sheet = st.selectbox(
                            f"Select Sheet for Bid Sheet {i + 1}",
                            excel_bid.sheet_names,
                            key=f'bid_sheet_selection_{i}'
                        )
                    except Exception as e:
                        st.error(f"Error reading Bid Sheet {i + 1}: {e}")
                        logger.error(f"Error reading Bid Sheet {i + 1}: {e}")
                if bid_file and supplier_name and bid_sheet:
                    bid_files_suppliers.append((bid_file, supplier_name, bid_sheet))
                    logger.info(f"Uploaded Bid Sheet {i + 1} for supplier '{supplier_name}' with sheet '{bid_sheet}'.")

            # Merge Data
            if st.button("Merge Data"):
                if validate_uploaded_file(baseline_file) and bid_files_suppliers:
                    if not baseline_sheet:
                        st.error("Please select a sheet for the baseline file.")
                    else:
                        merged_data = start_process(baseline_file, baseline_sheet, bid_files_suppliers)
                        if merged_data is not None:
                            st.session_state.merged_data = merged_data
                            st.session_state.original_merged_data = merged_data.copy()
                            st.session_state.columns = list(merged_data.columns)
                            st.session_state.baseline_data = load_baseline_data(baseline_file, baseline_sheet)
                            st.success("Data merged successfully. Please map the columns for analysis.")
                            logger.info("Data merged successfully.")

        else:
            # For Merged Data input method
            st.header("Upload Merged Data File")
            merged_file = st.file_uploader("Upload Merged Data File", type=["xlsx"], key='merged_data_file')
            merged_sheet = None
            if merged_file:
                try:
                    excel_merged = pd.ExcelFile(merged_file, engine='openpyxl')
                    merged_sheet = st.selectbox(
                        "Select Sheet",
                        excel_merged.sheet_names,
                        key='merged_sheet_selection'
                    )
                except Exception as e:
                    st.error(f"Error reading merged data file: {e}")
                    logger.error(f"Error reading merged data file: {e}")

            if merged_file and merged_sheet:
                try:
                    merged_data = pd.read_excel(merged_file, sheet_name=merged_sheet, engine='openpyxl')
                    # Normalize columns
                    merged_data = normalize_columns(merged_data)
                    st.session_state.merged_data = merged_data
                    st.session_state.original_merged_data = merged_data.copy()
                    st.session_state.columns = list(merged_data.columns)

                    st.success("Merged data loaded successfully. Please map the columns for analysis.")
                    logger.info("Merged data loaded successfully.")
                except Exception as e:
                    st.error(f"Error loading merged data: {e}")
                    logger.error(f"Error loading merged data: {e}")

        # Proceed to Column Mapping if merged_data is available
        if st.session_state.merged_data is not None:
            required_columns = ['Bid ID', 'Incumbent', 'Facility', 'Baseline Price', 'Current Price',
                    'Bid Volume', 'Bid Price', 'Supplier Capacity', 'Supplier Name']



            # Initialize column_mapping if not already initialized
            if 'column_mapping' not in st.session_state:
                st.session_state.column_mapping = {}

            # Map columns
            st.write("Map the following columns:")
            for col in required_columns:
                if col == 'Current Price':
                    options = ['None'] + list(st.session_state.merged_data.columns)
                    st.session_state.column_mapping[col] = st.selectbox(
                        f"Select Column for {col}",
                        options,
                        key=f"{col}_mapping"
                    )
                else:
                    st.session_state.column_mapping[col] = st.selectbox(
                        f"Select Column for {col}",
                        st.session_state.merged_data.columns,
                        key=f"{col}_mapping"
                    )


            # After mapping, check if all required columns are mapped
            missing_columns = [col for col in required_columns if col not in st.session_state.column_mapping or st.session_state.column_mapping[col] not in st.session_state.merged_data.columns]
            if missing_columns:
                st.error(f"The following required columns are not mapped or do not exist in the data: {', '.join(missing_columns)}")
            else:
                # After mapping, set 'Awarded Supplier' automatically
                st.session_state.merged_data['Awarded Supplier'] = st.session_state.merged_data[st.session_state.column_mapping['Supplier Name']]


            analyses_to_run = st.multiselect("Select Scenario Analyses to Run", [
                "As-Is",
                "Best of Best",
                "Best of Best Excluding Suppliers",
                "As-Is Excluding Suppliers",
                "Bid Coverage Report",
                "Customizable Analysis"
            ])

                # New Presentation Summaries multi-select
            presentation_options = ["Scenario Summary", "Bid Coverage Summary", "Supplier Comparison Summary"] 
            selected_presentations = st.multiselect("Presentation Summaries", options=presentation_options)


            # Exclusion rules for Best of Best Excluding Suppliers
            if "Best of Best Excluding Suppliers" in analyses_to_run:
                with st.expander("Configure Exclusion Rules for Best of Best Excluding Suppliers"):
                    st.header("Exclusion Rules for Best of Best Excluding Suppliers")

                    # Select Supplier to Exclude
                    supplier_name = st.selectbox(
                        "Select Supplier to Exclude",
                        st.session_state.merged_data['Awarded Supplier'].dropna().unique(),
                        key="supplier_name_excl_bob"
                    )

                    # Select Field for Rule
                    field = st.selectbox(
                        "Select Field for Rule",
                        st.session_state.merged_data.columns.drop('Awarded Supplier'),  # Exclude 'Awarded Supplier' if necessary
                        key="field_excl_bob"
                    )

                    # Select Logic
                    logic = st.selectbox(
                        "Select Logic (Equal to or Not equal to)",
                        ["Equal to", "Not equal to"],
                        key="logic_excl_bob"
                    )

                    # Select Value based on chosen field
                    unique_values = st.session_state.merged_data[field].dropna().unique()

                    if unique_values.size > 0:
                        value = st.selectbox(
                            "Select Value",
                            unique_values,
                            key="value_excl_bob"
                        )
                    else:
                        value = st.selectbox(
                            "Select Value",
                            options=[0],  # You can change 0 to any default value you prefer
                            index=0,
                            key="value_excl_bob"
                        )


                    # Checkbox to exclude all Bid IDs from the supplier
                    exclude_all = st.checkbox("Exclude from all Bid IDs", key="exclude_all_excl_bob")

                    # Button to add exclusion rule
                    if st.button("Add Exclusion Rule", key="add_excl_bob"):
                        if 'exclusions_bob' not in st.session_state:
                            st.session_state.exclusions_bob = []
                        # Append the exclusion rule as a tuple
                        st.session_state.exclusions_bob.append((supplier_name, field, logic, value, exclude_all))
                        logger.debug(f"Added exclusion rule for BOB Excl Suppliers: {supplier_name}, {field}, {logic}, {value}, Exclude All: {exclude_all}")
                        st.success("Exclusion rule added successfully!")

                    # Button to clear all exclusion rules
                    if st.button("Clear Exclusion Rules", key="clear_excl_bob"):
                        st.session_state.exclusions_bob = []
                        logger.debug("Cleared all exclusion rules for BOB Excl Suppliers.")
                        st.success("All exclusion rules cleared.")

                    # Display current exclusion rules
                    if 'exclusions_bob' in st.session_state and st.session_state.exclusions_bob:
                        st.write("**Current Exclusion Rules for Best of Best Excluding Suppliers:**")
                        for i, excl in enumerate(st.session_state.exclusions_bob):
                            st.write(f"{i + 1}. Supplier: {excl[0]}, Field: {excl[1]}, Logic: {excl[2]}, Value: {excl[3]}, Exclude All: {excl[4]}")

            # Exclusion rules for As-Is Excluding Suppliers
            if "As-Is Excluding Suppliers" in analyses_to_run:
                with st.expander("Configure Exclusion Rules for As-Is Excluding Suppliers"):
                    st.header("Exclusion Rules for As-Is Excluding Suppliers")

                    supplier_name_ais = st.selectbox("Select Supplier to Exclude", st.session_state.merged_data['Awarded Supplier'].unique(), key="supplier_name_excl_ais")
                    field_ais = st.selectbox("Select Field for Rule", st.session_state.merged_data.columns, key="field_excl_ais")
                    logic_ais = st.selectbox("Select Logic (Equal to or Not equal to)", ["Equal to", "Not equal to"], key="logic_excl_ais")
                    value_ais = st.selectbox("Select Value", st.session_state.merged_data[field_ais].unique(), key="value_excl_ais")
                    exclude_all_ais = st.checkbox("Exclude from all Bid IDs", key="exclude_all_excl_ais")

                    if st.button("Add Exclusion Rule", key="add_excl_ais"):
                        if 'exclusions_ais' not in st.session_state:
                            st.session_state.exclusions_ais = []
                        st.session_state.exclusions_ais.append((supplier_name_ais, field_ais, logic_ais, value_ais, exclude_all_ais))
                        logger.debug(f"Added exclusion rule for As-Is Excl Suppliers: {supplier_name_ais}, {field_ais}, {logic_ais}, {value_ais}, Exclude All: {exclude_all_ais}")

                    if st.button("Clear Exclusion Rules", key="clear_excl_ais"):
                        st.session_state.exclusions_ais = []
                        logger.debug("Cleared all exclusion rules for As-Is Excl Suppliers.")

                    if 'exclusions_ais' in st.session_state and st.session_state.exclusions_ais:
                        st.write("Current Exclusion Rules for As-Is Excluding Suppliers:")
                        for i, excl in enumerate(st.session_state.exclusions_ais):
                            st.write(f"{i + 1}. Supplier: {excl[0]}, Field: {excl[1]}, Logic: {excl[2]}, Value: {excl[3]}, Exclude All: {excl[4]}")

            # Bid Coverage Report Configuration
            if "Bid Coverage Report" in analyses_to_run:
                with st.expander("Configure Bid Coverage Report"):
                    st.header("Bid Coverage Report Configuration")

                    # Select variations
                    bid_coverage_variations = st.multiselect("Select Bid Coverage Report Variations", [
                        "Competitiveness Report",
                        "Supplier Coverage",
                        "Facility Coverage"
                    ], key="bid_coverage_variations")

                    # Select group by field
                    group_by_field = st.selectbox("Group by", st.session_state.merged_data.columns, key="bid_coverage_group_by")


            if "Customizable Analysis" in analyses_to_run:
                with st.expander("Configure Customizable Analysis"):
                    st.header("Customizable Analysis Configuration")
                    # Select Grouping Column
                    grouping_column_raw = st.selectbox(
                        "Select Grouping Column",
                        st.session_state.merged_data.columns,
                        key="customizable_grouping_column"
                    )
                    
                    # Check if the selected grouping column is already mapped
                    grouping_column_already_mapped = False
                    grouping_column_mapped = grouping_column_raw  # Default to raw name
                    for standard_col, mapped_col in st.session_state.column_mapping.items():
                        if mapped_col == grouping_column_raw:
                            grouping_column_mapped = standard_col
                            grouping_column_already_mapped = True
                            break
                    
                    # Store both the raw and mapped grouping column names and the flag
                    st.session_state.grouping_column_raw = grouping_column_raw
                    st.session_state.grouping_column_mapped = grouping_column_mapped
                    st.session_state.grouping_column_already_mapped = grouping_column_already_mapped


            if "Scenario Summary" in selected_presentations:
                with st.expander("Configure Grouping for Scenario Summary Slides"):
                    st.header("Scenario Summary Grouping")

                    grouping_options = st.session_state.merged_data.columns.tolist()
                    # This will store the selected grouping in st.session_state["scenario_detail_grouping"]
                    st.selectbox("Group by for Scenario Detail", grouping_options, key="scenario_detail_grouping")

                    # Just call st.toggle with a key. Do not assign to st.session_state again.
                    st.toggle("Include Sub-Scenario Summaries?", key="scenario_sub_summaries_on")

                    if st.session_state["scenario_sub_summaries_on"]:
                        scenario_summary_fields = st.session_state.merged_data.columns.tolist()
                        st.selectbox("Scenario Summaries Selections", scenario_summary_fields, key="scenario_summary_selections")

                        # st.pills with a key will automatically store the selected values in st.session_state["sub_summary_selections"]
                        st.pills(
                            "Select scenario sub-summaries", 
                            st.session_state.merged_data[st.session_state.scenario_summary_selections].unique(), 
                            selection_mode="multi",
                            key="sub_summary_selections"
                        )



            if "Bid Coverage Summary" in selected_presentations:
                with st.expander("Configure Grouping for Bid Coverage Slides"):
                    st.header("Bid Coverage Grouping")
               
                    # Select group by field
                    bid_coverage_slides_grouping = st.selectbox("Group by", st.session_state.merged_data.columns, key="bid_coverage_slides_grouping")

            if "Supplier Comparison Summary" in selected_presentations:
                with st.expander("Configure Grouping for Supplier Comparison Summary"):
                    st.header("Supplier Comparison Summary")
               
                    # Select group by field
                    supplier_comparison_summary_grouping = st.selectbox("Group by", st.session_state.merged_data.columns, key="supplier_comparison_summary_grouping")

            if st.button("Run Analysis"):
                with st.spinner("Running analysis..."):
                    # 1. Define required analyses for Scenario Summary with correct casing
                    REQUIRED_ANALYSES_FOR_SCENARIO_SUMMARY = [
                        "As-Is",
                        "Best of Best",
                        "Best of Best Excluding Suppliers",
                        "As-Is Excluding Suppliers"
                    ]

                    # 2. Initialize validation flag and message
                    is_valid_selection = True
                    validation_message = ""

                    # 3. Check if "Scenario Summary" is selected
                    if "Scenario Summary" in selected_presentations:
                        # Check if at least one required analysis is selected
                        if not any(analysis in analyses_to_run for analysis in REQUIRED_ANALYSES_FOR_SCENARIO_SUMMARY):
                            is_valid_selection = False
                            validation_message = (
                                "⚠️ **Error:** To generate the 'Scenario Summary' presentation, please select at least one of the following Excel analyses: "
                                + ", ".join(REQUIRED_ANALYSES_FOR_SCENARIO_SUMMARY) + "."
                            )

                    # 4. Display validation messages and stop execution if invalid
                    if not is_valid_selection:
                        st.error(validation_message)
                        st.stop()  # Prevent further execution

                    # 5. Informative message for Scenario Summary dependencies
                    if "Scenario Summary" in selected_presentations:
                        st.info(
                            " The 'Scenario Summary' presentation requires at least one of the following Excel analyses to be selected: "
                            + ", ".join(REQUIRED_ANALYSES_FOR_SCENARIO_SUMMARY) + "."
                        )

                    # 6. Proceed with analysis and presentation generation
                    excel_output = BytesIO()
                    ppt_output = BytesIO()
                    ppt_data = None
                    prs = None

                    # Determine if we need to produce an Excel file at all:
                    # We need Excel if we are running any analyses or if Scenario Summary is required.
                    need_excel = bool(analyses_to_run) or ("Scenario Summary" in selected_presentations)

                    try:
                        if need_excel:
                            # Perform the Excel-based analyses only if required
                            with pd.ExcelWriter(excel_output, engine='openpyxl') as writer:
                                baseline_data = st.session_state.baseline_data
                                original_merged_data = st.session_state.original_merged_data

                                # --- As-Is Analysis ---
                                if "As-Is" in analyses_to_run:
                                    as_is_df = as_is_analysis(st.session_state.merged_data, st.session_state.column_mapping)
                                    as_is_df = add_missing_bid_ids(as_is_df, original_merged_data, st.session_state.column_mapping, 'As-Is')
                                    as_is_df.to_excel(writer, sheet_name='#As-Is', index=False)
                                    if 'Baseline Savings' in as_is_df.columns:
                                        logger.info("[CHECKPOINT] As-Is 'Baseline Savings' sample: %s", as_is_df['Baseline Savings'].head(5).tolist())
                                    else:
                                        logger.warning("As-Is DF has no 'Baseline Savings' column. Columns = %s", as_is_df.columns.tolist())
                                    logger.info("As-Is analysis completed.")

                                # --- Best of Best Analysis ---
                                if "Best of Best" in analyses_to_run:
                                    best_of_best_df = best_of_best_analysis(st.session_state.merged_data, st.session_state.column_mapping)
                                    best_of_best_df = add_missing_bid_ids(best_of_best_df, original_merged_data, st.session_state.column_mapping, 'Best of Best')
                                    best_of_best_df.to_excel(writer, sheet_name='#Best of Best', index=False)
                                    if 'Baseline Savings' in best_of_best_df.columns:
                                         logger.info("[CHECKPOINT] Best-of-Best 'Baseline Savings' sample: %s", best_of_best_df['Baseline Savings'].head(5).tolist())
                                    logger.info("Best of Best analysis completed.")

                                # --- Best of Best Excluding Suppliers Analysis ---
                                if "Best of Best Excluding Suppliers" in analyses_to_run:
                                    # Retrieve exclusion rules from session state, or use an empty list
                                    exclusions_list_bob = st.session_state.exclusions_bob if 'exclusions_bob' in st.session_state else []
                                    
                                    # Ensure column names are stripped of leading/trailing spaces
                                    st.session_state.merged_data.columns = st.session_state.merged_data.columns.str.strip()
                                    
                                    # Call the updated best_of_best_excluding_suppliers function with column_mapping
                                    try:
                                        best_of_best_excl_df = best_of_best_excluding_suppliers(
                                            data=st.session_state.merged_data, 
                                            column_mapping=st.session_state.column_mapping, 
                                            excluded_conditions=exclusions_list_bob
                                        )
                                    except ValueError as ve:
                                        st.error(f"Error in Best of Best Excluding Suppliers Analysis: {ve}")
                                        logger.error(f"Best of Best Excluding Suppliers Analysis failed: {ve}")
                                    else:
                                        # Call add_missing_bid_ids with column_mapping
                                        try:
                                            best_of_best_excl_df = add_missing_bid_ids(
                                                best_of_best_excl_df, 
                                                original_merged_data, 
                                                st.session_state.column_mapping, 
                                                'BOB Excl Suppliers'
                                            )
                                        except Exception as e:
                                            st.error(f"Error in adding missing Bid IDs: {e}")
                                            logger.error(f"Adding Missing Bid IDs failed: {e}")
                                        else:
                                            # Export the result to Excel
                                            try:
                                                best_of_best_excl_df.to_excel(writer, sheet_name='#BOB Excl Suppliers', index=False)
                                                logger.info("Best of Best Excluding Suppliers analysis completed successfully.")
                                                st.success("Best of Best Excluding Suppliers analysis completed and exported to Excel.")
                                            except Exception as e:
                                                st.error(f"Error exporting Best of Best Excluding Suppliers Analysis to Excel: {e}")
                                                logger.error(f"Exporting to Excel failed: {e}")

                                # --- As-Is Excluding Suppliers Analysis ---
                                if "As-Is Excluding Suppliers" in analyses_to_run:
                                    exclusions_list_ais = st.session_state.exclusions_ais if 'exclusions_ais' in st.session_state else []
                                    as_is_excl_df = as_is_excluding_suppliers_analysis(
                                        st.session_state.merged_data, 
                                        st.session_state.column_mapping, 
                                        exclusions_list_ais
                                    )
                                    as_is_excl_df = add_missing_bid_ids(
                                        as_is_excl_df, 
                                        original_merged_data, 
                                        st.session_state.column_mapping, 
                                        'As-Is Excl Suppliers'
                                    )
                                    as_is_excl_df.to_excel(writer, sheet_name='#As-Is Excl Suppliers', index=False)
                                    logger.info("As-Is Excluding Suppliers analysis completed.")

                                # --- Bid Coverage Report Processing ---
                                if "Bid Coverage Report" in analyses_to_run:
                                    variations = st.session_state.bid_coverage_variations if 'bid_coverage_variations' in st.session_state else []
                                    group_by_field = st.session_state.bid_coverage_group_by if 'bid_coverage_group_by' in st.session_state else st.session_state.merged_data.columns[0]
                                    if variations:
                                        bid_coverage_reports = bid_coverage_report(
                                            st.session_state.merged_data, 
                                            st.session_state.column_mapping, 
                                            variations, 
                                            group_by_field
                                        )

                                        # Initialize startrow for Supplier Coverage sheet
                                        supplier_coverage_startrow = 0

                                        for report_name, report_df in bid_coverage_reports.items():
                                            if "Supplier Coverage" in report_name:
                                                sheet_name = "Supplier Coverage"
                                                if sheet_name not in writer.sheets:
                                                    report_df.to_excel(writer, sheet_name=sheet_name, startrow=supplier_coverage_startrow, index=False)
                                                    supplier_coverage_startrow += len(report_df) + 2
                                                else:
                                                    worksheet = writer.sheets[sheet_name]
                                                    worksheet.cell(row=supplier_coverage_startrow + 1, column=1, value=report_name)
                                                    supplier_coverage_startrow += 1
                                                    report_df.to_excel(writer, sheet_name=sheet_name, startrow=supplier_coverage_startrow, index=False)
                                                    supplier_coverage_startrow += len(report_df) + 2
                                                logger.info(f"{report_name} added to sheet '{sheet_name}'.")
                                            else:
                                                sheet_name_clean = report_name.replace(" ", "_")
                                                if len(sheet_name_clean) > 31:
                                                    sheet_name_clean = sheet_name_clean[:31]
                                                report_df.to_excel(writer, sheet_name=sheet_name_clean, index=False)
                                                logger.info(f"{report_name} generated and added to Excel.")
                                    else:
                                        st.warning("No Bid Coverage Report variations selected.")

                                # --- Customizable Analysis Processing ---
                                if "Customizable Analysis" in analyses_to_run:
                                    # Generate the customizable analysis DataFrame
                                    customizable_df = customizable_analysis(st.session_state.merged_data, st.session_state.column_mapping)

                                    # Define raw and mapped grouping column names
                                    grouping_column_raw = st.session_state.grouping_column_raw  # e.g., 'Facility'
                                    grouping_column_mapped = st.session_state.grouping_column_mapped  # e.g., 'Plant'

                                    # If grouping column was not already mapped, add it to customizable_df
                                    if not st.session_state.grouping_column_already_mapped:
                                        if grouping_column_mapped not in customizable_df.columns:
                                            # Add the grouping column next to 'Bid ID'
                                            bid_id_idx = customizable_df.columns.get_loc('Bid ID')
                                            customizable_df.insert(bid_id_idx + 1, grouping_column_mapped, '')

                                    # Write 'Customizable Template' sheet
                                    customizable_df.to_excel(writer, sheet_name='Customizable Template', index=False)

                                    # Prepare 'Customizable Reference' DataFrame
                                    customizable_reference_df = st.session_state.merged_data.copy()

                                    # If grouping column was not already mapped, add it to customizable_reference_df
                                    if not st.session_state.grouping_column_already_mapped:
                                        if grouping_column_mapped not in customizable_reference_df.columns:
                                            # Use 'grouping_column_mapped' instead of 'grouping_column_raw' to access the actual column name
                                            bid_id_to_grouping = st.session_state.merged_data.set_index(
                                                st.session_state.column_mapping['Bid ID']
                                            )[grouping_column_mapped].to_dict()
                                            customizable_reference_df[grouping_column_mapped] = customizable_reference_df[
                                                st.session_state.column_mapping['Bid ID']
                                            ].map(bid_id_to_grouping)

                                    # Write 'Customizable Reference' sheet
                                    customizable_reference_df.to_excel(writer, sheet_name='Customizable Reference', index=False)
                                    logger.info("Customizable Analysis data prepared.")

                                    # ======= Ensure 'Scenario Converter' Sheet Exists =======
                                    if 'Scenario Converter' not in writer.book.sheetnames:
                                        # Create 'Scenario Converter' sheet with 'Bid ID' and empty scenario columns
                                        bid_ids = st.session_state.merged_data[st.session_state.column_mapping['Bid ID']].unique()
                                        scenario_converter_df = pd.DataFrame({'Bid ID': bid_ids})
                                        for i in range(1, 8):
                                            scenario_converter_df[f'Scenario {i}'] = ''  # Initialize empty columns
                                        scenario_converter_df.to_excel(writer, sheet_name='Scenario Converter', index=False)
                                        logger.info("'Scenario Converter' sheet created.")
                                    else:
                                        logger.info("'Scenario Converter' sheet already exists.")

                                    # ======= Ensure Grouping Column Exists in 'merged_data' =======
                                    if not st.session_state.grouping_column_already_mapped:
                                        if grouping_column_mapped not in st.session_state.merged_data.columns:
                                            # Use 'grouping_column_mapped' instead of 'grouping_column_raw' to access the actual column name
                                            st.session_state.merged_data[grouping_column_mapped] = st.session_state.merged_data[
                                                st.session_state.column_mapping['Bid ID']
                                            ].map(bid_id_to_grouping)
                                            logger.info(f"Added grouping column '{grouping_column_mapped}' to 'merged_data'.")

                                    # ======= Access the Workbook and Sheets =======
                                    workbook = writer.book
                                    customizable_template_sheet = workbook['Customizable Template']
                                    customizable_reference_sheet = workbook['Customizable Reference']
                                    scenario_converter_sheet = workbook['Scenario Converter']

                                    # Get the max row numbers
                                    max_row_template = customizable_template_sheet.max_row
                                    max_row_reference = customizable_reference_sheet.max_row
                                    max_row_converter = scenario_converter_sheet.max_row

                                    # Map column names to letters in 'Customizable Reference' sheet
                                    reference_col_letter = {cell.value: get_column_letter(cell.column) for cell in customizable_reference_sheet[1]}
                                    # Map column names to letters in 'Customizable Template' sheet
                                    template_col_letter = {cell.value: get_column_letter(cell.column) for cell in customizable_template_sheet[1]}

                                    # Get the column letters for the grouping column in 'Customizable Reference' sheet
                                    ref_grouping_col = reference_col_letter.get(grouping_column_mapped)
                                    # Get the column letters for the grouping column in 'Customizable Template' sheet
                                    temp_grouping_col = template_col_letter.get(grouping_column_mapped)

                                    # Update 'Supplier Name' formulas in 'Customizable Template' sheet
                                    awarded_supplier_col_letter = template_col_letter['Awarded Supplier']
                                    supplier_name_col_letter = template_col_letter['Supplier Name']

                                    for row in range(2, max_row_template + 1):
                                        awarded_supplier_cell = f"{awarded_supplier_col_letter}{row}"
                                        supplier_name_cell = f"{supplier_name_col_letter}{row}"
                                        # Corrected formula syntax with closing parenthesis
                                        formula_supplier_name = f'=IF({awarded_supplier_cell}<>"", LEFT({awarded_supplier_cell}, FIND("(", {awarded_supplier_cell}) - 2), "")'
                                        customizable_template_sheet[supplier_name_cell].value = formula_supplier_name

                                    # Create supplier lists per Bid ID in a hidden sheet
                                    if 'SupplierLists' not in workbook.sheetnames:
                                        supplier_list_sheet = workbook.create_sheet("SupplierLists")
                                    else:
                                        supplier_list_sheet = workbook['SupplierLists']

                                    bid_id_col_reference = st.session_state.column_mapping['Bid ID']
                                    supplier_name_with_bid_price_col_reference = st.session_state.column_mapping.get('Supplier Name with Bid Price', 'Supplier Name with Bid Price')
                                    bid_id_supplier_list_ranges = {}
                                    current_row = 1

                                    bid_ids = st.session_state.merged_data[bid_id_col_reference].unique()
                                    data = st.session_state.merged_data

                                    for bid_id in bid_ids:
                                        bid_data = data[data[bid_id_col_reference] == bid_id]
                                        bid_data_filtered = bid_data[
                                            (bid_data[st.session_state.column_mapping['Bid Price']].notna()) &
                                            (bid_data[st.session_state.column_mapping['Bid Price']] != 0)
                                        ]
                                        if not bid_data_filtered.empty:
                                            bid_data_sorted = bid_data_filtered.sort_values(by=st.session_state.column_mapping['Bid Price'])
                                            suppliers = bid_data_sorted[supplier_name_with_bid_price_col_reference].dropna().tolist()
                                            start_row = current_row
                                            for supplier in suppliers:
                                                supplier_list_sheet.cell(row=current_row, column=1, value=supplier)
                                                current_row += 1
                                            end_row = current_row - 1
                                            bid_id_supplier_list_ranges[bid_id] = (start_row, end_row)
                                            current_row += 1  # Empty row for separation
                                        else:
                                            bid_id_supplier_list_ranges[bid_id] = None

                                    # Hide the 'SupplierLists' sheet
                                    supplier_list_sheet.sheet_state = 'hidden'

                                    # Set data validation and formulas in 'Customizable Template' sheet
                                    ref_sheet_name = 'Customizable Reference'
                                    temp_sheet_name = 'Customizable Template'

                                    ref_bid_id_col = reference_col_letter[st.session_state.column_mapping['Bid ID']]
                                    ref_supplier_name_col = reference_col_letter[st.session_state.column_mapping['Supplier Name']]
                                    ref_bid_price_col = reference_col_letter[st.session_state.column_mapping['Bid Price']]
                                    ref_supplier_capacity_col = reference_col_letter[st.session_state.column_mapping['Supplier Capacity']]
                                    ref_supplier_name_with_bid_price_col = reference_col_letter.get(supplier_name_with_bid_price_col_reference, 'A')  # Default to 'A' if not found

                                    temp_bid_id_col = template_col_letter['Bid ID']
                                    temp_supplier_name_col = template_col_letter['Supplier Name']
                                    temp_bid_volume_col = template_col_letter['Bid Volume']
                                    temp_baseline_price_col = template_col_letter['Baseline Price']

                                    for row in range(2, max_row_template + 1):
                                        bid_id_cell = f"{temp_bid_id_col}{row}"
                                        awarded_supplier_cell = f"{awarded_supplier_col_letter}{row}"
                                        supplier_name_cell = f"{temp_supplier_name_col}{row}"
                                        bid_volume_cell = f"{temp_bid_volume_col}{row}"
                                        baseline_price_cell = f"{temp_baseline_price_col}{row}"
                                        grouping_cell = f"{temp_grouping_col}{row}" if temp_grouping_col else None

                                        bid_id_value = customizable_template_sheet[bid_id_cell].value

                                        if bid_id_value in bid_id_supplier_list_ranges and bid_id_supplier_list_ranges[bid_id_value]:
                                            start_row, end_row = bid_id_supplier_list_ranges[bid_id_value]
                                            supplier_list_range = f"'SupplierLists'!$A${start_row}:$A${end_row}"

                                            # Set data validation for 'Awarded Supplier'
                                            dv = DataValidation(type="list", formula1=f"{supplier_list_range}", allow_blank=True)
                                            customizable_template_sheet.add_data_validation(dv)
                                            dv.add(customizable_template_sheet[awarded_supplier_cell])

                                            # Construct the formulas using SUMIFS
                                            # Awarded Supplier Price
                                            formula_price = (
                                                f"=IFERROR(SUMIFS('{ref_sheet_name}'!{ref_bid_price_col}:{ref_bid_price_col}, "
                                                f"'{ref_sheet_name}'!{ref_bid_id_col}:{ref_bid_id_col}, {bid_id_cell}, "
                                                f"'{ref_sheet_name}'!{ref_supplier_name_col}:{ref_supplier_name_col}, {supplier_name_cell}), \"\")"
                                            )
                                            awarded_supplier_price_cell = f"{template_col_letter['Awarded Supplier Price']}{row}"
                                            customizable_template_sheet[awarded_supplier_price_cell].value = formula_price

                                            # Awarded Supplier Capacity
                                            formula_supplier_capacity = (
                                                f"=IFERROR(SUMIFS('{ref_sheet_name}'!{ref_supplier_capacity_col}:{ref_supplier_capacity_col}, "
                                                f"'{ref_sheet_name}'!{ref_bid_id_col}:{ref_bid_id_col}, {bid_id_cell}, "
                                                f"'{ref_sheet_name}'!{ref_supplier_name_col}:{ref_supplier_name_col}, {supplier_name_cell}), \"\")"
                                            )
                                            awarded_supplier_capacity_cell = f"{template_col_letter['Awarded Supplier Capacity']}{row}"
                                            customizable_template_sheet[awarded_supplier_capacity_cell].value = formula_supplier_capacity

                                            # Awarded Volume
                                            awarded_volume_cell = f"{template_col_letter['Awarded Volume']}{row}"
                                            formula_awarded_volume = f"=IFERROR(MIN({bid_volume_cell}, {awarded_supplier_capacity_cell}), \"\")"
                                            customizable_template_sheet[awarded_volume_cell].value = formula_awarded_volume

                                            # Awarded Supplier Spend
                                            awarded_supplier_spend_cell = f"{template_col_letter['Awarded Supplier Spend']}{row}"
                                            formula_spend = f"=IF({awarded_supplier_price_cell}<>\"\", {awarded_supplier_price_cell}*{awarded_volume_cell}, \"\")"
                                            customizable_template_sheet[awarded_supplier_spend_cell].value = formula_spend

                                            # Savings
                                            savings_cell = f"{template_col_letter['Savings']}{row}"
                                            formula_savings = f"=IF({awarded_supplier_price_cell}<>\"\", ({baseline_price_cell}-{awarded_supplier_price_cell})*{awarded_volume_cell}, \"\")"
                                            customizable_template_sheet[savings_cell].value = formula_savings

                                            # Baseline Spend
                                            baseline_spend_cell = f"{template_col_letter['Baseline Spend']}{row}"
                                            formula_baseline_spend = f"={baseline_price_cell}*{bid_volume_cell}"
                                            customizable_template_sheet[baseline_spend_cell].value = formula_baseline_spend

                                            # If grouping column exists, populate it
                                            if grouping_cell and ref_grouping_col:
                                                col_index = column_index_from_string(ref_grouping_col) - column_index_from_string(ref_bid_id_col) + 1
                                                formula_grouping = (
                                                    f"=IFERROR(VLOOKUP({bid_id_cell}, '{ref_sheet_name}'!{ref_bid_id_col}:{ref_grouping_col}, "
                                                    f"{col_index}, FALSE), \"\")"
                                                )
                                                customizable_template_sheet[grouping_cell].value = formula_grouping
                                        else:
                                            pass  # No valid suppliers for this Bid ID

                                    # Apply formatting to 'Customizable Reference' sheet
                                    currency_columns_reference = ['Baseline Spend', 'Savings', st.session_state.column_mapping['Bid Price'], st.session_state.column_mapping['Baseline Price']]
                                    number_columns_reference = [st.session_state.column_mapping['Bid Volume'], st.session_state.column_mapping['Supplier Capacity']]

                                    for col_name in currency_columns_reference:
                                        col_letter = reference_col_letter.get(col_name)
                                        if col_letter:
                                            for row_num in range(2, max_row_reference + 1):
                                                cell = customizable_reference_sheet[f"{col_letter}{row_num}"]
                                                cell.number_format = '$#,##0.00'

                                    for col_name in number_columns_reference:
                                        col_letter = reference_col_letter.get(col_name)
                                        if col_letter:
                                            for row_num in range(2, max_row_reference + 1):
                                                cell = customizable_reference_sheet[f"{col_letter}{row_num}"]
                                                cell.number_format = '#,##0'

                                    # Apply formatting to 'Customizable Template' sheet
                                    currency_columns_template = ['Baseline Spend', 'Baseline Price', 'Awarded Supplier Price', 'Awarded Supplier Spend', 'Savings']
                                    number_columns_template = ['Bid Volume', 'Awarded Volume', 'Awarded Supplier Capacity']

                                    for col_name in currency_columns_template:
                                        col_letter = template_col_letter.get(col_name)
                                        if col_letter:
                                            for row_num in range(2, max_row_template + 1):
                                                cell = customizable_template_sheet[f"{col_letter}{row_num}"]
                                                cell.number_format = '$#,##0.00'

                                    for col_name in number_columns_template:
                                        col_letter = template_col_letter.get(col_name)
                                        if col_letter:
                                            for row_num in range(2, max_row_template + 1):
                                                cell = customizable_template_sheet[f"{col_letter}{row_num}"]
                                                cell.number_format = '#,##0'


                                    # ======= Create the Scenario Selector sheet =======
                                    # Continue with existing code for 'Scenario Selector' sheet
                                    scenario_selector_df = customizable_df.copy()
                                    grouping_column = st.session_state.grouping_column_raw 
                                    # Remove 'Supplier Name' column from 'Scenario Selector' sheet
                                    if 'Supplier Name' in scenario_selector_df.columns:
                                        scenario_selector_df.drop(columns=['Supplier Name'], inplace=True)

                                    # Ensure the grouping column is present
                                    if not st.session_state.grouping_column_already_mapped:
                                        if grouping_column not in scenario_selector_df.columns:
                                            # Add the grouping column next to 'Bid ID'
                                            bid_id_idx = scenario_selector_df.columns.get_loc('Bid ID')
                                            scenario_selector_df.insert(bid_id_idx + 1, grouping_column, '')

                                    scenario_selector_df.to_excel(writer, sheet_name='Scenario Selector', index=False)
                                    scenario_selector_sheet = workbook['Scenario Selector']

                                    # Map column names to letters in 'Scenario Selector' sheet
                                    scenario_selector_col_letter = {cell.value: get_column_letter(cell.column) for cell in scenario_selector_sheet[1]}
                                    max_row_selector = scenario_selector_sheet.max_row
                                    scenario_converter_sheet = workbook['Scenario Converter']
                                    max_row_converter = scenario_converter_sheet.max_row

                                    # Prepare formula components for 'Scenario Selector' sheet
                                    scenario_converter_data_range = f"'Scenario Converter'!$B$2:$H${max_row_converter}"
                                    scenario_converter_header_range = "'Scenario Converter'!$B$1:$H$1"
                                    scenario_bid_ids_range = f"'Scenario Converter'!$A$2:$A${max_row_converter}"

                                    bid_id_col_selector = scenario_selector_col_letter['Bid ID']
                                    awarded_supplier_col = scenario_selector_col_letter['Awarded Supplier']
                                    # Get the column letter for the grouping column in 'Scenario Selector' sheet
                                    selector_grouping_col = scenario_selector_col_letter.get(grouping_column)

                                    for row in range(2, max_row_selector + 1):
                                        bid_id_cell = f"{bid_id_col_selector}{row}"
                                        awarded_supplier_cell = f"{awarded_supplier_col}{row}"
                                        bid_volume_cell = f"{scenario_selector_col_letter['Bid Volume']}{row}"
                                        baseline_price_cell = f"{scenario_selector_col_letter['Baseline Price']}{row}"
                                        grouping_cell = f"{selector_grouping_col}{row}" if selector_grouping_col else None

                                        # Formula to pull 'Awarded Supplier' based on selected scenario
                                        formula_awarded_supplier = (
                                            f"=IFERROR(INDEX({scenario_converter_data_range}, MATCH({bid_id_cell}, {scenario_bid_ids_range}, 0), "
                                            f"MATCH('Scenario Reports'!$A$1, {scenario_converter_header_range}, 0)), \"\")"
                                        )
                                        scenario_selector_sheet[awarded_supplier_cell].value = formula_awarded_supplier

                                        # Use 'Awarded Supplier' directly in SUMIFS
                                        awarded_supplier_price_cell = f"{scenario_selector_col_letter['Awarded Supplier Price']}{row}"
                                        awarded_supplier_capacity_cell = f"{scenario_selector_col_letter['Awarded Supplier Capacity']}{row}"

                                        # Awarded Supplier Price
                                        formula_price = (
                                            f"=IFERROR(SUMIFS('{ref_sheet_name}'!{ref_bid_price_col}:{ref_bid_price_col}, "
                                            f"'{ref_sheet_name}'!{ref_bid_id_col}:{ref_bid_id_col}, {bid_id_cell}, "
                                            f"'{ref_sheet_name}'!{ref_supplier_name_col}:{ref_supplier_name_col}, {awarded_supplier_cell}), \"\")"
                                        )
                                        scenario_selector_sheet[awarded_supplier_price_cell].value = formula_price

                                        # Awarded Supplier Capacity
                                        formula_supplier_capacity = (
                                            f"=IFERROR(SUMIFS('{ref_sheet_name}'!{ref_supplier_capacity_col}:{ref_supplier_capacity_col}, "
                                            f"'{ref_sheet_name}'!{ref_bid_id_col}:{ref_bid_id_col}, {bid_id_cell}, "
                                            f"'{ref_sheet_name}'!{ref_supplier_name_col}:{ref_supplier_name_col}, {awarded_supplier_cell}), \"\")"
                                        )
                                        scenario_selector_sheet[awarded_supplier_capacity_cell].value = formula_supplier_capacity

                                        # Awarded Volume
                                        awarded_volume_cell = f"{scenario_selector_col_letter['Awarded Volume']}{row}"
                                        formula_awarded_volume = f"=IF({bid_volume_cell}=\"\", \"\", MIN({bid_volume_cell}, {awarded_supplier_capacity_cell}))"
                                        scenario_selector_sheet[awarded_volume_cell].value = formula_awarded_volume

                                        # Awarded Supplier Spend
                                        awarded_supplier_spend_cell = f"{scenario_selector_col_letter['Awarded Supplier Spend']}{row}"
                                        formula_spend = f"=IF({awarded_supplier_price_cell}<>\"\", {awarded_supplier_price_cell}*{awarded_volume_cell}, \"\")"
                                        scenario_selector_sheet[awarded_supplier_spend_cell].value = formula_spend

                                        # Savings
                                        savings_cell = f"{scenario_selector_col_letter['Savings']}{row}"
                                        formula_savings = f"=IF({awarded_supplier_price_cell}<>\"\", ({baseline_price_cell}-{awarded_supplier_price_cell})*{awarded_volume_cell}, \"\")"
                                        scenario_selector_sheet[savings_cell].value = formula_savings

                                        # Baseline Spend
                                        baseline_spend_cell = f"{scenario_selector_col_letter['Baseline Spend']}{row}"
                                        formula_baseline_spend = f"={baseline_price_cell}*{bid_volume_cell}"
                                        scenario_selector_sheet[baseline_spend_cell].value = formula_baseline_spend

                                        # If grouping column exists, populate it
                                        if grouping_cell and ref_grouping_col:
                                            col_index = column_index_from_string(ref_grouping_col) - column_index_from_string(ref_bid_id_col) + 1
                                            formula_grouping = (
                                                f"=IFERROR(VLOOKUP({bid_id_cell}, '{ref_sheet_name}'!{ref_bid_id_col}:{ref_grouping_col}, "
                                                f"{col_index}, FALSE), \"\")"
                                            )
                                            scenario_selector_sheet[grouping_cell].value = formula_grouping

                                    # Apply formatting to 'Scenario Selector' sheet
                                    currency_columns_selector = ['Baseline Spend', 'Baseline Price', 'Awarded Supplier Price', 'Awarded Supplier Spend', 'Savings']
                                    number_columns_selector = ['Bid Volume', 'Awarded Volume', 'Awarded Supplier Capacity']

                                    for col_name in currency_columns_selector:
                                        col_letter = scenario_selector_col_letter.get(col_name)
                                        if col_letter:
                                            for row_num in range(2, max_row_selector + 1):
                                                cell = scenario_selector_sheet[f"{col_letter}{row_num}"]
                                                cell.number_format = '$#,##0.00'

                                    for col_name in number_columns_selector:
                                        col_letter = scenario_selector_col_letter.get(col_name)
                                        if col_letter:
                                            for row_num in range(2, max_row_selector + 1):
                                                cell = scenario_selector_sheet[f"{col_letter}{row_num}"]
                                                cell.number_format = '#,##0'


                                                # ======= Create the Scenario Reports sheet =======
                                                # Check if 'Scenario Reports' sheet exists; if not, create it
                                                if 'Scenario Reports' not in workbook.sheetnames:
                                                    scenario_reports_sheet = workbook.create_sheet('Scenario Reports')
                                                    logger.info("'Scenario Reports' sheet created.")
                                                else:
                                                    scenario_reports_sheet = workbook['Scenario Reports']
                                                    logger.info("'Scenario Reports' sheet already exists.")

                                                # Define the starting row for group summaries
                                                starting_row = 4  # As per user instruction, first summary starts at A4

                                                # Determine the number of unique suppliers (can be dynamic)
                                                unique_suppliers = st.session_state.merged_data[st.session_state.column_mapping['Supplier Name']].unique()
                                                num_suppliers = len(unique_suppliers)

                                                # Get the grouping column name
                                                grouping_column = st.session_state.grouping_column_raw  # e.g., 'Plant'

                                                # ======= Create a Separate DataFrame for Scenario Reports =======
                                                # **Change:** Use 'customizable_reference_df' as the source instead of 'merged_data'
                                                if grouping_column not in customizable_reference_df.columns:
                                                    st.error(f"Grouping column '{grouping_column}' not found in merged data.")
                                                    logger.error(f"Grouping column '{grouping_column}' not found in merged data.")
                                                else:
                                                    # Create scenario_reports_df with 'Bid ID', grouping column, and 'Supplier Name' from 'Customizable Reference'
                                                    scenario_reports_df = customizable_reference_df[[st.session_state.column_mapping['Bid ID'], grouping_column, st.session_state.column_mapping['Supplier Name']]].copy()

                                                    # Get unique grouping values
                                                    unique_groups = scenario_reports_df[grouping_column].dropna().unique()

                                                    # ======= Ensure 'Scenario Selector' Sheet Exists =======
                                                    if 'Scenario Selector' not in workbook.sheetnames:
                                                        # Create 'Scenario Selector' sheet with headers if it doesn't exist
                                                        scenario_selector_sheet = workbook.create_sheet('Scenario Selector')
                                                        headers = ['Bid ID', grouping_column, st.session_state.column_mapping['Supplier Name'],
                                                                'Awarded Supplier Price', 'Awarded Supplier Capacity',
                                                                'Bid Volume', 'Baseline Price', 'Savings']  # Add other necessary headers
                                                        for col_num, header in enumerate(headers, start=1):
                                                            cell = scenario_selector_sheet.cell(row=1, column=col_num, value=header)
                                                            cell.font = Font(bold=True)
                                                        logger.info("'Scenario Selector' sheet created with headers.")
                                                    else:
                                                        scenario_selector_sheet = workbook['Scenario Selector']
                                                        scenario_selector_headers = [cell.value for cell in scenario_selector_sheet[1]]
                                                        logger.info("'Scenario Selector' sheet already exists.")

                                                    # ======= Detect Presence of Grouping Column in 'Scenario Selector' =======
                                                    # Assuming 'grouping_column_mapped' is the name of the grouping column in 'Scenario Selector'
                                                    if st.session_state.grouping_column_mapped in scenario_selector_headers:
                                                        column_offset = 1  # Shift references by 1 to the right
                                                        logger.info(f"Grouping column '{st.session_state.grouping_column_mapped}' detected in 'Scenario Selector'. Column references will be shifted by {column_offset}.")
                                                    else:
                                                        column_offset = 0  # No shift needed
                                                        logger.info("No grouping column detected in 'Scenario Selector'. Column references remain unchanged.")


                                                    # Helper function to shift column letters based on header row count
                                                    def shift_column(col_letter, scenario_selector_sheet, header_row=1):
                                                        """
                                                        Shifts a column letter by 1 if there are 13 column headers in the specified header row
                                                        of 'Scenario Selector' sheet, does not shift if there are 12 column headers.
                                                        
                                                        Parameters:
                                                        - col_letter (str): The original column letter (e.g., 'B').
                                                        - scenario_selector_sheet (Worksheet): The 'Scenario Selector' worksheet object.
                                                        - header_row (int): The row number where headers are located. Default is 1.
                                                        
                                                        Returns:
                                                        - str: The shifted column letter.
                                                        """
                                                        # Extract the header row
                                                        header_cells = scenario_selector_sheet[header_row]
                                                        
                                                        # Count non-empty header cells
                                                        header_col_count = sum(1 for cell in header_cells if cell.value is not None)
                                                        
                                                        if header_col_count == 13:
                                                            offset = 1
                                                            logger.info(f"Detected {header_col_count} column headers in 'Scenario Selector' sheet. Shifting columns by 1.")
                                                        elif header_col_count == 12:
                                                            offset = 0
                                                            logger.info(f"Detected {header_col_count} column headers in 'Scenario Selector' sheet. No shift applied.")
                                                        else:
                                                            offset = 0  # Default shift
                                                            logger.warning(f"Unexpected number of column headers ({header_col_count}) in 'Scenario Selector' sheet. No shift applied.")
                                                        
                                                        # Convert column letter to index
                                                        try:
                                                            col_index = column_index_from_string(col_letter)
                                                        except ValueError:
                                                            logger.error(f"Invalid column letter provided: {col_letter}")
                                                            raise ValueError(f"Invalid column letter: {col_letter}")
                                                        
                                                        # Calculate new column index
                                                        new_col_index = col_index + offset
                                                        
                                                        # Ensure the new column index is within Excel's limits (1 to 16384)
                                                        if not 1 <= new_col_index <= 16384:
                                                            logger.error(f"Shifted column index {new_col_index} out of Excel's column range.")
                                                            raise ValueError(f"Shifted column index {new_col_index} out of Excel's column range.")
                                                        
                                                        # Convert back to column letter
                                                        new_col_letter = get_column_letter(new_col_index)
                                                        
                                                        logger.debug(f"Column '{col_letter}' shifted by {offset} to '{new_col_letter}'.")
                                                        
                                                        return new_col_letter



                                                    for group in unique_groups:
                                                        # Insert Group Label
                                                        scenario_reports_sheet[f"A{starting_row}"] = group
                                                        scenario_reports_sheet[f"A{starting_row}"].font = Font(bold=True)

                                                        # Insert Header Row
                                                        headers = [
                                                            'Supplier Name',
                                                            'Awarded Volume',
                                                            '% of Business',
                                                            'Baseline Avg',
                                                            'Avg Bid Price',
                                                            '%Δ b/w Baseline and Avg Bid',
                                                            'RFP Savings'
                                                        ]
                                                        for col_num, header in enumerate(headers, start=1):
                                                            cell = scenario_reports_sheet.cell(row=starting_row + 1, column=col_num, value=header)
                                                            cell.font = Font(bold=True)

                                                        # ======= Insert Formula for Supplier Name in the First Supplier Row =======
                                                        supplier_name_cell = f"A{starting_row + 2}"  # Assigning to column A
                                                        # Dynamic reference to the group label cell within the 'Scenario Reports' sheet
                                                        group_label_cell = f"'Scenario Reports'!$A${starting_row}"

                                                        try:
                                                            # Original column references in 'Scenario Selector' sheet
                                                            original_supplier_name_col = 'G'
                                                            original_group_col = 'B'

                                                            # Adjust column references based on column_offset
                                                            adjusted_supplier_name_col = shift_column(original_supplier_name_col, scenario_selector_sheet)
                                                            adjusted_group_col = shift_column(original_group_col, scenario_selector_sheet)

                                                            # Construct the formula with IFERROR to handle potential errors, without the leading '='
                                                            formula_supplier_name = (
                                                                f"IFERROR(UNIQUE(FILTER('Scenario Selector'!{adjusted_supplier_name_col}:{adjusted_supplier_name_col}, "
                                                                f"('Scenario Selector'!{adjusted_supplier_name_col}:{adjusted_supplier_name_col}<>\"\") * ('Scenario Selector'!{original_group_col}:{original_group_col}={group_label_cell}))), \"\")"
                                                            )

                                                            # Optional: Log the formula for debugging purposes
                                                            logger.info(f"Assigning formula to {supplier_name_cell}: {formula_supplier_name}")

                                                            # Assign the formula as text to the specified cell in the 'Scenario Reports' sheet
                                                            scenario_reports_sheet[supplier_name_cell].value = formula_supplier_name
                                                            logger.info(f"Successfully assigned formula as text to {supplier_name_cell}")
                                                        except Exception as e:
                                                            logger.error(f"Failed to assign formula to {supplier_name_cell}: {e}")
                                                            st.error(f"An error occurred while assigning formulas to {supplier_name_cell}: {e}")

                                                        # ======= Insert Formulas for Other Columns in the First Supplier Row =======
                                                        awarded_volume_cell = f"B{starting_row + 2}"
                                                        percent_business_cell = f"C{starting_row + 2}"
                                                        avg_baseline_price_cell = f"D{starting_row + 2}"
                                                        avg_bid_price_cell = f"E{starting_row + 2}"
                                                        percent_delta_cell = f"F{starting_row + 2}"
                                                        rfp_savings_cell = f"G{starting_row + 2}"

                                                        try:
                                                            # Original column references for other formulas
                                                            original_awarded_volume_col = 'I'
                                                            original_bid_volume_col = 'I'  # Assuming 'Bid Volume' is in column 'I'
                                                            original_bid_price_col = 'D'
                                                            original_avg_bid_price_col = 'H'
                                                            original_savings_col = 'L'

                                                            # Adjust column references based on column_offset
                                                            adjusted_awarded_volume_col = shift_column(original_awarded_volume_col, scenario_selector_sheet)
                                                            adjusted_bid_volume_col = shift_column(original_bid_volume_col, scenario_selector_sheet)
                                                            adjusted_bid_price_col = shift_column(original_bid_price_col, scenario_selector_sheet)
                                                            adjusted_avg_bid_price_col = shift_column(original_avg_bid_price_col, scenario_selector_sheet)
                                                            adjusted_savings_col = shift_column(original_savings_col, scenario_selector_sheet)

                                                            # Awarded Volume Formula
                                                            formula_awarded_volume = (
                                                                f"=IF({supplier_name_cell}=\"\", \"\", SUMIFS('Scenario Selector'!${adjusted_awarded_volume_col}:${adjusted_awarded_volume_col}, "
                                                                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}, 'Scenario Selector'!${adjusted_supplier_name_col}:${adjusted_supplier_name_col}, {supplier_name_cell}))"
                                                            )
                                                            scenario_reports_sheet[awarded_volume_cell].value = formula_awarded_volume

                                                            # % of Business Formula
                                                            formula_percent_business = (
                                                                f"=IF({awarded_volume_cell}=0, \"\", {awarded_volume_cell}/SUMIFS('Scenario Selector'!${adjusted_awarded_volume_col}:${adjusted_awarded_volume_col}, "
                                                                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}))"
                                                            )
                                                            scenario_reports_sheet[percent_business_cell].value = formula_percent_business

                                                            # Avg Baseline Price Formula
                                                            formula_avg_baseline_price = (
                                                                f"=IF({supplier_name_cell}=\"\", \"\", AVERAGEIFS('Scenario Selector'!${adjusted_bid_price_col}:{adjusted_bid_price_col}, "
                                                                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}, 'Scenario Selector'!${adjusted_supplier_name_col}:${adjusted_supplier_name_col}, {supplier_name_cell}))"
                                                            )
                                                            scenario_reports_sheet[avg_baseline_price_cell].value = formula_avg_baseline_price

                                                            # Avg Bid Price Formula
                                                            formula_avg_bid_price = (
                                                                f"=IF({supplier_name_cell}=\"\", \"\", AVERAGEIFS('Scenario Selector'!${adjusted_avg_bid_price_col}:{adjusted_avg_bid_price_col}, "
                                                                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}, 'Scenario Selector'!${adjusted_supplier_name_col}:${adjusted_supplier_name_col}, {supplier_name_cell}))"
                                                            )
                                                            scenario_reports_sheet[avg_bid_price_cell].value = formula_avg_bid_price

                                                            # % Delta between Baseline and Avg Bid Formula
                                                            formula_percent_delta = (
                                                                f"=IF(AND({avg_baseline_price_cell}>0, {avg_bid_price_cell}>0), "
                                                                f"({avg_baseline_price_cell}-{avg_bid_price_cell})/{avg_baseline_price_cell}, \"\")"
                                                            )
                                                            scenario_reports_sheet[percent_delta_cell].value = formula_percent_delta

                                                            # RFP Savings Formula
                                                            formula_rfp_savings = (
                                                                f"=IFERROR(IF({supplier_name_cell}=\"\", \"\", SUMIFS('Scenario Selector'!${adjusted_savings_col}:${adjusted_savings_col}, "
                                                                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}, 'Scenario Selector'!${adjusted_supplier_name_col}:${adjusted_supplier_name_col}, {supplier_name_cell})),\"\")"
                                                            )
                                                            scenario_reports_sheet[rfp_savings_cell].value = formula_rfp_savings

                                                            logger.info(f"Successfully assigned formulas to columns B-G in row {starting_row + 2}")
                                                        except Exception as e:
                                                            logger.error(f"Failed to assign formulas to columns B-G in row {starting_row + 2}: {e}")
                                                            st.error(f"An error occurred while assigning formulas to columns B-G in row {starting_row + 2}: {e}")

                                                        # ======= Advance the starting row =======
                                                        # Since UNIQUE spills the results, we'll need to estimate the number of suppliers.
                                                        # For simplicity, we'll add a fixed number of rows per group.
                                                        # Adjust 'max_suppliers_per_group' as needed based on your data.
                                                        max_suppliers_per_group = 10  # Example: Allow up to 10 suppliers per group
                                                        starting_row += 2 + max_suppliers_per_group + 3  # Group label + header + supplier rows + spacing

                                                                                    # ======= Add Drop-Down to 'Scenario Reports' Sheet =======
                                                        # Restore the drop-down menu in cell A1
                                                        try:
                                                            dv_scenario = DataValidation(type="list", formula1="'Scenario Converter'!$B$1:$H$1", allow_blank=True)
                                                            scenario_reports_sheet.add_data_validation(dv_scenario)
                                                            dv_scenario.add(scenario_reports_sheet['A1'])  # Placing drop-down in A1
                                                            logger.info("Scenario Reports sheet created with drop-down in cell A1.")
                                                        except Exception as e:
                                                            st.error(f"Error adding drop-down to Scenario Reports sheet: {e}")
                                                            logger.error(f"Error adding drop-down to Scenario Reports sheet: {e}")

                                                        logger.info(f"Advanced starting row to {starting_row}")

                        if need_excel:
                            excel_output.seek(0)
                            excel_data = excel_output.getvalue()
                            st.session_state.excel_data = excel_data
                        else:
                            # No Excel data generated
                            st.session_state.excel_data = None
                            excel_data = None

                        # If we needed Excel data for Scenario Summary, we read it now:
                        scenario_sheets_loaded = False
                        if "Scenario Summary" in selected_presentations and need_excel:
                            try:
                                scenario_excel_file = pd.ExcelFile(BytesIO(excel_output.getvalue()))
                                scenario_sheet_names = [sheet_name for sheet_name in scenario_excel_file.sheet_names if sheet_name.startswith('#')]
                                scenario_dataframes = {}
                                for sheet_name in scenario_sheet_names:
                                    df = pd.read_excel(scenario_excel_file, sheet_name=sheet_name)
                                            # Log a sample of "Savings" to see if it changed after saving/reading.
                                    if 'Baseline Savings' in df.columns:
                                        logger.info("[CHECKPOINT] After re-reading '%s': 'Baseline Savings' sample: %s",
                                                    sheet_name, df['Baseline Savings'].head(5).tolist())
                                    else:
                                        logger.warning("Sheet '%s' has no 'Baseline Savings' column. Columns = %s", sheet_name, df.columns.tolist())
                                    scenario_dataframes[sheet_name] = df
                                scenario_sheets_loaded = True
                            except Exception as e:
                                st.error(f"Failed to read the generated Excel file for scenario summary: {e}")
                                logger.error(f"Failed to read the generated Excel file for scenario summary: {e}")
                                st.stop()

                        # Generate PowerPoint File if presentations are selected
                        if "Scenario Summary" in selected_presentations:
                            try:
                                # Construct the template file path
                                script_dir = os.path.dirname(os.path.abspath(__file__))
                                template_file_path = os.path.join(script_dir, 'Slide template.pptx')

                                # Read the Excel file into a pandas ExcelFile object
                                excel_file = pd.ExcelFile(BytesIO(excel_data))

                                # Use 'original_merged_data' from your existing variables
                                # Ensure that 'original_merged_data' is available in this scope
                                if 'original_merged_data' in globals() or 'original_merged_data' in locals():
                                    original_df = original_merged_data.copy()
                                else:
                                    st.error("The 'original_merged_data' DataFrame is not available.")
                                    original_df = None  # Set to None to handle the error later

                                # Retrieve the grouping field selected by the user
                                scenario_detail_grouping = st.session_state.get('scenario_detail_grouping', None)
                                scenario_sub_summaries_on = st.session_state.get("scenario_sub_summaries_on", False)
                                scenario_summary_selections = st.session_state.get("scenario_summary_selections", None)
                                sub_summaries_list = st.session_state.get("sub_summary_selections", [])

                                # Get list of sheet names starting with '#'
                                scenario_sheet_names = [sheet_name for sheet_name in excel_file.sheet_names if sheet_name.startswith('#')]

                                scenario_dataframes = {}
                                for sheet_name in scenario_sheet_names:
                                    df = pd.read_excel(excel_file, sheet_name=sheet_name)

                                    # Ensure 'Bid ID' exists in df
                                    if 'Bid ID' not in df.columns:
                                        st.error(f"'Bid ID' is not present in the scenario data for '{sheet_name}'. Skipping this sheet.")
                                        continue  # Skip this scenario

                                    # -----------------------------------------------------------------
                                    # Merge scenario_detail_grouping if required (Option D approach):
                                    # -----------------------------------------------------------------
                                    if scenario_detail_grouping and scenario_detail_grouping not in df.columns:
                                        # Attempt to pull grouping from original_df, but ensuring one row per Bid ID
                                        if original_df is not None and 'Bid ID' in original_df.columns:
                                            if scenario_detail_grouping in original_df.columns:
                                                # 1) Make sure both have str Bid ID
                                                df['Bid ID'] = df['Bid ID'].astype(str)
                                                original_df['Bid ID'] = original_df['Bid ID'].astype(str)
                                                # 2) Build a unique_map from original_df to pick the FIRST non-blank grouping row
                                                unique_map = (
                                                    original_df.copy()
                                                    .assign(
                                                        _blank=lambda x: (
                                                            x[scenario_detail_grouping].isna() 
                                                            | (x[scenario_detail_grouping] == "")
                                                        )
                                                    )
                                                    # Sort so that non-blanks (_blank=False) come first, blanks go last
                                                    .sort_values("_blank")
                                                    # Then group by Bid ID, keep only the first row
                                                    .groupby("Bid ID", as_index=False)
                                                    .first()
                                                )
                                                # 3) Merge that single row per Bid ID
                                                df = df.merge(unique_map[['Bid ID', scenario_detail_grouping]], on='Bid ID', how='left')
                                                if scenario_detail_grouping not in df.columns:
                                                    st.error(f"Failed to merge the grouping field '{scenario_detail_grouping}' into '{sheet_name}'. Skipping this scenario.")
                                                    continue
                                            else:
                                                st.warning(
                                                    f"The selected grouping field '{scenario_detail_grouping}' is not in 'original_merged_data'. "
                                                    "No detail slides will be created for this scenario."
                                                )
                                                # It's okay to proceed without details if user still wants scenario summary slides
                                        else:
                                            st.warning(
                                                "The 'original_merged_data' is not available or 'Bid ID' is missing in 'original_merged_data'. "
                                                f"Cannot merge grouping for '{sheet_name}'."
                                            )
                                            # We can still proceed without scenario details

                                    # Merge scenario_summary_selections if sub-summaries are on and the column not present
                                    if scenario_sub_summaries_on and scenario_summary_selections:
                                        if scenario_summary_selections not in df.columns:
                                            # Attempt to merge from original_df
                                            if original_df is not None and 'Bid ID' in original_df.columns:
                                                if scenario_summary_selections in original_df.columns:
                                                    df['Bid ID'] = df['Bid ID'].astype(str)
                                                    original_df['Bid ID'] = original_df['Bid ID'].astype(str)
                                                    df = df.merge(
                                                        original_df[['Bid ID', scenario_summary_selections]], 
                                                        on='Bid ID', 
                                                        how='left'
                                                    )
                                                    if scenario_summary_selections not in df.columns:
                                                        st.warning(
                                                            f"Failed to merge the sub-summary field '{scenario_summary_selections}' into '{sheet_name}'. "
                                                            "Sub-summaries may not be created."
                                                        )
                                                else:
                                                    st.warning(
                                                        f"The sub-summary field '{scenario_summary_selections}' is not in 'original_merged_data'. "
                                                        f"No sub-summaries for '{sheet_name}'."
                                                    )
                                            else:
                                                st.warning(
                                                    "The 'original_merged_data' is not available or 'Bid ID' missing for merging sub-summary selections."
                                                )

                                    # Store in scenario_dataframes
                                    scenario_dataframes[sheet_name] = df

                                if not scenario_dataframes:
                                    st.error("No valid scenario dataframes were created. Please check your data.")
                                    ppt_data = None  # Ensure ppt_data is set to None if generation fails
                                else:
                                    # Generate the presentation (this now includes sub-summaries if toggled on)

                                    for sn, df in scenario_dataframes.items():
                                        if 'Baseline Savings' in df.columns:
                                            logger.info("[CHECKPOINT] Pre-PPT '%s': 'Baseline Savings' sample: %s",
                                                        sn, df['Baseline Savings'].head(5).tolist())
                                        else:
                                            logger.warning("[CHECKPOINT] Pre-PPT '%s' has no 'Baseline Savings' col. Columns: %s",
                                                        sn, df.columns.tolist())

                                    prs = create_scenario_summary_presentation(scenario_dataframes, template_file_path)

                                    if not prs:
                                        st.error("Failed to generate Scenario Summary presentation.")
                                        ppt_data = None  # Ensure ppt_data is set to None if generation fails
                                    else:
                                        # Save the presentation to BytesIO
                                        prs.save(ppt_output)
                                        ppt_data = ppt_output.getvalue()
                            except Exception as e:
                                st.error(f"An error occurred while generating the presentation: {e}")
                                logger.error(f"Error generating presentation: {e}")
                                ppt_data = None  # Ensure ppt_data is set to None if generation fails
                        else:
                            ppt_data = None  # No presentations selected




                            # --- Supplier Comparison Summary Presentation ---
                        if "Supplier Comparison Summary" in selected_presentations:
                            try:
                                supplier_comparison_summary_grouping = st.session_state.get('supplier_comparison_summary_grouping', None)
                                if not supplier_comparison_summary_grouping:
                                    st.error("Please select a grouping field for the Supplier Comparison Summary.")
                                else:
                                    # Make a copy of merged_data for supplier comparison summary
                                    if 'merged_data' not in st.session_state:
                                        st.error("No merged data available for Supplier Comparison Summary.")
                                    else:
                                        sc_df = st.session_state.merged_data.copy()
                        
                                        # If no existing presentation (prs) has been created yet, create one
                                        if prs is None:
                                            script_dir = os.path.dirname(os.path.abspath(__file__))
                                            template_file_path = os.path.join(script_dir, 'Slide template.pptx')
                                            prs = Presentation(template_file_path)
                        
                                        # Add the Supplier Comparison Summary slide
                                        prs = create_supplier_comparison_summary_slide(prs, sc_df, supplier_comparison_summary_grouping)
                        
                                        # Re-save presentation with new slides
                                        ppt_output = BytesIO()
                                        prs.save(ppt_output)
                                        ppt_output.seek(0)
                                        ppt_data = ppt_output.getvalue()
                        
                            except Exception as e:
                                st.error(f"An error occurred while generating the Supplier Comparison Summary presentation: {e}")
                                logger.error(f"Error generating Supplier Comparison Summary presentation: {e}")

                        # --- Bid Coverage Summary Presentation ---
                        if "Bid Coverage Summary" in selected_presentations:
                            try:
                                bid_coverage_slides_grouping = st.session_state.get('bid_coverage_slides_grouping', None)
                                if not bid_coverage_slides_grouping:
                                    st.error("Please select a grouping field for the Bid Coverage slides.")
                                else:
                                    bc_df = st.session_state.merged_data.copy()
                                    if prs is None:
                                        script_dir = os.path.dirname(os.path.abspath(__file__))
                                        template_file_path = os.path.join(script_dir, 'Slide template.pptx')
                                        prs = Presentation(template_file_path)

                                    prs = create_bid_coverage_summary_slides(prs, bc_df, bid_coverage_slides_grouping)
                                    ppt_output = BytesIO()
                                    prs.save(ppt_output)
                                    ppt_output.seek(0)
                                    ppt_data = ppt_output.getvalue()
                            except Exception as e:
                                st.error(f"An error occurred while generating the Bid Coverage Summary presentation: {e}")
                                logger.error(f"Error generating Bid Coverage Summary presentation: {e}")

                    except Exception as e:
                        st.error(f"An error occurred during analysis: {e}")
                        logger.error(f"Error during analysis: {e}")
                        st.stop()

                    # If we never needed Excel, excel_data is None. Otherwise, we have excel_output.
                    if need_excel:
                        excel_output.seek(0)
                        excel_data = excel_output.getvalue()
                        st.session_state.excel_data = excel_data
                    else:
                        # No Excel data generated
                        st.session_state.excel_data = None

                    st.session_state.ppt_data = ppt_data

                    st.success("Analysis completed successfully. Please download your files below.")

                    # Excel download button (only if excel_data is generated)
                    if st.session_state.excel_data:
                        st.download_button(
                            label="Download Analysis Results (Excel)",
                            data=st.session_state.excel_data,
                            file_name="scenario_analysis_results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        logger.info("Analysis results prepared for download.")
                    else:
                        # If user selected no excel analyses and no scenario summary, no excel is expected
                        logger.info("No Excel analysis results available for download.")
                    
                    # PowerPoint download button (only if data is available)
                    if st.session_state.ppt_data:
                        st.download_button(
                            label="Download Presentation (PowerPoint)",
                            data=st.session_state.ppt_data,
                            file_name="presentation_summary.pptx",
                            mime="application/vnd.openxmlformats-officedocument.presentationml.presentation"
                        )
                        logger.info("Presentation prepared for download.")
                    else:
                        # If no ppt_data is available, either user didn't select presentations
                        # or required conditions for certain presentations were not met.
                        logger.info("No presentation available for download.")





    elif section == 'project_folder':
        if st.session_state.selected_project and st.session_state.selected_subfolder:
            st.title(f"{st.session_state.selected_project} - {st.session_state.selected_subfolder}")
            project_dir = BASE_PROJECTS_DIR / st.session_state.username / st.session_state.selected_project / st.session_state.selected_subfolder

            if not project_dir.exists():
                st.error("Selected folder does not exist.")
                logger.error(f"Folder {project_dir} does not exist.")
            else:
                # Path Navigation Buttons
                st.markdown("---")
                st.subheader("Navigation")
                path_components = [st.session_state.selected_project, st.session_state.selected_subfolder]
                path_keys = ['path_button_project', 'path_button_subfolder']
                path_buttons = st.columns(len(path_components))
                for i, (component, key) in enumerate(zip(path_components, path_keys)):
                    if i < len(path_components) - 1:
                        # Clickable button for higher-level folders
                        if path_buttons[i].button(component, key=key):
                            st.session_state.selected_subfolder = None
                    else:
                        # Disabled button for current folder
                        path_buttons[i].button(component, disabled=True, key=key+'_current')

                # List existing files with download buttons
                st.write(f"Contents of {st.session_state.selected_subfolder}:")
                files = [f.name for f in project_dir.iterdir() if f.is_file()]
                if files:
                    for file in files:
                        file_path = project_dir / file
                        # Provide a download link for each file
                        with open(file_path, "rb") as f:
                            data = f.read()
                        st.download_button(
                            label=f"Download {file}",
                            data=data,
                            file_name=file,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.write("No files found in this folder.")

                st.markdown("---")
                st.subheader("Upload Files to This Folder")

                # File uploader for the selected subfolder
                uploaded_files = st.file_uploader(
                    "Upload Excel Files",
                    type=["xlsx"],
                    accept_multiple_files=True,
                    key='file_uploader_project_folder'
                )

                if uploaded_files:
                    for uploaded_file in uploaded_files:
                        if validate_uploaded_file(uploaded_file):
                            file_path = project_dir / uploaded_file.name
                            if file_path.exists():
                                overwrite = st.checkbox(f"Overwrite existing file '{uploaded_file.name}'?", key=f"overwrite_{uploaded_file.name}")
                                if not overwrite:
                                    st.warning(f"File '{uploaded_file.name}' not uploaded. It already exists.")
                                    logger.warning(f"User chose not to overwrite existing file '{uploaded_file.name}'.")
                                    continue
                            try:
                                with open(file_path, "wb") as f:
                                    f.write(uploaded_file.getbuffer())
                                st.success(f"File '{uploaded_file.name}' uploaded successfully.")
                                logger.info(f"File '{uploaded_file.name}' uploaded to '{project_dir}'.")
                            except Exception as e:
                                st.error(f"Failed to upload file '{uploaded_file.name}': {e}")
                                logger.error(f"Failed to upload file '{uploaded_file.name}': {e}")

        elif st.session_state.selected_project:
            # Display project button as disabled
            st.markdown("---")
            st.subheader("Navigation")
            project_button = st.button(st.session_state.selected_project, disabled=True, key='path_button_project_main')

            # List subfolders
            project_dir = BASE_PROJECTS_DIR / st.session_state.username / st.session_state.selected_project
            subfolders = ["Baseline", "Round 1 Analysis", "Round 2 Analysis", "Supplier Feedback", "Negotiations"]
            st.write("Subfolders:")
            for subfolder in subfolders:
                if st.button(subfolder, key=f"subfolder_{subfolder}"):
                    st.session_state.selected_subfolder = subfolder

        else:
            st.error("No project selected..")
            logger.error("No project selected..")


    elif section == 'upload database':
        st.title('Upload Database Files')
        st.write("Upload your Excel files to Supabase.")
        



    elif section == 'about':
        st.title("About")

        # Step 1: Read the entire markdown file
        with open("docs/report_documentation.md", "r", encoding="utf-8") as f:
            doc_text = f.read()


        # Step 2: Identify a consistent pattern in your markdown headings.
        # For example, if each report section starts with "##" followed by the report name,
        # we can split on that pattern or use a more robust parsing approach.
        #
        # In this example, let's assume your markdown is structured with distinct 
        # second-level headings (##) for each report type, like:
        #
        # ## "As-Is" Report
        # (content...)
        #
        # ## "As-Is + Exclusions" Report
        # (content...)
        #
        # and so forth.
        
        # Split the text by '## ' to separate each section
        sections = doc_text.split('## ')
        # The first split might be text before the first "##", so we can ignore it if empty
        sections = [sec.strip() for sec in sections if sec.strip()]

        # Each element in 'sections' should now start with something like:
        # '"As-Is" Report\n\n**What it does:** ...'
        # We can map each section to a title and content by splitting on the first newline.
        report_dict = {}
        for sec in sections:
            # Split on the first newline to separate the title line from the content
            lines = sec.split('\n', 1)
            title_line = lines[0].strip()
            content = lines[1].strip() if len(lines) > 1 else ""
            report_dict[title_line] = content

        # Step 3: Create tabs. The keys in report_dict should match your reports.
        # Extract just the titles you're interested in, ensuring they match your markdown headings.
        report_titles = [
            'Reports Overview',
            'As-Is',
            'As-Is + Exclusions',
            'Best of Best',
            'Best of Best + Exclusions',
            'Customizable Analysis',
            'Bid Coverage Report'
        ]
        
        st.markdown(
            """
            <style>
            /* Allow the tabs container to scroll horizontally */
            div[data-baseweb="tab-list"] {
            overflow-x: auto;
            white-space: nowrap;
            }
            </style>
            """,
            unsafe_allow_html=True
        )

        tab_objects = st.tabs(report_titles)

        # Step 4: Display each report’s content in the corresponding tab dynamically
        for tab, title in zip(tab_objects, report_titles):
            with tab:
                # Safely access the dictionary (if there's a mismatch, provide a fallback)
                content = report_dict.get(title, "*Content not found.*")
                # Write the content as markdown
                st.markdown("## " + title)
                st.markdown(content)

    else:
        st.title('Home')
        st.write("This section is under construction.")

if __name__ == '__main__':
    main()                                       
