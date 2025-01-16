import pandas as pd
from .config import logger
from io import BytesIO
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font
import numpy as np
import re
# Scenario Analysis

import numpy as np  # Ensure this import is present to handle np.nan
import pandas as pd  # Make sure you have imported pandas as well

def add_missing_bid_ids(best_of_best_excl_df, original_merged_data, column_mapping, sheet_name):
    # """
    # Add missing Bid IDs to the analysis results by creating 'Unallocated' rows.

    # This version uses pd.concat instead of the deprecated DataFrame.append().
    
    # Parameters:
    #     best_of_best_excl_df (pd.DataFrame): The current analysis DataFrame.
    #     original_merged_data (pd.DataFrame): The original merged DataFrame before exclusions.
    #     column_mapping (dict): A mapping of standard column names to actual DataFrame column names.
    #     sheet_name (str): The name of the sheet (for logging or further processing).
    
    # Returns:
    #     pd.DataFrame: Updated analysis DataFrame with missing Bid IDs handled.
    # """

    # Extract mapped column names
    bid_id_col = column_mapping['Bid ID']
    facility_col = column_mapping['Facility']
    incumbent_col = column_mapping['Incumbent']
    baseline_price_col = column_mapping['Baseline Price']
    bid_volume_col = column_mapping['Bid Volume']

    # Check whether a 'Current Price' mapping exists
    has_current_price = (
        'Current Price' in column_mapping
        and column_mapping['Current Price'] is not None
        and column_mapping['Current Price'].lower() != 'none'
    )
    current_price_col = column_mapping.get('Current Price') if has_current_price else None

    # Identify Bid IDs missing in the analysis DataFrame
    all_bid_ids = original_merged_data[bid_id_col].unique()
    processed_ids = best_of_best_excl_df[bid_id_col].unique()
    missing_bids = set(all_bid_ids) - set(processed_ids)

    for bid_id in missing_bids:
        # Pick the first row with that missing Bid ID
        original_row = original_merged_data.loc[original_merged_data[bid_id_col] == bid_id].iloc[0]

        # Recalculate 'Baseline Spend' on the fly
        volume_val = original_row[bid_volume_col]
        base_price_val = original_row[baseline_price_col]
        if pd.notna(volume_val) and pd.notna(base_price_val):
            baseline_spend = volume_val * base_price_val
        else:
            baseline_spend = np.nan

        # Handle Current Price if applicable
        if current_price_col and current_price_col in original_merged_data.columns:
            current_price_val = original_row[current_price_col]
        else:
            current_price_val = np.nan

        # Construct a dictionary for the missing/unallocated row
        row_dict = {
            'Bid ID': bid_id,
            'Bid ID Split': 'A',
            'Facility': original_row[facility_col],
            'Incumbent': original_row[incumbent_col],
            'Baseline Price': base_price_val,
            'Current Price': current_price_val if has_current_price else np.nan,
            'Bid Volume': volume_val,
            'Baseline Spend': baseline_spend,
            'Awarded Supplier': 'Unallocated',
            'Awarded Supplier Price': np.nan,
            'Awarded Volume': np.nan,
            'Awarded Supplier Spend': np.nan,
            'Awarded Supplier Capacity': np.nan,
            'Baseline Savings': np.nan,
            'Current Price Savings': np.nan
        }

        # Convert row_dict into a DataFrame
        row_df = pd.DataFrame([row_dict])
        # Concat the new row onto the existing DataFrame
        best_of_best_excl_df = pd.concat([best_of_best_excl_df, row_df], ignore_index=True)

        logger.debug(
            f"[{sheet_name}] Added unallocated row for missing Bid ID {bid_id}, "
            f"Baseline Spend={baseline_spend}."
        )

    return best_of_best_excl_df

    """
    Add missing Bid IDs to the analysis results by creating 'Unallocated' rows.
 
    This version uses pd.concat instead of the deprecated DataFrame.append().
   
    Parameters:
        best_of_best_excl_df (pd.DataFrame): The current analysis DataFrame.
        original_merged_data (pd.DataFrame): The original merged DataFrame before exclusions.
        column_mapping (dict): A mapping of standard column names to actual DataFrame column names.
        sheet_name (str): The name of the sheet (for logging or further processing).
   
    Returns:
        pd.DataFrame: Updated analysis DataFrame with missing Bid IDs handled.
    """
 
    # Extract mapped column names
    bid_id_col = column_mapping['Bid ID']
    facility_col = column_mapping['Facility']
    incumbent_col = column_mapping['Incumbent']
    baseline_price_col = column_mapping['Baseline Price']
    bid_volume_col = column_mapping['Bid Volume']
 
    # Check whether a 'Current Price' mapping exists
    has_current_price = (
        'Current Price' in column_mapping
        and column_mapping['Current Price'] is not None
        and column_mapping['Current Price'].lower() != 'none'
    )
    current_price_col = column_mapping.get('Current Price') if has_current_price else None
 
    # Identify Bid IDs missing in the analysis DataFrame
    all_bid_ids = original_merged_data[bid_id_col].unique()
    processed_ids = best_of_best_excl_df[bid_id_col].unique()
    missing_bids = set(all_bid_ids) - set(processed_ids)
 
    for bid_id in missing_bids:
        # Pick the first row with that missing Bid ID
        original_row = original_merged_data.loc[original_merged_data[bid_id_col] == bid_id].iloc[0]
 
        # Recalculate 'Baseline Spend' on the fly
        volume_val = original_row[bid_volume_col]
        base_price_val = original_row[baseline_price_col]
        if pd.notna(volume_val) and pd.notna(base_price_val):
            baseline_spend = volume_val * base_price_val
        else:
            baseline_spend = np.nan
 
        # Handle Current Price if applicable
        if current_price_col and current_price_col in original_merged_data.columns:
            current_price_val = original_row[current_price_col]
        else:
            current_price_val = np.nan
 
        # Construct a dictionary for the missing/unallocated row
        row_dict = {
            'Bid ID': bid_id,
            'Bid ID Split': 'A',
            'Facility': original_row[facility_col],
            'Incumbent': original_row[incumbent_col],
            'Baseline Price': base_price_val,
            'Current Price': current_price_val if has_current_price else np.nan,
            'Bid Volume': volume_val,
            'Baseline Spend': baseline_spend,
            'Awarded Supplier': 'Unallocated',
            'Awarded Supplier Price': np.nan,
            'Awarded Volume': np.nan,
            'Awarded Supplier Spend': np.nan,
            'Awarded Supplier Capacity': np.nan,
            'Baseline Savings': np.nan,
            'Current Price Savings': np.nan
        }
 
        # Convert row_dict into a DataFrame
        row_df = pd.DataFrame([row_dict])
        # Concat the new row onto the existing DataFrame
        best_of_best_excl_df = pd.concat([best_of_best_excl_df, row_df], ignore_index=True)
 
        logger.debug(
            f"[{sheet_name}] Added unallocated row for missing Bid ID {bid_id}, "
            f"Baseline Spend={baseline_spend}."
        )
 
    return best_of_best_excl_df
 
    """
    Add missing Bid IDs to the analysis results by creating 'Unallocated' rows.
 
    This version uses pd.concat instead of the deprecated DataFrame.append().
   
    Parameters:
        best_of_best_excl_df (pd.DataFrame): The current analysis DataFrame.
        original_merged_data (pd.DataFrame): The original merged DataFrame before exclusions.
        column_mapping (dict): A mapping of standard column names to actual DataFrame column names.
        sheet_name (str): The name of the sheet (for logging or further processing).
   
    Returns:
        pd.DataFrame: Updated analysis DataFrame with missing Bid IDs handled.
    """
 
    # Extract mapped column names
    bid_id_col = column_mapping['Bid ID']
    facility_col = column_mapping['Facility']
    incumbent_col = column_mapping['Incumbent']
    baseline_price_col = column_mapping['Baseline Price']
    bid_volume_col = column_mapping['Bid Volume']
 
    # Check whether a 'Current Price' mapping exists
    has_current_price = (
        'Current Price' in column_mapping
        and column_mapping['Current Price'] is not None
        and column_mapping['Current Price'].lower() != 'none'
    )
    current_price_col = column_mapping.get('Current Price') if has_current_price else None
 
    # Identify Bid IDs missing in the analysis DataFrame
    all_bid_ids = original_merged_data[bid_id_col].unique()
    processed_ids = best_of_best_excl_df[bid_id_col].unique()
    missing_bids = set(all_bid_ids) - set(processed_ids)
 
    for bid_id in missing_bids:
        # Pick the first row with that missing Bid ID
        original_row = original_merged_data.loc[original_merged_data[bid_id_col] == bid_id].iloc[0]
 
        # Recalculate 'Baseline Spend' on the fly
        volume_val = original_row[bid_volume_col]
        base_price_val = original_row[baseline_price_col]
        if pd.notna(volume_val) and pd.notna(base_price_val):
            baseline_spend = volume_val * base_price_val
        else:
            baseline_spend = np.nan
 
        # Handle Current Price if applicable
        if current_price_col and current_price_col in original_merged_data.columns:
            current_price_val = original_row[current_price_col]
        else:
            current_price_val = np.nan
 
        # Construct a dictionary for the missing/unallocated row
        row_dict = {
            'Bid ID': bid_id,
            'Bid ID Split': 'A',
            'Facility': original_row[facility_col],
            'Incumbent': original_row[incumbent_col],
            'Baseline Price': base_price_val,
            'Current Price': current_price_val if has_current_price else np.nan,
            'Bid Volume': volume_val,
            'Baseline Spend': baseline_spend,
            'Awarded Supplier': 'Unallocated',
            'Awarded Supplier Price': np.nan,
            'Awarded Volume': np.nan,
            'Awarded Supplier Spend': np.nan,
            'Awarded Supplier Capacity': np.nan,
            'Baseline Savings': np.nan,
            'Current Price Savings': np.nan
        }
        best_of_best_excl_df = best_of_best_excl_df.append(row_dict, ignore_index=True)
        logger.debug(f"Added unallocated Bid ID {bid_id} to the analysis.")
   
    return best_of_best_excl_df
 
    """Add missing bid IDs to the analysis output with baseline info and 'Unallocated'."""
    # Extract required column names from the mapping
    bid_id_col = column_mapping['Bid ID']
    bid_volume_col = column_mapping['Bid Volume']
    baseline_price_col = column_mapping['Baseline Price']
    facility_col = column_mapping['Facility']
    incumbent_col = column_mapping['Incumbent']
 
    # Identify missing bid IDs in the analysis data
    missing_bid_ids = original_df[~original_df[bid_id_col].isin(analysis_df[bid_id_col])]
 
    # Ensure we only have one row per missing Bid ID
    missing_bid_ids = missing_bid_ids.drop_duplicates(subset=[bid_id_col])
 
    # Fill missing bid IDs with baseline data and 'Unallocated' in the award sections
    if not missing_bid_ids.empty:
        missing_rows = []
        for _, row in missing_bid_ids.iterrows():
            bid_id = row[bid_id_col]
            bid_volume = row[bid_volume_col]
            baseline_price = row[baseline_price_col]
            baseline_spend = bid_volume * baseline_price
            facility = row[facility_col]
            incumbent = row[incumbent_col]
 
            missing_row = {
                'Bid ID': bid_id,
                'Bid ID Split': 'A',
                'Facility': facility,
                'Incumbent': incumbent,
                'Baseline Price': baseline_price,
                'Bid Volume': bid_volume,
                'Baseline Spend': baseline_spend,
                'Awarded Supplier': 'Unallocated',
                'Awarded Supplier Price': None,
                'Awarded Volume': None,
                'Awarded Supplier Spend': None,
                'Awarded Supplier Capacity': None,
                'Savings': None
            }
            missing_rows.append(missing_row)
            logger.debug(f"Added missing Bid ID {bid_id} back into analysis.")
 
        missing_df = pd.DataFrame(missing_rows)
 
        # Concatenate missing_df to analysis_df
        analysis_df = pd.concat([analysis_df, missing_df], ignore_index=True)
 
    return analysis_df
 
def as_is_analysis(data, column_mapping):
    """Perform 'As-Is' analysis with normalized fields, including Current Price Savings."""
    logger.info("Starting As-Is analysis.")
    bid_price_col = column_mapping['Bid Price']
    bid_volume_col = column_mapping['Bid Volume']
    baseline_price_col = column_mapping['Baseline Price']
    supplier_capacity_col = column_mapping['Supplier Capacity']
    bid_id_col = column_mapping['Bid ID']
    incumbent_col = column_mapping['Incumbent']
    supplier_name_col = column_mapping['Supplier Name']
    facility_col = column_mapping['Facility']
 
    # Check if 'Current Price' is mapped and not 'None'
    has_current_price = 'Current Price' in column_mapping and column_mapping['Current Price'] != 'None'
    if has_current_price:
        current_price_col = column_mapping['Current Price']
        data['Current Price'] = data[current_price_col]
 
    data[supplier_name_col] = data[supplier_name_col].str.title()
    data[incumbent_col] = data[incumbent_col].str.title()
    data['Baseline Spend'] = data[bid_volume_col] * data[baseline_price_col]
 
    # Treat bids with Bid Price NaN or 0 as 'No Bid'
    data['Valid Bid'] = data[bid_price_col].notna() & (data[bid_price_col] != 0)
 
    as_is_list = []
    bid_ids = data[bid_id_col].unique()
    for bid_id in bid_ids:
        bid_rows = data[(data[bid_id_col] == bid_id) & data['Valid Bid']]
        incumbent = data.loc[data[bid_id_col] == bid_id, incumbent_col].iloc[0]
        incumbent_bid = bid_rows[bid_rows[supplier_name_col] == incumbent]
 
        if incumbent_bid.empty:
            bid_row = data[data[bid_id_col] == bid_id].iloc[0]
            row_dict = {
                'Bid ID': bid_id,
                'Bid ID Split': 'A',
                'Facility': bid_row[facility_col],
                'Incumbent': incumbent,
                'Baseline Price': bid_row[baseline_price_col],
                'Current Price': None if has_current_price else bid_row[baseline_price_col],  # Optional
                'Bid Volume': bid_row[bid_volume_col],
                'Baseline Spend': bid_row['Baseline Spend'],
                'Awarded Supplier': 'No Bid from Incumbent',
                'Awarded Supplier Price': None,
                'Awarded Volume': None,
                'Awarded Supplier Spend': None,
                'Awarded Supplier Capacity': None,
                'Baseline Savings': None,
                'Current Price Savings': None if has_current_price else None
            }
            as_is_list.append(row_dict)
            logger.debug(f"No valid bid from incumbent for Bid ID {bid_id}.")
            continue
 
        remaining_volume = incumbent_bid.iloc[0][bid_volume_col]
        split_index = 'A'
        for i, row in incumbent_bid.iterrows():
            supplier_capacity = row[supplier_capacity_col] if pd.notna(row[supplier_capacity_col]) else remaining_volume
            awarded_volume = min(remaining_volume, supplier_capacity)
            baseline_volume = awarded_volume
            baseline_spend = baseline_volume * row[baseline_price_col]
            as_is_spend = awarded_volume * row[bid_price_col]
            baseleline_as_is_savings = baseline_spend - as_is_spend
 
 
            row_dict = {
                'Bid ID': row[bid_id_col],
                'Bid ID Split': split_index,
                'Facility': row[facility_col],
                'Incumbent': incumbent,
                'Baseline Price': row[baseline_price_col],
                'Bid Volume': baseline_volume,
                'Baseline Spend': baseline_spend,
                'Awarded Supplier': row[supplier_name_col],
                'Awarded Supplier Price': row[bid_price_col],
                'Awarded Volume': awarded_volume,
                'Awarded Supplier Spend': as_is_spend,
                'Awarded Supplier Capacity': supplier_capacity,
                'Baseline Savings': baseleline_as_is_savings
            }
 
            if has_current_price:
                current_price = data.loc[data[bid_id_col] == bid_id, 'Current Price'].iloc[0]
                row_dict['Current Price'] = current_price
                if row_dict['Awarded Supplier Price'] is not None and row_dict['Bid Volume'] is not None:
                    row_dict['Current Price Savings'] = (current_price - row_dict['Awarded Supplier Price']) * row_dict['Bid Volume']
                else:
                    row_dict['Current Price Savings'] = None
 
            as_is_list.append(row_dict)
            logger.debug(f"As-Is analysis for Bid ID {bid_id}, Split {split_index}: Awarded Volume = {awarded_volume}")
            remaining_volume -= awarded_volume
            if remaining_volume > 0:
                split_index = chr(ord(split_index) + 1)
            else:
                break
 
    as_is_df = pd.DataFrame(as_is_list)
 
    # Define desired column order
    desired_columns = [
        'Bid ID', 'Bid ID Split', 'Facility', 'Incumbent',
        'Baseline Price'
    ]
 
    if has_current_price:
        desired_columns.append('Current Price')
 
    desired_columns.extend([
        'Bid Volume', 'Baseline Spend',
        'Awarded Supplier', 'Awarded Supplier Price',
        'Awarded Volume', 'Awarded Supplier Spend',
        'Awarded Supplier Capacity', 'Baseline Savings'
    ])
 
    if has_current_price:
        desired_columns.append('Current Price Savings')
 
    # Reorder columns
    as_is_df = as_is_df.reindex(columns=desired_columns)
 
    return as_is_df
 
def best_of_best_analysis(data, column_mapping):
    """Perform 'Best of Best' analysis with normalized fields, including Current Price Savings."""
    logger.info("Starting Best of Best analysis.")
    bid_price_col = column_mapping['Bid Price']
    bid_volume_col = column_mapping['Bid Volume']
    baseline_price_col = column_mapping['Baseline Price']
    supplier_capacity_col = column_mapping['Supplier Capacity']
    bid_id_col = column_mapping['Bid ID']
    facility_col = column_mapping['Facility']
    incumbent_col = column_mapping['Incumbent']
    supplier_name_col = column_mapping['Supplier Name']
 
    # Check if 'Current Price' is mapped and not 'None'
    has_current_price = 'Current Price' in column_mapping and column_mapping['Current Price'] != 'None'
    if has_current_price:
        current_price_col = column_mapping['Current Price']
        data['Current Price'] = data[current_price_col]
 
    # Treat bids with Bid Price NaN or 0 as 'No Bid'
    data['Valid Bid'] = data[bid_price_col].notna() & (data[bid_price_col] != 0)
 
    bid_data = data.loc[data['Valid Bid']]
    bid_data = bid_data.sort_values([bid_id_col, bid_price_col])
    data['Baseline Spend'] = data[bid_volume_col] * data[baseline_price_col]
    best_of_best_list = []
    bid_ids = data[bid_id_col].unique()
    for bid_id in bid_ids:
        bid_rows = bid_data[bid_data[bid_id_col] == bid_id]
        if bid_rows.empty:
            bid_row = data[data[bid_id_col] == bid_id].iloc[0]
            row_dict = {
                'Bid ID': bid_id,
                'Bid ID Split': 'A',
                'Facility': bid_row[facility_col],
                'Incumbent': bid_row[incumbent_col],
                'Baseline Price': bid_row[baseline_price_col],
                'Current Price': None if not has_current_price else bid_row[current_price_col],
                'Bid Volume': bid_row[bid_volume_col],
                'Baseline Spend': bid_row['Baseline Spend'],
                'Awarded Supplier': 'No Bids',
                'Awarded Supplier Price': None,
                'Awarded Volume': None,
                'Awarded Supplier Spend': None,
                'Awarded Supplier Capacity': None,
                'Baseline Savings': None,
                'Current Price Savings': None if not has_current_price else None
            }
            best_of_best_list.append(row_dict)
            logger.debug(f"No valid bids for Bid ID {bid_id}.")
            continue
 
        remaining_volume = bid_rows.iloc[0][bid_volume_col]
        split_index = 'A'
        for i, row in bid_rows.iterrows():
            supplier_capacity = row[supplier_capacity_col] if pd.notna(row[supplier_capacity_col]) else remaining_volume
            awarded_volume = min(remaining_volume, supplier_capacity)
            baseline_volume = awarded_volume
            baseline_spend = baseline_volume * row[baseline_price_col]
            best_of_best_spend = awarded_volume * row[bid_price_col]
            baseline_savings = baseline_spend - best_of_best_spend
 
            row_dict = {
                'Bid ID': row[bid_id_col],
                'Bid ID Split': split_index,
                'Facility': row[facility_col],
                'Incumbent': row[incumbent_col],
                'Baseline Price': row[baseline_price_col],
                'Bid Volume': baseline_volume,
                'Baseline Spend': baseline_spend,
                'Awarded Supplier': row[supplier_name_col],
                'Awarded Supplier Price': row[bid_price_col],
                'Awarded Volume': awarded_volume,
                'Awarded Supplier Spend': best_of_best_spend,
                'Awarded Supplier Capacity': supplier_capacity,
                'Baseline Savings': baseline_savings
            }
 
            if has_current_price:
                current_price = data.loc[data[bid_id_col] == bid_id, 'Current Price'].iloc[0]
                row_dict['Current Price'] = current_price
                if row_dict['Awarded Supplier Price'] is not None and row_dict['Bid Volume'] is not None:
                    row_dict['Current Price Savings'] = (current_price - row_dict['Awarded Supplier Price']) * row_dict['Bid Volume']
                else:
                    row_dict['Current Price Savings'] = None
 
            best_of_best_list.append(row_dict)
            logger.debug(f"Best of Best analysis for Bid ID {bid_id}, Split {split_index}: Awarded Volume = {awarded_volume}")
            remaining_volume -= awarded_volume
            if remaining_volume > 0:
                split_index = chr(ord(split_index) + 1)
            else:
                break
 
    best_of_best_df = pd.DataFrame(best_of_best_list)
 
    # Define desired column order
    desired_columns = [
        'Bid ID', 'Bid ID Split', 'Facility', 'Incumbent',
        'Baseline Price'
    ]
 
    if has_current_price:
        desired_columns.append('Current Price')
 
    desired_columns.extend([
        'Bid Volume', 'Baseline Spend',
        'Awarded Supplier', 'Awarded Supplier Price',
        'Awarded Volume', 'Awarded Supplier Spend',
        'Awarded Supplier Capacity', 'Baseline Savings'
    ])
 
    if has_current_price:
        desired_columns.append('Current Price Savings')
 
    # Reorder columns
    best_of_best_df = best_of_best_df.reindex(columns=desired_columns)
 
    return best_of_best_df
 
def best_of_best_excluding_suppliers(data, column_mapping, excluded_conditions):
    """
    Perform 'Best of Best Excluding Suppliers' analysis, including Current Price and Savings.
 
    Parameters:
        data (pd.DataFrame): The input data containing bid information.
        column_mapping (dict): A mapping of standard column names to actual DataFrame column names.
        excluded_conditions (list of tuples): Conditions to exclude certain suppliers or bids.
 
    Returns:
        pd.DataFrame: A DataFrame containing the analysis results.
 
    Raises:
        ValueError: If required columns are missing in column_mapping or data.
    """
    logger.info("Starting Best of Best Excluding Suppliers analysis.")
 
    # Define required columns
    required_columns = [
        'Bid Price', 'Bid Volume', 'Baseline Price', 'Supplier Capacity',
        'Bid ID', 'Incumbent', 'Supplier Name', 'Facility'
    ]
    has_current_price = 'Current Price' in column_mapping and column_mapping['Current Price'] != 'None'
    if has_current_price:
        required_columns.append('Current Price')
 
    # Validate column_mapping
    missing_columns = [col for col in required_columns if col not in column_mapping]
    if missing_columns:
        logger.error(f"Missing columns in column_mapping: {missing_columns}")
        raise ValueError(f"Missing columns in column_mapping: {missing_columns}")
 
    # Validate data columns
    data_columns = [column_mapping[col] for col in required_columns]
    missing_data_columns = [col for col in data_columns if col not in data.columns]
    if missing_data_columns:
        logger.error(f"Missing columns in data: {missing_data_columns}")
        raise ValueError(f"Missing columns in data: {missing_data_columns}")
 
    # Extract column names from column_mapping
    bid_price_col = column_mapping['Bid Price']
    bid_volume_col = column_mapping['Bid Volume']
    baseline_price_col = column_mapping['Baseline Price']
    supplier_capacity_col = column_mapping['Supplier Capacity']
    bid_id_col = column_mapping['Bid ID']
    incumbent_col = column_mapping['Incumbent']
    supplier_name_col = column_mapping['Supplier Name']
    facility_col = column_mapping['Facility']
    current_price_col = column_mapping.get('Current Price', None)
 
    # Convert 'Bid Price' to numeric by removing '$' and handling non-numeric entries
    data[bid_price_col] = data[bid_price_col].astype(str).str.replace('$', '').str.strip()
    data[bid_price_col] = pd.to_numeric(data[bid_price_col], errors='coerce')
 
    # Convert 'Current Price' to numeric if applicable
    if has_current_price:
        data['Current Price'] = data[current_price_col].astype(str).str.replace('$', '').str.strip()
        data['Current Price'] = pd.to_numeric(data['Current Price'], errors='coerce')
 
    # Standardize supplier and incumbent names and strip leading/trailing spaces
    data[supplier_name_col] = data[supplier_name_col].astype(str).str.title().str.strip()
    data[incumbent_col] = data[incumbent_col].astype(str).str.title().str.strip()
    data[facility_col] = data[facility_col].astype(str).str.strip()
 
    # Calculate Baseline Spend
    data['Baseline Spend'] = data[bid_volume_col] * data[baseline_price_col]
 
    # Determine Valid Bids
    data['Valid Bid'] = data[bid_price_col].notna() & (data[bid_price_col] != 0)
 
    # Apply Exclusion Rules (Case-Insensitive)
    for condition in excluded_conditions:
        if len(condition) != 5:
            logger.error(f"Invalid condition format: {condition}")
            raise ValueError(f"Each excluded condition must have 5 elements, got: {condition}")
       
        supplier, field, logic, value, exclude_all = condition
 
        # Standardize supplier name for case-insensitive comparison
        supplier_standard = supplier.strip().title()
 
        if exclude_all:
            # Exclude all bids from the supplier (case-insensitive)
            data = data[data[supplier_name_col].str.lower() != supplier_standard.lower()]
            logger.debug(f"Excluding all bids from supplier '{supplier_standard}'.")
        else:
            # Ensure that the field exists in the data
            if field not in data.columns:
                logger.warning(f"Field '{field}' does not exist in data. Skipping this condition.")
                continue
 
            # If the field is of string type, standardize it for consistent comparison
            if pd.api.types.is_string_dtype(data[field]):
                data[field] = data[field].astype(str).str.title().str.strip()
 
            if logic == "Equal to":
                # Exclude bids where supplier matches and field equals the value (case-insensitive for supplier)
                condition_mask = (data[supplier_name_col].str.lower() == supplier_standard.lower()) & (data[field] == value)
                data = data[~condition_mask]
                logger.debug(f"Excluding bids from supplier '{supplier_standard}' where '{field}' == '{value}'.")
            elif logic == "Not equal to":
                # Exclude bids where supplier matches and field does not equal the value
                condition_mask = (data[supplier_name_col].str.lower() == supplier_standard.lower()) & (data[field] != value)
                data = data[~condition_mask]
                logger.debug(f"Excluding bids from supplier '{supplier_standard}' where '{field}' != '{value}'.")
            else:
                logger.warning(f"Unknown logic '{logic}' in condition: {condition}")
 
    # Filter valid bids after exclusions
    bid_data = data.loc[data['Valid Bid']].copy()
    bid_data = bid_data.sort_values([bid_id_col, bid_price_col])
    best_of_best_excl_list = []
    bid_ids = bid_data[bid_id_col].unique()
 
    for bid_id in bid_ids:
        bid_rows = bid_data[bid_data[bid_id_col] == bid_id]
        if bid_rows.empty:
            # Handle unallocated bids
            original_bid = data[data[bid_id_col] == bid_id].iloc[0]
            row_dict = {
                'Bid ID': bid_id,
                'Bid ID Split': 'A',
                'Facility': original_bid[facility_col],
                'Incumbent': original_bid[incumbent_col],
                'Baseline Price': original_bid[baseline_price_col],
                'Current Price': np.nan if not has_current_price else original_bid['Current Price'],
                'Bid Volume': original_bid[bid_volume_col],
                'Baseline Spend': original_bid['Baseline Spend'],
                'Awarded Supplier': 'Unallocated',
                'Awarded Supplier Price': np.nan,
                'Awarded Volume': np.nan,
                'Awarded Supplier Spend': np.nan,
                'Awarded Supplier Capacity': np.nan,
                'Baseline Savings': np.nan,
                'Current Price Savings': np.nan if not has_current_price else np.nan
            }
            best_of_best_excl_list.append(row_dict)
            logger.debug(f"No valid bids for Bid ID {bid_id}. Marked as Unallocated.")
            continue
 
        remaining_volume = bid_rows.iloc[0][bid_volume_col]
        split_index = 'A'
 
        for _, row in bid_rows.iterrows():
            supplier_capacity = row[supplier_capacity_col] if pd.notna(row[supplier_capacity_col]) else remaining_volume
            awarded_volume = min(remaining_volume, supplier_capacity)
            baseline_spend = awarded_volume * row[baseline_price_col]
            awarded_spend = awarded_volume * row[bid_price_col]
            baseline_savings = baseline_spend - awarded_spend
 
            # Calculate Current Price Savings if applicable
            if has_current_price:
                current_price = row['Current Price']
                current_price_savings = (current_price - row[bid_price_col]) * awarded_volume if pd.notna(current_price) else np.nan
            else:
                current_price_savings = np.nan
 
            row_dict = {
                'Bid ID': row[bid_id_col],
                'Bid ID Split': split_index,
                'Facility': row[facility_col],
                'Incumbent': row[incumbent_col],
                'Baseline Price': row[baseline_price_col],
                'Current Price': row['Current Price'] if has_current_price else np.nan,
                'Bid Volume': awarded_volume,
                'Baseline Spend': baseline_spend,
                'Awarded Supplier': row[supplier_name_col],
                'Awarded Supplier Price': row[bid_price_col],
                'Awarded Volume': awarded_volume,
                'Awarded Supplier Spend': awarded_spend,
                'Awarded Supplier Capacity': supplier_capacity,
                'Baseline Savings': baseline_savings,
                'Current Price Savings': current_price_savings
            }
 
            best_of_best_excl_list.append(row_dict)
            logger.debug(
                f"Processed Bid ID {bid_id}, Split {split_index}: "
                f"Awarded Volume = {awarded_volume}"
            )
            remaining_volume -= awarded_volume
            if remaining_volume <= 0:
                break
            split_index = chr(ord(split_index) + 1)
 
    # Create DataFrame from the list of dictionaries
    best_of_best_excl_df = pd.DataFrame(best_of_best_excl_list)
 
    # Define desired column order
    desired_columns = [
        'Bid ID', 'Bid ID Split', 'Facility', 'Incumbent',
        'Baseline Price'
    ]
 
    if has_current_price:
        desired_columns.append('Current Price')
 
    desired_columns.extend([
        'Bid Volume', 'Baseline Spend',
        'Awarded Supplier', 'Awarded Supplier Price',
        'Awarded Volume', 'Awarded Supplier Spend',
        'Awarded Supplier Capacity', 'Baseline Savings'
    ])
 
    if has_current_price:
        desired_columns.append('Current Price Savings')
 
    # Reorder columns
    best_of_best_excl_df = best_of_best_excl_df.reindex(columns=desired_columns)
 
    logger.info("Completed Best of Best Excluding Suppliers analysis.")
    return best_of_best_excl_df
 
def as_is_excluding_suppliers_analysis(data, column_mapping, excluded_conditions):
    """Perform 'As-Is Excluding Suppliers' analysis with exclusion rules, including Current Price Savings."""
    logger.info("Starting As-Is Excluding Suppliers analysis.")
   
    # Strip leading/trailing spaces from all column names
    data.columns = data.columns.str.strip()
   
    # Column mappings
    bid_price_col = column_mapping['Bid Price']
    bid_volume_col = column_mapping['Bid Volume']
    supplier_capacity_col = column_mapping['Supplier Capacity']
    bid_id_col = column_mapping['Bid ID']
    incumbent_col = column_mapping['Incumbent']
    supplier_name_col = column_mapping['Supplier Name']
    facility_col = column_mapping['Facility']
    baseline_price_col = column_mapping['Baseline Price']
   
    # Check if 'Current Price' is mapped and not 'None'
    has_current_price = 'Current Price' in column_mapping and column_mapping['Current Price'] != 'None'
    if has_current_price:
        current_price_col = column_mapping['Current Price']
        data['Current Price'] = data[current_price_col]
   
    # Standardize supplier and incumbent names
    data[supplier_name_col] = data[supplier_name_col].astype(str).str.title().str.strip()
    data[incumbent_col] = data[incumbent_col].astype(str).str.title().str.strip()
   
    # Treat bids with Bid Price NaN or 0 as 'No Bid'
    data['Valid Bid'] = data[bid_price_col].notna() & (data[bid_price_col] != 0)
   
    # Apply exclusion rules specific to this analysis
    for condition in excluded_conditions:
        supplier, field, logic, value, exclude_all = condition
        supplier = supplier.strip().title()
        if isinstance(value, str):
            value = value.strip().title()
       
        if exclude_all:
            data = data[data[supplier_name_col].str.lower() != supplier.lower()]
            logger.debug(f"Excluding all bids from supplier {supplier} in As-Is Excluding Suppliers analysis.")
        else:
            if field not in data.columns:
                logger.warning(f"Field '{field}' does not exist in data. Skipping this condition.")
                continue
            if logic == "Equal to":
                if pd.api.types.is_string_dtype(data[field]):
                    data = data[~((data[supplier_name_col].str.lower() == supplier.lower()) &
                                 (data[field].astype(str).str.lower().str.strip() == value.lower()))]
                else:
                    data = data[~((data[supplier_name_col].str.lower() == supplier.lower()) &
                                 (data[field] == value))]
                logger.debug(f"Excluding bids from supplier {supplier} where {field} == {value}.")
            elif logic == "Not equal to":
                if pd.api.types.is_string_dtype(data[field]):
                    data = data[~((data[supplier_name_col].str.lower() == supplier.lower()) &
                                 (data[field].astype(str).str.lower().str.strip() != value.lower()))]
                else:
                    data = data[~((data[supplier_name_col].str.lower() == supplier.lower()) &
                                 (data[field] != value))]
                logger.debug(f"Excluding bids from supplier {supplier} where {field} != {value}.")
            else:
                logger.warning(f"Unknown logic '{logic}' in condition: {condition}")
   
    bid_data = data.loc[data['Valid Bid']]
    data['Baseline Spend'] = data[bid_volume_col] * data[baseline_price_col]
    as_is_excl_list = []
    bid_ids = data[bid_id_col].unique()
   
    for bid_id in bid_ids:
        bid_rows = bid_data[bid_data[bid_id_col] == bid_id]
        all_rows = data[data[bid_id_col] == bid_id]
        if all_rows.empty:
            continue  # No data for this bid_id
        incumbent = all_rows[incumbent_col].iloc[0]
        facility = all_rows[facility_col].iloc[0]
        baseline_price = all_rows[baseline_price_col].iloc[0]
        bid_volume = all_rows[bid_volume_col].iloc[0]
        baseline_spend = bid_volume * baseline_price
 
        # Check if incumbent is excluded
        incumbent_excluded = False
        for condition in excluded_conditions:
            supplier, field, logic, value, exclude_all = condition
            supplier = supplier.strip().title()
            if supplier != incumbent:
                continue
            if exclude_all:
                incumbent_excluded = True
                break
            if field not in all_rows.columns:
                logger.warning(f"Field '{field}' does not exist in data. Skipping this condition for incumbent.")
                continue
            incumbent_field_value = all_rows[field].iloc[0]
            if pd.api.types.is_string_dtype(all_rows[field]):
                incumbent_field_value = str(incumbent_field_value).strip().title()
                value_comp = str(value).strip().title()
                if logic == "Equal to" and incumbent_field_value == value_comp:
                    incumbent_excluded = True
                    break
                elif logic == "Not equal to" and incumbent_field_value != value_comp:
                    incumbent_excluded = True
                    break
            else:
                if logic == "Equal to" and incumbent_field_value == value:
                    incumbent_excluded = True
                    break
                elif logic == "Not equal to" and incumbent_field_value != value:
                    incumbent_excluded = True
                    break
 
        if not incumbent_excluded:
            # Incumbent is not excluded
            incumbent_bid = bid_rows[bid_rows[supplier_name_col] == incumbent]
            if not incumbent_bid.empty:
                # Incumbent did bid
                row = incumbent_bid.iloc[0]
                supplier_capacity = row[supplier_capacity_col] if pd.notna(row[supplier_capacity_col]) else bid_volume
                awarded_volume = min(bid_volume, supplier_capacity)
                awarded_spend = awarded_volume * row[bid_price_col]
                baseline_savings = baseline_spend - awarded_spend
 
                row_dict = {
                    'Bid ID': bid_id,
                    'Bid ID Split': 'A',
                    'Facility': facility,
                    'Incumbent': incumbent,
                    'Baseline Price': baseline_price,
                    'Bid Volume': awarded_volume,
                    'Baseline Spend': awarded_volume * baseline_price,
                    'Awarded Supplier': incumbent,
                    'Awarded Supplier Price': row[bid_price_col],
                    'Awarded Volume': awarded_volume,
                    'Awarded Supplier Spend': awarded_spend,
                    'Awarded Supplier Capacity': supplier_capacity,
                    'Baseline Savings': baseline_savings  # Renamed from 'Savings'
                }
 
                if has_current_price:
                    current_price = data.loc[data[bid_id_col] == bid_id, current_price_col].iloc[0]
                    row_dict['Current Price'] = current_price
                    if row_dict['Awarded Supplier Price'] is not None and row_dict['Bid Volume'] is not None:
                        row_dict['Current Price Savings'] = (current_price - row_dict['Awarded Supplier Price']) * row_dict['Bid Volume']
                    else:
                        row_dict['Current Price Savings'] = None
 
                as_is_excl_list.append(row_dict)
                logger.debug(f"As-Is Excl analysis for Bid ID {bid_id}: Awarded to incumbent.")
 
                remaining_volume = bid_volume - awarded_volume
                if remaining_volume > 0:
                    # Remaining volume is unallocated
                    row_dict_unallocated = {
                        'Bid ID': bid_id,
                        'Bid ID Split': 'B',
                        'Facility': facility,
                        'Incumbent': incumbent,
                        'Baseline Price': baseline_price,
                        'Bid Volume': remaining_volume,
                        'Baseline Spend': remaining_volume * baseline_price,
                        'Awarded Supplier': 'Unallocated',
                        'Awarded Supplier Price': None,
                        'Awarded Volume': remaining_volume,
                        'Awarded Supplier Spend': None,
                        'Awarded Supplier Capacity': None,
                        'Baseline Savings': None  # Renamed from 'Savings'
                    }
 
                    if has_current_price:
                        row_dict_unallocated['Current Price'] = None
                        row_dict_unallocated['Current Price Savings'] = None
 
                    as_is_excl_list.append(row_dict_unallocated)
                    logger.debug(f"Remaining volume for Bid ID {bid_id} is unallocated after awarding to incumbent.")
            else:
                # Incumbent did not bid or bid is invalid
                row_dict = {
                    'Bid ID': bid_id,
                    'Bid ID Split': 'A',
                    'Facility': facility,
                    'Incumbent': incumbent,
                    'Baseline Price': baseline_price,
                    'Bid Volume': bid_volume,
                    'Baseline Spend': baseline_spend,
                    'Awarded Supplier': 'Unallocated',
                    'Awarded Supplier Price': None,
                    'Awarded Volume': bid_volume,
                    'Awarded Supplier Spend': None,
                    'Awarded Supplier Capacity': None,
                    'Baseline Savings': None,  # Renamed from 'Savings'
                }
 
                if has_current_price:
                    row_dict['Current Price'] = None
                    row_dict['Current Price Savings'] = None
 
                as_is_excl_list.append(row_dict)
                logger.debug(f"Incumbent did not bid or invalid bid for Bid ID {bid_id}. Entire volume is unallocated.")
        else:
            # Incumbent is excluded
            # Allocate to the lowest priced suppliers
            valid_bids = bid_rows[bid_rows[supplier_name_col] != incumbent]
            valid_bids = valid_bids.sort_values(by=bid_price_col)
            remaining_volume = bid_volume
            split_index = 'A'
 
            if valid_bids.empty:
                # No valid bids, mark as Unallocated
                row_dict = {
                    'Bid ID': bid_id,
                    'Bid ID Split': split_index,
                    'Facility': facility,
                    'Incumbent': incumbent,
                    'Baseline Price': baseline_price,
                    'Bid Volume': bid_volume,
                    'Baseline Spend': baseline_spend,
                    'Awarded Supplier': 'Unallocated',
                    'Awarded Supplier Price': None,
                    'Awarded Volume': bid_volume,
                    'Awarded Supplier Spend': None,
                    'Awarded Supplier Capacity': None,
                    'Baseline Savings': None  # Renamed from 'Savings'
                }
 
                if has_current_price:
                    current_price = data.loc[data[bid_id_col] == bid_id, current_price_col].iloc[0]
                    row_dict['Current Price'] = current_price
                    row_dict['Current Price Savings'] = None
 
                as_is_excl_list.append(row_dict)
                logger.debug(f"No valid bids for Bid ID {bid_id} after exclusions. Entire volume is unallocated.")
                continue
 
            for _, row in valid_bids.iterrows():
                supplier_capacity = row[supplier_capacity_col] if pd.notna(row[supplier_capacity_col]) else remaining_volume
                awarded_volume = min(remaining_volume, supplier_capacity)
                awarded_spend = awarded_volume * row[bid_price_col]
                baseline_spend_allocated = awarded_volume * baseline_price
                baseline_savings = baseline_spend_allocated - awarded_spend
 
                row_dict = {
                    'Bid ID': row[bid_id_col],
                    'Bid ID Split': split_index,
                    'Facility': facility,
                    'Incumbent': incumbent,
                    'Baseline Price': baseline_price,
                    'Bid Volume': awarded_volume,
                    'Baseline Spend': baseline_spend_allocated,
                    'Awarded Supplier': row[supplier_name_col],
                    'Awarded Supplier Price': row[bid_price_col],
                    'Awarded Volume': awarded_volume,
                    'Awarded Supplier Spend': awarded_spend,
                    'Awarded Supplier Capacity': supplier_capacity,
                    'Baseline Savings': baseline_savings  # Renamed from 'Savings'
                }
 
                if has_current_price:
                    current_price = data.loc[data[bid_id_col] == bid_id, current_price_col].iloc[0]
                    row_dict['Current Price'] = current_price
                    if row_dict['Awarded Supplier Price'] is not None and row_dict['Bid Volume'] is not None:
                        row_dict['Current Price Savings'] = (current_price - row_dict['Awarded Supplier Price']) * row_dict['Bid Volume']
                    else:
                        row_dict['Current Price Savings'] = None
 
                as_is_excl_list.append(row_dict)
                logger.debug(f"As-Is Excl analysis for Bid ID {bid_id}, Split {split_index}: Awarded Volume = {awarded_volume} to {row[supplier_name_col]}")
 
                remaining_volume -= awarded_volume
                if remaining_volume <= 0:
                    break
                split_index = chr(ord(split_index) + 1)
 
            if remaining_volume > 0:
                # Remaining volume is unallocated
                row_dict_unallocated = {
                    'Bid ID': bid_id,
                    'Bid ID Split': split_index,
                    'Facility': facility,
                    'Incumbent': incumbent,
                    'Baseline Price': baseline_price,
                    'Bid Volume': remaining_volume,
                    'Baseline Spend': remaining_volume * baseline_price,
                    'Awarded Supplier': 'Unallocated',
                    'Awarded Supplier Price': None,
                    'Awarded Volume': remaining_volume,
                    'Awarded Supplier Spend': None,
                    'Awarded Supplier Capacity': None,
                    'Baseline Savings': None  # Renamed from 'Savings'
                }
 
                if has_current_price:
                    row_dict_unallocated['Current Price'] = current_price
                    row_dict_unallocated['Current Price Savings'] = None
 
                as_is_excl_list.append(row_dict_unallocated)
                logger.debug(f"Remaining volume for Bid ID {bid_id} is unallocated after allocating to suppliers.")
 
    as_is_excl_df = pd.DataFrame(as_is_excl_list)
 
    # Define desired column order
    desired_columns = [
        'Bid ID', 'Bid ID Split', 'Facility', 'Incumbent',
        'Baseline Price'
    ]
 
    if has_current_price:
        desired_columns.append('Current Price')
 
    desired_columns.extend([
        'Bid Volume', 'Baseline Spend',
        'Awarded Supplier', 'Awarded Supplier Price',
        'Awarded Volume', 'Awarded Supplier Spend',
        'Awarded Supplier Capacity', 'Baseline Savings'
    ])
 
    if has_current_price:
        desired_columns.append('Current Price Savings')
 
    # Reorder columns
    as_is_excl_df = as_is_excl_df.reindex(columns=desired_columns)
 
    return as_is_excl_df
 
 
def customizable_analysis(data, column_mapping):
    """Perform 'Customizable Analysis' and prepare data for Excel output."""
    bid_price_col = column_mapping['Bid Price']
    bid_volume_col = column_mapping['Bid Volume']
    baseline_price_col = column_mapping['Baseline Price']
    supplier_capacity_col = column_mapping['Supplier Capacity']
    bid_id_col = column_mapping['Bid ID']
    facility_col = column_mapping['Facility']
    incumbent_col = column_mapping['Incumbent']
    supplier_name_col = column_mapping['Supplier Name']
 
    # Ensure necessary columns are numeric
    data[bid_volume_col] = pd.to_numeric(data[bid_volume_col], errors='coerce')
    data[supplier_capacity_col] = pd.to_numeric(data[supplier_capacity_col], errors='coerce')
    data[bid_price_col] = pd.to_numeric(data[bid_price_col], errors='coerce')
    data[baseline_price_col] = pd.to_numeric(data[baseline_price_col], errors='coerce')
 
    # Calculate Savings
    data['Savings'] = (data[baseline_price_col] - data[bid_price_col]) * data[bid_volume_col]
 
    # Create Supplier Name with Bid Price
    data['Supplier Name with Bid Price'] = data[supplier_name_col] + " ($" + data[bid_price_col].round(2).astype(str) + ")"
 
    # Calculate Baseline Spend
    data['Baseline Spend'] = data[bid_volume_col] * data[baseline_price_col]
 
    # Get unique Bid IDs
    bid_ids = data[bid_id_col].unique()
 
# Prepare the customizable analysis DataFrame
    customizable_list = []
    for bid_id in bid_ids:
        bid_row = data[data[bid_id_col] == bid_id].iloc[0]
        customizable_list.append({
            'Bid ID': bid_id,
            'Facility': bid_row[facility_col],
            'Incumbent': bid_row[incumbent_col],
            'Baseline Price': bid_row[baseline_price_col],
            'Bid Volume': bid_row[bid_volume_col],
            'Baseline Spend': bid_row['Baseline Spend'],
            'Awarded Supplier': '',  # To be selected via data validation in Excel
            'Supplier Name': '',     # New column added here
            'Awarded Supplier Price': None,  # Formula-based
            'Awarded Volume': None,  # Formula-based
            'Awarded Supplier Spend': None,  # Formula-based
            'Awarded Supplier Capacity': None,  # Formula-based
            'Savings': None  # Formula-based
        })
    customizable_df = pd.DataFrame(customizable_list)
    return customizable_df
 
 
def run_customizable_analysis_processing(
    st,
    writer,
    customizable_analysis,
    logger,
    pd,
    workbook,
    data,
    column_mapping,
    grouping_column_raw,
    grouping_column_mapped,
    grouping_column_already_mapped
):
    """
    Runs the entire 'Customizable Analysis' block that was previously inline in main().
    Modifies the ExcelWriter workbook sheets to add 'Customizable Template',
    'Customizable Reference', 'Scenario Converter', 'SupplierLists', 'Scenario Selector', and
    'Scenario Reports' as needed.
 
    Parameters:
        st: Streamlit instance (for st.session_state, st.error, etc.)
        writer: The pd.ExcelWriter instance.
        customizable_analysis: Reference to the 'customizable_analysis' function.
        logger: Logger instance for logging info, warnings, and errors.
        pd: Pandas module reference.
        workbook: The openpyxl workbook object from writer.book.
        data: The main merged_data DataFrame from st.session_state.
        column_mapping: The dictionary of mapped columns from st.session_state.
        grouping_column_raw: The raw grouping column name chosen by the user in Streamlit.
        grouping_column_mapped: The mapped grouping column name if the user’s column is already mapped, or the raw name otherwise.
        grouping_column_already_mapped: Boolean indicating if the selected grouping column is already recognized in column_mapping.
    """
 
    #########################################
    # 1) Prepare the 'Customizable Template' and 'Customizable Reference' sheets
    #########################################
 
    # Generate the customizable analysis DataFrame
    customizable_df = customizable_analysis(data, column_mapping)
    logger.info("Customizable analysis DataFrame created successfully.")
 
    # If grouping column was not already mapped, add it to customizable_df
    if not grouping_column_already_mapped:
        if grouping_column_mapped not in customizable_df.columns:
            bid_id_idx = customizable_df.columns.get_loc('Bid ID')
            customizable_df.insert(bid_id_idx + 1, grouping_column_mapped, '')
 
    # Write 'Customizable Template' sheet
    customizable_df.to_excel(writer, sheet_name='Customizable Template', index=False)
 
    # Prepare 'Customizable Reference' DataFrame
    customizable_reference_df = data.copy()  # the merged_data from st.session_state
 
    # If grouping column was not already mapped, add it to customizable_reference_df
    if not grouping_column_already_mapped:
        if grouping_column_mapped not in customizable_reference_df.columns:
            # Use 'grouping_column_mapped' instead of 'grouping_column_raw'
            bid_id_col = column_mapping['Bid ID']
            bid_id_to_grouping = data.set_index(bid_id_col)[grouping_column_mapped].to_dict()
            customizable_reference_df[grouping_column_mapped] = customizable_reference_df[bid_id_col].map(bid_id_to_grouping)
 
    # Write 'Customizable Reference' sheet
    customizable_reference_df.to_excel(writer, sheet_name='Customizable Reference', index=False)
    logger.info("Customizable Reference sheet written successfully.")
 
    #########################################
    # 2) Ensure 'Scenario Converter' Sheet
    #########################################
    workbook = writer.book  # Re-acquire the workbook from the writer
 
    # Retrieve the mapped columns for convenience
    bid_id_col = column_mapping['Bid ID']
    incumbent_col = column_mapping['Incumbent']
 
    # Build a dictionary mapping each Bid ID to its Incumbent
    # If you have multiple rows per Bid ID with different incumbents,
    #   you can decide whether to take the first, last, etc.
    # Here, we'll assume the first non-null Incumbent is correct.
    incumbent_map = (
        data
        .groupby(bid_id_col)[incumbent_col]
        .first()      # picks the first Incumbent per Bid ID
        .dropna()     # optionally drop missing
        .to_dict()
    )
 
    if 'Scenario Converter' not in workbook.sheetnames:
        logger.info("Creating new 'Scenario Converter' sheet...")
 
        # Gather unique Bid IDs
        bid_ids = data[bid_id_col].unique()
 
        # Build initial DataFrame
        scenario_converter_df = pd.DataFrame({'Bid ID': bid_ids})
 
        scenario_converter_df['As-is'] = ''  # We'll populate this with the incumbent name
        for i in range(2, 8):
            scenario_converter_df[f'Scenario {i}'] = ''  # Initialize empty columns
 
        # Now fill the 'As-is' column with the incumbent supplier name
        scenario_converter_df['As-is'] = scenario_converter_df['Bid ID'].map(incumbent_map)
 
        # Write to Excel
        scenario_converter_df.to_excel(writer, sheet_name='Scenario Converter', index=False)
        logger.info("'Scenario Converter' sheet created with incumbent names in 'Scenario 1' column.")
    else:
        logger.info("'Scenario Converter' sheet already exists.")
 
    #########################################
    # 3) Ensure grouping column in st.session_state.merged_data
    #########################################
    if not grouping_column_already_mapped:
        if grouping_column_mapped not in data.columns:
            bid_id_col = column_mapping['Bid ID']
            # Use the same 'bid_id_to_grouping' map
            bid_id_to_grouping = data.set_index(bid_id_col)[grouping_column_mapped].to_dict()
            data[grouping_column_mapped] = data[bid_id_col].map(bid_id_to_grouping)
            logger.info(f"Added grouping column '{grouping_column_mapped}' to 'merged_data'.")
 
    # Re-acquire sheets
    customizable_template_sheet = workbook['Customizable Template']
    customizable_reference_sheet = workbook['Customizable Reference']
    scenario_converter_sheet = workbook['Scenario Converter']
 
    # Get max row numbers
    max_row_template = customizable_template_sheet.max_row
    max_row_reference = customizable_reference_sheet.max_row
    max_row_converter = scenario_converter_sheet.max_row
 
    # Map column names to letters in 'Customizable Reference' sheet
    reference_col_letter = {
        cell.value: get_column_letter(cell.column) for cell in customizable_reference_sheet[1]
    }
    # Map column names to letters in 'Customizable Template' sheet
    template_col_letter = {
        cell.value: get_column_letter(cell.column) for cell in customizable_template_sheet[1]
    }
 
    # Grab column letters for the grouping column
    ref_grouping_col = reference_col_letter.get(grouping_column_mapped)
    temp_grouping_col = template_col_letter.get(grouping_column_mapped)
 
    # Update 'Supplier Name' formulas in 'Customizable Template' sheet
    awarded_supplier_col_letter = template_col_letter['Awarded Supplier']
    supplier_name_col_letter = template_col_letter['Supplier Name']
    for row in range(2, max_row_template + 1):
        awarded_supplier_cell = f"{awarded_supplier_col_letter}{row}"
        supplier_name_cell = f"{supplier_name_col_letter}{row}"
        formula_supplier_name = f'=IF({awarded_supplier_cell}<>"", LEFT({awarded_supplier_cell}, FIND("(", {awarded_supplier_cell}) - 2), "")'
        customizable_template_sheet[supplier_name_cell].value = formula_supplier_name
 
    # Create or reuse 'SupplierLists' sheet
    if 'SupplierLists' not in workbook.sheetnames:
        supplier_list_sheet = workbook.create_sheet("SupplierLists")
    else:
        supplier_list_sheet = workbook['SupplierLists']
 
    bid_id_col_reference = column_mapping['Bid ID']
    supplier_name_with_bid_price_col_reference = column_mapping.get('Supplier Name with Bid Price', 'Supplier Name with Bid Price')
    bid_id_supplier_list_ranges = {}
    current_row = 1
 
    # Build out supplier lists per Bid ID in the hidden 'SupplierLists' sheet
    bid_ids_unique = data[bid_id_col_reference].unique()
    for b_id in bid_ids_unique:
        subset = data[data[bid_id_col_reference] == b_id]
        # Filter valid bids only
        valid_bids = subset[
            (subset[column_mapping['Bid Price']].notna()) &
            (subset[column_mapping['Bid Price']] != 0)
        ]
        if not valid_bids.empty:
            valid_bids_sorted = valid_bids.sort_values(by=column_mapping['Bid Price'])
            sup_list = valid_bids_sorted[supplier_name_with_bid_price_col_reference].dropna().tolist()
            start_row = current_row
            for sup in sup_list:
                supplier_list_sheet.cell(row=current_row, column=1, value=sup)
                current_row += 1
            end_row = current_row - 1
            bid_id_supplier_list_ranges[b_id] = (start_row, end_row)
            current_row += 1  # Empty line for separation
        else:
            bid_id_supplier_list_ranges[b_id] = None
 
    supplier_list_sheet.sheet_state = 'hidden'
 
    # Set data validation and formulas in 'Customizable Template'
    ref_sheet_name = 'Customizable Reference'
    temp_sheet_name = 'Customizable Template'
    ref_bid_id_col = reference_col_letter[column_mapping['Bid ID']]
    ref_supplier_name_col = reference_col_letter[column_mapping['Supplier Name']]
    ref_bid_price_col = reference_col_letter[column_mapping['Bid Price']]
    ref_supplier_capacity_col = reference_col_letter[column_mapping['Supplier Capacity']]
    temp_bid_id_col = template_col_letter['Bid ID']
    temp_supplier_name_col = template_col_letter['Supplier Name']
    temp_bid_volume_col = template_col_letter['Bid Volume']
    temp_baseline_price_col = template_col_letter['Baseline Price']
 
    # Loop over each row in Customizable Template to set formulas
    for row in range(2, max_row_template + 1):
        bid_id_cell = f"{temp_bid_id_col}{row}"
        awarded_supplier_cell = f"{awarded_supplier_col_letter}{row}"
        supplier_name_cell = f"{temp_supplier_name_col}{row}"
        bid_volume_cell = f"{temp_bid_volume_col}{row}"
        baseline_price_cell = f"{temp_baseline_price_col}{row}"
        grouping_cell = f"{temp_grouping_col}{row}" if temp_grouping_col else None
 
        b_id_value = customizable_template_sheet[bid_id_cell].value
        if b_id_value in bid_id_supplier_list_ranges and bid_id_supplier_list_ranges[b_id_value]:
            start_row, end_row = bid_id_supplier_list_ranges[b_id_value]
            supplier_list_range = f"'SupplierLists'!$A${start_row}:$A${end_row}"
 
            # Data validation for Awarded Supplier
            dv = DataValidation(type="list", formula1=f"{supplier_list_range}", allow_blank=True)
            customizable_template_sheet.add_data_validation(dv)
            dv.add(customizable_template_sheet[awarded_supplier_cell])
 
            # Construct SUMIFS-based formulas
            awarded_supplier_price_cell = f"{template_col_letter['Awarded Supplier Price']}{row}"
            formula_price = (
                f"=IFERROR(SUMIFS('{ref_sheet_name}'!{ref_bid_price_col}:{ref_bid_price_col}, "
                f"'{ref_sheet_name}'!{ref_bid_id_col}:{ref_bid_id_col}, {bid_id_cell}, "
                f"'{ref_sheet_name}'!{ref_supplier_name_col}:{ref_supplier_name_col}, {supplier_name_cell}), \"\")"
            )
            customizable_template_sheet[awarded_supplier_price_cell].value = formula_price
 
            awarded_supplier_capacity_cell = f"{template_col_letter['Awarded Supplier Capacity']}{row}"
            formula_supplier_capacity = (
                f"=IFERROR(SUMIFS('{ref_sheet_name}'!{ref_supplier_capacity_col}:{ref_supplier_capacity_col}, "
                f"'{ref_sheet_name}'!{ref_bid_id_col}:{ref_bid_id_col}, {bid_id_cell}, "
                f"'{ref_sheet_name}'!{ref_supplier_name_col}:{ref_supplier_name_col}, {supplier_name_cell}), \"\")"
            )
            customizable_template_sheet[awarded_supplier_capacity_cell].value = formula_supplier_capacity
 
            awarded_volume_cell = f"{template_col_letter['Awarded Volume']}{row}"
            formula_awarded_volume = f"=IFERROR(MIN({bid_volume_cell}, {awarded_supplier_capacity_cell}), \"\")"
            customizable_template_sheet[awarded_volume_cell].value = formula_awarded_volume
 
            awarded_supplier_spend_cell = f"{template_col_letter['Awarded Supplier Spend']}{row}"
            formula_spend = f"=IF({awarded_supplier_price_cell}<>\"\", {awarded_supplier_price_cell}*{awarded_volume_cell}, \"\")"
            customizable_template_sheet[awarded_supplier_spend_cell].value = formula_spend
 
            savings_cell = f"{template_col_letter['Savings']}{row}"
            formula_savings = f"=IF({awarded_supplier_price_cell}<>\"\", ({baseline_price_cell}-{awarded_supplier_price_cell})*{awarded_volume_cell}, \"\")"
            customizable_template_sheet[savings_cell].value = formula_savings
 
            baseline_spend_cell = f"{template_col_letter['Baseline Spend']}{row}"
            formula_baseline_spend = f"={baseline_price_cell}*{bid_volume_cell}"
            customizable_template_sheet[baseline_spend_cell].value = formula_baseline_spend
 
            # If grouping column exists, populate it
            if grouping_cell and ref_grouping_col:
                col_index = column_index_from_string(ref_grouping_col) - column_index_from_string(ref_bid_id_col) + 1
                formula_grouping = (
                    f"=IFERROR(VLOOKUP({bid_id_cell}, '{ref_sheet_name}'!{ref_bid_id_col}:{ref_grouping_col}, "
                    f"{col_index}, FALSE), \"\")"
                )
                customizable_template_sheet[grouping_cell].value = formula_grouping
 
    #########################################
    # 4) Apply Formatting
    #########################################
    # Format 'Customizable Reference' sheet
    currency_columns_reference = [
        'Baseline Spend', 'Savings', column_mapping['Bid Price'], column_mapping['Baseline Price']
    ]
    number_columns_reference = [
        column_mapping['Bid Volume'], column_mapping['Supplier Capacity']
    ]
    for col_name in currency_columns_reference:
        col_letter = reference_col_letter.get(col_name)
        if col_letter:
            for row_num in range(2, max_row_reference + 1):
                cell = customizable_reference_sheet[f"{col_letter}{row_num}"]
                cell.number_format = '$#,##0.00'
    for col_name in number_columns_reference:
        col_letter = reference_col_letter.get(col_name)
        if col_letter:
            for row_num in range(2, max_row_reference + 1):
                cell = customizable_reference_sheet[f"{col_letter}{row_num}"]
                cell.number_format = '#,##0'
 
    # Format 'Customizable Template' sheet
    currency_columns_template = [
        'Baseline Spend', 'Baseline Price', 'Awarded Supplier Price', 'Awarded Supplier Spend', 'Savings'
    ]
    number_columns_template = ['Bid Volume', 'Awarded Volume', 'Awarded Supplier Capacity']
    for col_name in currency_columns_template:
        col_letter = template_col_letter.get(col_name)
        if col_letter:
            for row_num in range(2, max_row_template + 1):
                cell = customizable_template_sheet[f"{col_letter}{row_num}"]
                cell.number_format = '$#,##0.00'
    for col_name in number_columns_template:
        col_letter = template_col_letter.get(col_name)
        if col_letter:
            for row_num in range(2, max_row_template + 1):
                cell = customizable_template_sheet[f"{col_letter}{row_num}"]
                cell.number_format = '#,##0'
 
 
    #########################################
    # 5) Create the Scenario Selector sheet
    #########################################
    scenario_selector_df = customizable_df.copy()
    # Remove 'Supplier Name' column from 'Scenario Selector' if it exists
    if 'Supplier Name' in scenario_selector_df.columns:
        scenario_selector_df.drop(columns=['Supplier Name'], inplace=True)
 
    # Ensure grouping column is present
    if not grouping_column_already_mapped:
        if grouping_column_raw not in scenario_selector_df.columns:
            bid_id_idx = scenario_selector_df.columns.get_loc('Bid ID')
            scenario_selector_df.insert(bid_id_idx + 1, grouping_column_raw, '')
 
    scenario_selector_df.to_excel(writer, sheet_name='Scenario Selector', index=False)
    scenario_selector_sheet = workbook['Scenario Selector']
 
    # Map column names to letters in 'Scenario Selector' sheet
    scenario_selector_col_letter = {
        cell.value: get_column_letter(cell.column) for cell in scenario_selector_sheet[1]
    }
    max_row_selector = scenario_selector_sheet.max_row
 
    scenario_converter_sheet = workbook['Scenario Converter']
    max_row_converter = scenario_converter_sheet.max_row
 
    scenario_converter_data_range = f"'Scenario Converter'!$B$2:$H${max_row_converter}"
    scenario_converter_header_range = "'Scenario Converter'!$B$1:$H$1"
    scenario_bid_ids_range = f"'Scenario Converter'!$A$2:$A${max_row_converter}"
 
    bid_id_col_selector = scenario_selector_col_letter['Bid ID']
    awarded_supplier_col = scenario_selector_col_letter['Awarded Supplier']
    # Check if the grouping column is in scenario_selector_col_letter
    selector_grouping_col = scenario_selector_col_letter.get(grouping_column_raw, None)
 
 
    for row in range(2, max_row_selector + 1):
        bid_id_cell = f"{bid_id_col_selector}{row}"
        awarded_supplier_cell = f"{awarded_supplier_col}{row}"
        bid_volume_cell = f"{scenario_selector_col_letter['Bid Volume']}{row}"
        baseline_price_cell = f"{scenario_selector_col_letter['Baseline Price']}{row}"
        grouping_cell = f"{selector_grouping_col}{row}" if selector_grouping_col else None
 
        # Awarded Supplier
        formula_awarded_supplier = (
            f"=IFERROR(INDEX({scenario_converter_data_range}, MATCH({bid_id_cell}, {scenario_bid_ids_range}, 0), "
            f"MATCH('Scenario Reports'!$A$1, {scenario_converter_header_range}, 0)), \"\")"
        )
        scenario_selector_sheet[awarded_supplier_cell].value = formula_awarded_supplier
 
        # Awarded Supplier Price
        awarded_supplier_price_cell = f"{scenario_selector_col_letter['Awarded Supplier Price']}{row}"
        awarded_supplier_capacity_cell = f"{scenario_selector_col_letter['Awarded Supplier Capacity']}{row}"
 
        formula_price = (
            f"=IFERROR(SUMIFS('{ref_sheet_name}'!{ref_bid_price_col}:{ref_bid_price_col}, "
            f"'{ref_sheet_name}'!{ref_bid_id_col}:{ref_bid_id_col}, {bid_id_cell}, "
            f"'{ref_sheet_name}'!{ref_supplier_name_col}:{ref_supplier_name_col}, {awarded_supplier_cell}), \"\")"
        )
        scenario_selector_sheet[awarded_supplier_price_cell].value = formula_price
 
        # Awarded Supplier Capacity
        formula_supplier_capacity = (
            f"=IFERROR(SUMIFS('{ref_sheet_name}'!{ref_supplier_capacity_col}:{ref_supplier_capacity_col}, "
            f"'{ref_sheet_name}'!{ref_bid_id_col}:{ref_bid_id_col}, {bid_id_cell}, "
            f"'{ref_sheet_name}'!{ref_supplier_name_col}:{ref_supplier_name_col}, {awarded_supplier_cell}), \"\")"
        )
        scenario_selector_sheet[awarded_supplier_capacity_cell].value = formula_supplier_capacity
 
        # Awarded Volume
        awarded_volume_cell = f"{scenario_selector_col_letter['Awarded Volume']}{row}"
        formula_awarded_volume = f"=IF({bid_volume_cell}=\"\", \"\", MIN({bid_volume_cell}, {awarded_supplier_capacity_cell}))"
        scenario_selector_sheet[awarded_volume_cell].value = formula_awarded_volume
 
        # Awarded Supplier Spend
        awarded_supplier_spend_cell = f"{scenario_selector_col_letter['Awarded Supplier Spend']}{row}"
        formula_spend = f"=IF({awarded_supplier_price_cell}<>\"\", {awarded_supplier_price_cell}*{awarded_volume_cell}, \"\")"
        scenario_selector_sheet[awarded_supplier_spend_cell].value = formula_spend
 
        # Savings
        savings_cell = f"{scenario_selector_col_letter['Savings']}{row}"
        formula_savings = f"=IF({awarded_supplier_price_cell}<>\"\", ({baseline_price_cell}-{awarded_supplier_price_cell})*{awarded_volume_cell}, \"\")"
        scenario_selector_sheet[savings_cell].value = formula_savings
 
        # Baseline Spend
        baseline_spend_cell = f"{scenario_selector_col_letter['Baseline Spend']}{row}"
        formula_baseline_spend = f"={baseline_price_cell}*{bid_volume_cell}"
        scenario_selector_sheet[baseline_spend_cell].value = formula_baseline_spend
 
        # If grouping column exists, populate it
        if grouping_cell and ref_grouping_col:
            col_index = column_index_from_string(ref_grouping_col) - column_index_from_string(ref_bid_id_col) + 1
            formula_grouping = (
                f"=IFERROR(VLOOKUP({bid_id_cell}, '{ref_sheet_name}'!{ref_bid_id_col}:{ref_grouping_col}, "
                f"{col_index}, FALSE), \"\")"
            )
            scenario_selector_sheet[grouping_cell].value = formula_grouping
 
    # Apply formatting to 'Scenario Selector' sheet
    currency_columns_selector = [
        'Baseline Spend', 'Baseline Price', 'Awarded Supplier Price', 'Awarded Supplier Spend', 'Savings'
    ]
    number_columns_selector = ['Bid Volume', 'Awarded Volume', 'Awarded Supplier Capacity']
    for col_name in currency_columns_selector:
        col_letter = scenario_selector_col_letter.get(col_name)
        if col_letter:
            for row_num in range(2, max_row_selector + 1):
                cell = scenario_selector_sheet[f"{col_letter}{row_num}"]
                cell.number_format = '$#,##0.00'
    for col_name in number_columns_selector:
        col_letter = scenario_selector_col_letter.get(col_name)
        if col_letter:
            for row_num in range(2, max_row_selector + 1):
                cell = scenario_selector_sheet[f"{col_letter}{row_num}"]
                cell.number_format = '#,##0'
 
    #########################################
    # 6) Create the Scenario Reports sheet
    #########################################
    if 'Scenario Reports' not in workbook.sheetnames:
        scenario_reports_sheet = workbook.create_sheet('Scenario Reports')
        logger.info("'Scenario Reports' sheet created.")
    else:
        scenario_reports_sheet = workbook['Scenario Reports']
        logger.info("'Scenario Reports' sheet already exists.")
 
    starting_row = 4
    unique_suppliers = data[column_mapping['Supplier Name']].unique()
    num_suppliers = len(unique_suppliers)
    if grouping_column_raw not in customizable_reference_df.columns:
        st.error(f"Grouping column '{grouping_column_raw}' not found in merged data.")
        logger.error(f"Grouping column '{grouping_column_raw}' not found in merged data.")
        return
 
    # Create scenario_reports_df from 'Customizable Reference'
    scenario_reports_df = customizable_reference_df[
        [column_mapping['Bid ID'], grouping_column_raw, column_mapping['Supplier Name']]
    ].copy()
    unique_groups = scenario_reports_df[grouping_column_raw].dropna().unique()
 
    # If 'Scenario Selector' doesn't exist, create minimal version
    if 'Scenario Selector' not in workbook.sheetnames:
        scenario_selector_sheet = workbook.create_sheet('Scenario Selector')
        headers = [
            'Bid ID', grouping_column_raw, column_mapping['Supplier Name'], 'Awarded Supplier Price',
            'Awarded Supplier Capacity', 'Bid Volume', 'Baseline Price', 'Savings'
        ]
        for col_num, hdr in enumerate(headers, start=1):
            cell = scenario_selector_sheet.cell(row=1, column=col_num, value=hdr)
            cell.font = Font(bold=True)
        logger.info("'Scenario Selector' sheet created with minimal headers.")
    else:
        scenario_selector_sheet = workbook['Scenario Selector']
        scenario_selector_headers = [cell.value for cell in scenario_selector_sheet[1]]
        logger.info("'Scenario Selector' sheet already exists.")
 
    # Detect grouping column in 'Scenario Selector'
    if grouping_column_mapped in scenario_selector_headers:
        column_offset = 1
        logger.info(f"Grouping column '{grouping_column_mapped}' found in 'Scenario Selector'. Shifting columns by 1.")
    else:
        column_offset = 0
        logger.info("No grouping column found in 'Scenario Selector'. No shift applied to references.")
 
    # Helper function to shift columns ...
    def shift_column(col_letter, scenario_selector_sheet, header_row=1):
        header_cells = scenario_selector_sheet[header_row]
        header_col_count = sum(1 for cell in header_cells if cell.value is not None)
        if header_col_count == 13:
            offset = 1
            logger.info(f"13 column headers in 'Scenario Selector'. Shift columns by 1.")
        elif header_col_count == 12:
            offset = 0
            logger.info(f"12 column headers in 'Scenario Selector'. No shift.")
        else:
            offset = 0
            logger.warning(f"Unexpected column count {header_col_count}. No shift applied.")
        try:
            col_idx = column_index_from_string(col_letter)
        except ValueError:
            logger.error(f"Invalid column letter: {col_letter}")
            raise ValueError(f"Invalid column letter: {col_letter}")
 
        new_idx = col_idx + offset
        if not 1 <= new_idx <= 16384:
            logger.error(f"Shifted column index {new_idx} is out of range.")
            raise ValueError(f"Shifted column index {new_idx} out of Excel range.")
        new_col_letter = get_column_letter(new_idx)
        logger.debug(f"Column '{col_letter}' shifted to '{new_col_letter}'.")
        return new_col_letter
 
 
 
    def limit_rows_in_column(col_ref: str, first_row: int = 2, last_row: int = 999) -> str:
        """
        Takes a column reference like '$G:$G' and returns '$G$2:$G$999'.
        If 'col_ref' doesn't match the pattern of '$LETTER:$LETTER', we return it unchanged.
        """
        # This regex looks for something like '$G:$G' or 'G:G' capturing just the letter(s).
        match = re.match(r'^\$?([A-Z]+)\$?:\$?([A-Z]+)$', col_ref)
        if not match:
            # Not in the pattern we expect, return as-is
            return col_ref
 
        # Typically group(1) == group(2) if it's something like '$G:$G'
        col_letters_1, col_letters_2 = match.groups()
        if col_letters_1 != col_letters_2:
            # For safety, just handle the left column
            pass
 
        # Use the first match's letters
        col_letters = col_letters_1
        # Build the final string, e.g. '$G$2:$G$999'
        return f"${col_letters}${first_row}:${col_letters}${last_row}"
 
 
    # For each grouping
    for group in unique_groups:
        scenario_reports_sheet[f"A{starting_row}"] = group
        scenario_reports_sheet[f"A{starting_row}"].font = Font(bold=True)
        headers = [
            'Supplier Name', 'Awarded Volume', '% of Business',
            'Baseline Avg', 'Avg Bid Price', '%Δ b/w Baseline and Avg Bid', 'RFP Savings'
        ]
        for col_num, hdr in enumerate(headers, start=1):
            c = scenario_reports_sheet.cell(row=starting_row + 1, column=col_num, value=hdr)
            c.font = Font(bold=True)
 
        # Insert Formula for Supplier Name in the row below
        supplier_name_cell = f"A{starting_row + 2}"
        group_label_cell = f"'Scenario Reports'!$A${starting_row}"
 
 
        try:
            # Original references
            original_supplier_name_col = 'G'
            original_group_col = 'B'
            adjusted_supplier_name_col = shift_column(original_supplier_name_col, scenario_selector_sheet)
            adjusted_group_col = shift_column(original_group_col, scenario_selector_sheet)
 
                        # Suppose shift_column() returned "$G:$G"
            adjusted_supplier_name_col = shift_column(original_supplier_name_col, scenario_selector_sheet)
 
            # Then we replace entire-column with row-limited references
            adj_supplier_range = limit_rows_in_column(adjusted_supplier_name_col, first_row=2, last_row=999)
            adj_group_range = limit_rows_in_column(adjusted_group_col,first_row=2, last_row=999)
 
            formula_supplier_name = (
                f"=IFERROR(UNIQUE("
                f"FILTER('Scenario Selector'!{adj_supplier_range},"
                f"('Scenario Selector'!{adj_supplier_range}<>\"\") * "
                f"('Scenario Selector'!{adj_group_range}={group_label_cell})"
                f")"
                f"), \"\")"
            )
            scenario_reports_sheet[supplier_name_cell].value = formula_supplier_name
            logger.info(f"Assigned formula to {supplier_name_cell}: {formula_supplier_name}")
        except Exception as e:
            logger.error(f"Failed to assign formula to {supplier_name_cell}: {e}")
            st.error(f"An error occurred while assigning formula to {supplier_name_cell}: {e}")
 
        awarded_volume_cell = f"B{starting_row + 2}"
        percent_business_cell = f"C{starting_row + 2}"
        avg_baseline_price_cell = f"D{starting_row + 2}"
        avg_bid_price_cell = f"E{starting_row + 2}"
        percent_delta_cell = f"F{starting_row + 2}"
        rfp_savings_cell = f"G{starting_row + 2}"
 
        try:
            # Original references for others
            original_awarded_volume_col = 'I'
            original_bid_volume_col = 'I'
            original_bid_price_col = 'D'
            original_avg_bid_price_col = 'H'
            original_savings_col = 'L'
 
            adjusted_awarded_volume_col = shift_column(original_awarded_volume_col, scenario_selector_sheet)
            adjusted_bid_volume_col = shift_column(original_bid_volume_col, scenario_selector_sheet)
            adjusted_bid_price_col = shift_column(original_bid_price_col, scenario_selector_sheet)
            adjusted_avg_bid_price_col = shift_column(original_avg_bid_price_col, scenario_selector_sheet)
            adjusted_savings_col = shift_column(original_savings_col, scenario_selector_sheet)
 
            # Awarded Volume
            formula_awarded_volume = (
                f"=IF({supplier_name_cell}=\"\", \"\", SUMIFS('Scenario Selector'!${adjusted_awarded_volume_col}:${adjusted_awarded_volume_col}, "
                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}, "
                f"'Scenario Selector'!${adjusted_supplier_name_col}:${adjusted_supplier_name_col}, {supplier_name_cell}))"
            )
            scenario_reports_sheet[awarded_volume_cell].value = formula_awarded_volume
 
            # % of Business
            formula_percent_business = (
                f"=IF({awarded_volume_cell}=0, \"\", {awarded_volume_cell}/SUMIFS('Scenario Selector'!${adjusted_awarded_volume_col}:${adjusted_awarded_volume_col}, "
                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}))"
            )
            scenario_reports_sheet[percent_business_cell].value = formula_percent_business
 
            # Avg Baseline Price
            formula_avg_baseline_price = (
                f"=IF({supplier_name_cell}=\"\", \"\", AVERAGEIFS('Scenario Selector'!${adjusted_bid_price_col}:{adjusted_bid_price_col}, "
                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}, "
                f"'Scenario Selector'!${adjusted_supplier_name_col}:{adjusted_supplier_name_col}, {supplier_name_cell}))"
            )
            scenario_reports_sheet[avg_baseline_price_cell].value = formula_avg_baseline_price
 
            # Avg Bid Price
            formula_avg_bid_price = (
                f"=IF({supplier_name_cell}=\"\", \"\", AVERAGEIFS('Scenario Selector'!${adjusted_avg_bid_price_col}:{adjusted_avg_bid_price_col}, "
                f"'Scenario Selector'!${original_group_col}:{original_group_col}, {group_label_cell}, "
                f"'Scenario Selector'!${adjusted_supplier_name_col}:{adjusted_supplier_name_col}, {supplier_name_cell}))"
            )
            scenario_reports_sheet[avg_bid_price_cell].value = formula_avg_bid_price
 
            # % Delta
            formula_percent_delta = (
                f"=IF(AND({avg_baseline_price_cell}>0, {avg_bid_price_cell}>0), "
                f"({avg_baseline_price_cell}-{avg_bid_price_cell})/{avg_baseline_price_cell}, \"\")"
            )
            scenario_reports_sheet[percent_delta_cell].value = formula_percent_delta
 
            # RFP Savings
            formula_rfp_savings = (
                f"=IFERROR(IF({supplier_name_cell}=\"\", \"\", SUMIFS('Scenario Selector'!${adjusted_savings_col}:${adjusted_savings_col}, "
                f"'Scenario Selector'!${original_group_col}:${original_group_col}, {group_label_cell}, "
                f"'Scenario Selector'!${adjusted_supplier_name_col}:${adjusted_supplier_name_col}, {supplier_name_cell})),\"\")"
            )
            scenario_reports_sheet[rfp_savings_cell].value = formula_rfp_savings
 
            logger.info(f"Successfully assigned formulas to columns B-G in row {starting_row + 2}")
        except Exception as e:
            logger.error(f"Failed to assign formulas to columns B-G in row {starting_row + 2}: {e}")
            st.error(f"An error occurred while assigning formulas to columns B-G in row {starting_row + 2}: {e}")
 
        # Advance the starting row
        max_suppliers_per_group = 10
        starting_row += 2 + max_suppliers_per_group + 3  # offset
 
        # Add drop-down to Scenario Reports (cell A1)
        try:
            dv_scenario = DataValidation(type="list", formula1="'Scenario Converter'!$B$1:$H$1", allow_blank=True)
            scenario_reports_sheet.add_data_validation(dv_scenario)
            dv_scenario.add(scenario_reports_sheet['A1'])
            # Set the default value in A1 to whatever is in 'Scenario Converter'!B1:
            scenario_reports_sheet['A1'].value = "='Scenario Converter'!B1"
            logger.info("Scenario Reports sheet created with drop-down in cell A1.")
        except Exception as e:
            st.error(f"Error adding drop-down to Scenario Reports sheet: {e}")
            logger.error(f"Error adding drop-down to Scenario Reports sheet: {e}")
 
        logger.info(f"Advanced starting row to {starting_row}")
 
    logger.info("Customizable Analysis processing completed.")
 
 
# Bid Coverage Reporting
 
# Function for Competitiveness Report
def competitiveness_report(data, column_mapping, group_by_field):
    """Generate Competitiveness Report with corrected calculations."""
    logger.info(f"Generating Competitiveness Report grouped by {group_by_field}.")
 
    # Extract column names from column_mapping
    bid_price_col = column_mapping['Bid Price']
    bid_id_col = column_mapping['Bid ID']
    incumbent_col = column_mapping['Incumbent']
    supplier_name_col = 'Awarded Supplier'  # Use 'Awarded Supplier' directly
 
    # Prepare data
    suppliers = data[supplier_name_col].unique()
    total_suppliers = len(suppliers)
 
    # Treat bids with Bid Price NaN or 0 as 'No Bid'
    data['Valid Bid'] = data[bid_price_col].notna() & (data[bid_price_col] != 0)
 
    grouped = data.groupby(group_by_field)
    report_rows = []
 
    for group, group_data in grouped:
        unique_bid_ids = group_data[bid_id_col].unique()
        total_bid_ids = len(unique_bid_ids)
        possible_bids = total_suppliers * total_bid_ids
 
        bids_received = group_data[group_data['Valid Bid']].shape[0]
 
        bid_ids_with_no_bids = total_bid_ids - group_data[group_data['Valid Bid']][bid_id_col].nunique()
 
        bid_ids_multiple_bids = group_data[group_data['Valid Bid']].groupby(bid_id_col)[supplier_name_col].nunique()
        percent_multiple_bids = (bid_ids_multiple_bids > 1).sum() / total_bid_ids * 100 if total_bid_ids > 0 else 0
 
        # Incumbent not bidding
        bid_ids_incumbent_no_bid = []
        for bid_id in unique_bid_ids:
            bid_rows = group_data[group_data[bid_id_col] == bid_id]
            incumbent = bid_rows[incumbent_col].iloc[0]
            incumbent_bid = bid_rows[(bid_rows[supplier_name_col] == incumbent) & (bid_rows['Valid Bid'])]
            if incumbent_bid.empty:
                bid_ids_incumbent_no_bid.append(bid_id)
        num_incumbent_no_bid = len(bid_ids_incumbent_no_bid)
        bid_ids_incumbent_no_bid_list = ', '.join(map(str, bid_ids_incumbent_no_bid))
 
        report_rows.append({
            'Group': group,
            '# of Possible Bids': possible_bids,
            '# of Bids Received': bids_received,
            'Bid IDs with No Bids': bid_ids_with_no_bids,
            '% of Bid IDs with Multiple Bids': f"{percent_multiple_bids:.0f}%",
            '# of Bid IDs Where Incumbent Did Not Bid': num_incumbent_no_bid,
            'List of Bid IDs Where Incumbent Did Not Bid': bid_ids_incumbent_no_bid_list
        })
 
    report_df = pd.DataFrame(report_rows)
    return report_df
 
# Function for Supplier Coverage Report
def supplier_coverage_report(data, column_mapping, group_by_field):
    """Generate Supplier Coverage Report with All Bids and grouped tables."""
    logger.info(f"Generating Supplier Coverage Report grouped by {group_by_field}.")
 
    # Extract column names from column_mapping
    bid_price_col = column_mapping['Bid Price']
    bid_id_col = column_mapping['Bid ID']
    supplier_name_col = 'Awarded Supplier'  # Use 'Awarded Supplier' directly
 
    # Treat bids with Bid Price NaN or 0 as 'No Bid'
    data['Valid Bid'] = data[bid_price_col].notna() & (data[bid_price_col] != 0)
 
    total_bid_ids = data[bid_id_col].nunique()
    suppliers = data[supplier_name_col].unique()
    all_bids_rows = []
    for supplier in suppliers:
        bids_provided = data[(data[supplier_name_col] == supplier) & (data['Valid Bid'])][bid_id_col].nunique()
        coverage = (bids_provided / total_bid_ids) * 100 if total_bid_ids > 0 else 0
        all_bids_rows.append({
            'Supplier': supplier,
            '# of Bid IDs': total_bid_ids,
            '# of Bids Provided': bids_provided,
            '% Coverage': f"{coverage:.0f}%"
        })
    all_bids_df = pd.DataFrame(all_bids_rows)
 
    # Grouped Tables
    grouped_tables = {}
    groups = data[group_by_field].unique()
    for group in groups:
        group_data = data[data[group_by_field] == group]
        group_total_bid_ids = group_data[bid_id_col].nunique()
        group_rows = []
        for supplier in suppliers:
            bids_provided = group_data[(group_data[supplier_name_col] == supplier) & (group_data['Valid Bid'])][bid_id_col].nunique()
            coverage = (bids_provided / group_total_bid_ids) * 100 if group_total_bid_ids > 0 else 0
            group_rows.append({
                'Supplier': supplier,
                '# of Bid IDs': group_total_bid_ids,
                '# of Bids Provided': bids_provided,
                '% Coverage': f"{coverage:.0f}%"
            })
        group_df = pd.DataFrame(group_rows)
        grouped_tables[f"Supplier Coverage - {group}"] = group_df
 
    return {'Supplier Coverage - All Bids': all_bids_df, **grouped_tables}
 
# Function for Facility Coverage Report
def facility_coverage_report(data, column_mapping, group_by_field):
    """Generate Facility Coverage Report grouped by the specified field."""
    logger.info(f"Generating Facility Coverage Report grouped by {group_by_field}.")
 
    facility_col = column_mapping['Facility']
    supplier_name_col = 'Awarded Supplier'  # Use 'Awarded Supplier' directly
    bid_price_col = column_mapping['Bid Price']
    bid_id_col = column_mapping['Bid ID']
 
    facilities = data[facility_col].unique()
    suppliers = data[supplier_name_col].unique()
    report = pd.DataFrame({'Supplier': suppliers})
    report.set_index('Supplier', inplace=True)
 
    # Treat bids with Bid Price NaN or 0 as 'No Bid'
    data['Valid Bid'] = data[bid_price_col].notna() & (data[bid_price_col] != 0)
 
    for facility in facilities:
        facility_bids = data.loc[(data[facility_col] == facility) & (data['Valid Bid'])]
        total_bid_ids = data[data[facility_col] == facility][bid_id_col].nunique()
        coverage = facility_bids.groupby(supplier_name_col)[bid_id_col].nunique() / total_bid_ids
        coverage = coverage.reindex(suppliers).fillna(0) * 100  # Ensure alignment with suppliers
        report[facility] = coverage
    report.reset_index(inplace=True)
    return report
 
# Function to handle Bid Coverage Report
def bid_coverage_report(data, column_mapping, variations, group_by_field):
    """Generate Bid Coverage Reports based on selected variations and grouping."""
    logger.info(f"Running Bid Coverage Report with variations: {variations} and grouping by {group_by_field}.")
    reports = {}
    if "Competitiveness Report" in variations:
        competitiveness = competitiveness_report(data, column_mapping, group_by_field)
        reports['Competitiveness Report'] = competitiveness
        logger.info("Competitiveness Report generated.")
    if "Supplier Coverage" in variations:
        supplier_coverage = supplier_coverage_report(data, column_mapping, group_by_field)
        reports.update(supplier_coverage)  # Include all tables
        logger.info("Supplier Coverage Report generated.")
    if "Facility Coverage" in variations:
        facility_coverage = facility_coverage_report(data, column_mapping, group_by_field)
        reports['Facility Coverage'] = facility_coverage
        logger.info("Facility Coverage Report generated.")
    return reports

